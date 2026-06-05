from __future__ import annotations

import datetime as dt
import base64
import csv
import gzip
import hashlib
import json
import importlib.util
import math
import os
import re
import shutil
import sys
import tempfile
import threading
import time
import tkinter as tk
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from dataclasses import dataclass
from pathlib import Path
from queue import Empty, Queue
from tkinter import filedialog, messagebox, ttk
from xml.etree import ElementTree as ET

_YF_MODULE = None
_NP_MODULE = None
_PD_MODULE = None
_OPENPYXL_LOAD_WORKBOOK = None


def _load_yfinance():
    global _YF_MODULE
    if _YF_MODULE is not None:
        return _YF_MODULE

    try:
        spec = importlib.util.find_spec("yfinance")
        if spec is None or spec.loader is None:  # pragma: no cover - runtime only
            _YF_MODULE = False
            return None

        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
    except Exception:  # pragma: no cover - runtime only
        _YF_MODULE = False
        return None

    _YF_MODULE = module
    return module


def _load_numpy():
    global _NP_MODULE
    if _NP_MODULE is not None:
        return _NP_MODULE
    import numpy as np  # type: ignore

    _NP_MODULE = np
    return np


def _load_pandas():
    global _PD_MODULE
    if _PD_MODULE is not None:
        return _PD_MODULE
    import pandas as pd  # type: ignore

    _PD_MODULE = pd
    return pd


def _load_workbook_loader():
    global _OPENPYXL_LOAD_WORKBOOK
    if _OPENPYXL_LOAD_WORKBOOK is not None:
        return _OPENPYXL_LOAD_WORKBOOK
    from openpyxl import load_workbook

    _OPENPYXL_LOAD_WORKBOOK = load_workbook
    return load_workbook


META_FILE_SUFFIX = "108nodes new meta.xlsx"
META_DATE_RE = re.compile(r"(?P<date>\d{2}\.\d{2}\.\d{4})\s+108nodes new meta\.xlsx$", re.IGNORECASE)
META_START_DATE = dt.date(2026, 4, 24)
EMA_MODEL_PREDICTION_START_DATE = dt.date(2026, 5, 1)
META_SHEET_NAME = "basic"
META_START_ROW = 6
META_COLS_MULTI = {
    "ticker": "B",
    "1w_pred": "BW", "1w_std": "CG",
    "4w_pred": "BY", "4w_std": "CH",
    "13w_pred": "CA", "13w_std": "CI",
    "26w_pred": "CC", "26w_std": "CJ",
    "52w_pred": "CE", "52w_std": "CK"
}
HORIZON_SCORE_WEIGHTS = {
    "1w": 1.0,
    "4w": 1.0,
    "13w": 1.0,
    "26w": 1.0,
    "52w": 1.0,
}
AGGREGATION_MODEL_VERSION = 7
DEFAULT_EMA_ALPHAS = {
    "1w": 1.0,
    "4w": 0.9,
    "13w": 0.8,
    "26w": 0.7,
    "52w": 0.5,
}
DEFAULT_FALLBACK_UNCERTAINTY = 1.0
DEFAULT_RISK_EXPONENT = 1.0
MIN_VALID_RETURN = -1.0
ENABLE_HOLDING_FALLBACK_PREDICTIONS = True
MIDWEEK_REPRICING_CONFIG = {
    "1w": {"damping": 0.45, "cap": 0.05},
    "4w": {"damping": 0.65, "cap": 0.06},
    "13w": {"damping": 0.80, "cap": 0.08},
    "26w": {"damping": 0.90, "cap": 0.10},
    "52w": {"damping": 1.00, "cap": 0.12},
}
DEFAULT_REMOTE_PREDICTION_MANIFEST_URL = (
    "https://raw.githubusercontent.com/vacpekar-cell/Portfolio-Manager/main/manifest.json"
)
DEFAULT_REMOTE_PREDICTIONS_ENABLED = True
DISPLAY_METRIC_HORIZON = "52w"


def application_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        executable = getattr(sys, "executable", "")
        if executable:
            return Path(executable).resolve().parent
    return Path(__file__).resolve().parent


@dataclass
class ObjectiveConfig:
    risk_exponent: float = DEFAULT_RISK_EXPONENT


@dataclass
class StockRecord:
    ticker: str
    sharpe: float
    forecast_pct: float
    std_pct: float
    upside_pct: float = 0.0
    is_synthetic: bool = False
    horizon_data: dict | None = None
    score: float = 0.0
    cvar10_pct: float | None = None



@dataclass
class TradeLine:
    ticker: str
    current_czk: float
    target_czk: float
    delta_czk: float
    action: str
    sharpe: float
    forecast_pct: float
    std_pct: float
    upside_pct: float = 0.0
    horizon_data: dict | None = None
    score: float = 0.0


@dataclass
class OptimizationResult:
    records: list[StockRecord]
    weights: np.ndarray
    sharpe: float
    expected_return: float
    volatility: float
    sold_czk: float
    penalty_czk: float
    invested_czk: float
    trades: list[TradeLine]
    target_total_czk: float
    current_total_czk: float
    extra_cash_czk: float
    position_changes: int
    min_trade_czk: float


@dataclass(frozen=True)
class PositionCountPlan:
    requested: int
    planned: int
    min_required_by_max_weight: int
    max_allowed_by_min_size: int | None
    infeasible_reason: str = ""


def _safe_float(value, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, str):
        cleaned = value.replace("%", "").replace(" ", "").replace(",", ".").strip()
        if not cleaned:
            return default
        return float(cleaned)
    return float(value)


def _is_finite_number(value) -> bool:
    try:
        return math.isfinite(float(value))
    except (TypeError, ValueError):
        return False


def _is_valid_return(value) -> bool:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return False
    return math.isfinite(numeric) and numeric > MIN_VALID_RETURN


def _effective_target_positions_for_min_size(
    target_positions: int,
    total_capital_czk: float,
    min_position_czk: float,
) -> int:
    target = max(1, int(target_positions))
    minimum = max(0.0, float(min_position_czk))
    capital = max(0.0, float(total_capital_czk))
    if minimum <= 0.0 or capital <= 0.0:
        return target
    affordable = max(1, int(capital // minimum))
    return max(1, min(target, affordable))


def _position_count_plan_for_constraints(
    target_positions: int,
    total_capital_czk: float,
    min_position_czk: float,
    max_position_weight: float | None = None,
) -> PositionCountPlan:
    target = max(1, int(target_positions))
    capital = max(0.0, float(total_capital_czk))
    minimum = max(0.0, float(min_position_czk))

    min_required = 1
    weight = None
    if max_position_weight is not None:
        try:
            weight = float(max_position_weight)
        except (TypeError, ValueError):
            weight = None
        if weight is not None and math.isfinite(weight) and weight > 0.0:
            min_required = max(1, int(math.ceil((1.0 - 1e-9) / min(1.0, weight))))
        else:
            weight = None

    max_allowed = None
    if minimum > 0.0:
        max_allowed = int(capital // minimum) if capital > 0.0 else 0
        if max_allowed < 1:
            return PositionCountPlan(
                requested=target,
                planned=1,
                min_required_by_max_weight=min_required,
                max_allowed_by_min_size=max_allowed,
                infeasible_reason=(
                    f"Minimální velikost pozice {minimum:,.0f} CZK je vyšší než "
                    f"celý dostupný kapitál {capital:,.0f} CZK."
                ),
            )

    planned = max(target, min_required)
    if max_allowed is not None:
        if max_allowed < min_required:
            cap_czk = capital * weight if weight is not None else None
            if cap_czk is not None:
                reason = (
                    f"Minimální velikost pozice {minimum:,.0f} CZK dovolí nejvýše "
                    f"{max_allowed} pozic, ale max váha {weight * 100:.2f} % vyžaduje "
                    f"alespoň {min_required} pozic (max {cap_czk:,.0f} CZK na jednu pozici)."
                )
            else:
                reason = (
                    f"Minimální velikost pozice {minimum:,.0f} CZK dovolí nejvýše "
                    f"{max_allowed} pozic."
                )
            return PositionCountPlan(
                requested=target,
                planned=max(1, max_allowed),
                min_required_by_max_weight=min_required,
                max_allowed_by_min_size=max_allowed,
                infeasible_reason=reason,
            )
        planned = min(planned, max_allowed)

    return PositionCountPlan(
        requested=target,
        planned=max(1, planned),
        min_required_by_max_weight=min_required,
        max_allowed_by_min_size=max_allowed,
    )


TICKER_ALIASES = {
    # Trading 212 can keep stale broker symbols after corporate actions.
    "AMEH": "ASTH",   # Apollo Medical Holdings -> Astrana Health, Feb 2024
    "FB": "META",     # Facebook -> Meta Platforms ticker change
    "CRDB": "CRD-B",  # Crawford & Company class B appears in brokers without separator
    "CRD/B": "CRD-B",
    "AAXN": "AXON",   # Axon Enterprise ticker change (formerly AAXN)
}

ISIN_TICKER_ALIASES = {
    "US03763A2078": "ASTH",
    "US30303M1027": "META",
    "US2246331076": "CRD-B",
    "USN070592100": "ASML",
    "US06684L1035": "BZUN",
    "US88557W1018": "QFIN",
    "US00724F1012": "ADBE",
    "US01973R1014": "ALSN",
    "US02079K3059": "GOOGL",
    "US0231351067": "AMZN",
    "US04914Y1029": "ATLC",
    "CH0334081137": "CRSP",
    "US22658D1000": "CRCT",
    "US26603R1068": "DUOL",
    "US2854091087": "ELMD",
    "IL0011582033": "FVRR",
    "US5763231090": "MTZ",
    "US64110L1061": "NFLX",
    "US70450Y1038": "PYPL",
    "US79466L3024": "CRM",
    "KYG851581069": "STNE",
    "US88160R1014": "TSLA",
    "US98923T1043": "ZDGE",
    "CA06849F1080": "B",
}

NAME_TICKER_ALIASES = {
    "ASTRANA HEALTH": "ASTH",
    "META PLATFORMS": "META",
    "CRAWFORD & CO": "CRD-B",
    "CRAWFORD & COMPANY": "CRD-B",
    "ZEDGE": "ZDGE",
    "ZEDGE INC": "ZDGE",
    "BARRICK MINING": "B",
    "BARRICK GOLD": "B",
}

TRADING212_BLOCKED_BROKER_TICKERS = {
    # Stale/removed Trading 212 instruments observed in live pie updates.
    "B_US_EQ",
    "WLKP_US_EQ",
}

TRADING212_APP_TICKER_BROKER_ALIASES = {
    # Trading 212 UI can show Barrick as NYSE:B, while the API metadata/pie
    # instrument identifier is still GOLD_US_EQ.
    "B": "GOLD_US_EQ",
}


def normalize_ticker_symbol(ticker: str) -> str:
    symbol = str(ticker or "").strip().upper()
    if re.search(r"_[A-Z]{2,3}_(EQ|ETF|STOCK)$", symbol):
        symbol = re.sub(r"_[A-Z]{2,3}_(?:EQ|ETF|STOCK)$", "", symbol)
    elif re.search(r"_(EQ|ETF|STOCK)_[A-Z]{2,3}$", symbol):
        symbol = re.sub(r"_(?:EQ|ETF|STOCK)_[A-Z]{2,3}$", "", symbol)
    if symbol in TICKER_ALIASES:
        return TICKER_ALIASES[symbol]
    normalized = symbol.replace("/", "-").replace(".", "-").replace("_", "-")
    return TICKER_ALIASES.get(normalized, normalized)


def resolve_instrument_ticker(ticker: str, name: str = "", isin: str = "") -> str:
    isin_key = str(isin or "").strip().upper()
    name_key = re.sub(r"\s+", " ", str(name or "").strip().upper())
    for prefix, mapped in NAME_TICKER_ALIASES.items():
        if name_key.startswith(prefix):
            return mapped
    if isin_key in ISIN_TICKER_ALIASES:
        return ISIN_TICKER_ALIASES[isin_key]
    
    allow_online_isin_lookup = str(os.environ.get("PORTFOLIO_ONLINE_ISIN_LOOKUP", "")).strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }

    # Optional online ISIN to ticker resolution. Disabled by default because bulk
    # Trading 212 mapping would otherwise make thousands of slow network calls.
    if allow_online_isin_lookup and isin_key and re.match(r"^[A-Z]{2}[A-Z0-9]{9}\d$", isin_key):
        import urllib.request
        import json
        try:
            url = f"https://query2.finance.yahoo.com/v1/finance/search?q={isin_key}"
            req = urllib.request.Request(
                url, 
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
            )
            with urllib.request.urlopen(req, timeout=3) as response:
                data = json.loads(response.read().decode("utf-8"))
                quotes = data.get("quotes", [])
                best_symbol = None
                for q in quotes:
                    symbol = q.get("symbol", "")
                    exchange = q.get("exchange", "")
                    if exchange in ["NMS", "NYQ", "ASE", "NGS"] or not "." in symbol:
                        best_symbol = symbol
                        break
                if not best_symbol and quotes:
                    best_symbol = quotes[0].get("symbol", "")
                
                if best_symbol:
                    base = best_symbol.split(".")[0]
                    cleaned = re.sub(r"^\d+", "", base)
                    if cleaned:
                        ISIN_TICKER_ALIASES[isin_key] = cleaned
                        return cleaned
        except Exception:
            pass

    normalized = normalize_ticker_symbol(ticker)
    if normalized in TICKER_ALIASES:
        return TICKER_ALIASES[normalized]
    return normalized


def _downside_tail_risk(tail_return: float, floor: float = 1e-6) -> float:
    """Downside risk for STARR from an absolute lower-tail return."""
    return max(-float(tail_return), float(floor))


def _starr_score(expected_return: float, downside_risk: float, negative_penalty: float = 1.0) -> float:
    """STARR-like score robust for negative expected returns.

    For positive returns: classic return / downside-risk.
    For negative returns: adds explicit downside penalty so larger risk is always worse.
    """
    tail = max(float(downside_risk), 1e-6)
    exp_ret = float(expected_return)
    if exp_ret >= 0:
        return exp_ret / tail
    return (exp_ret / tail) - negative_penalty * tail


def _coerce_horizon_values(value, default: tuple[float, float, float, float] | None = None) -> tuple[float, float, float, float]:
    if default is None:
        default = (0.0, 0.0, 0.0, 0.0)
    try:
        if isinstance(value, dict):
            p = value.get("pred", value.get("mean", default[0]))
            cvar10 = value.get("cvar10", value.get("q10", default[1]))
            q90 = value.get("q90", value.get("upside", default[2]))
            uncertainty = value.get("uncertainty", value.get("nejistota", default[3]))
            return float(p), float(cvar10), float(q90), float(uncertainty or 0.0)
        if isinstance(value, (list, tuple)):
            p = float(value[0]) if len(value) > 0 else default[0]
            cvar10 = float(value[1]) if len(value) > 1 else default[1]
            q90 = float(value[2]) if len(value) > 2 else default[2]
            uncertainty = float(value[3]) if len(value) > 3 and value[3] is not None else default[3]
            return p, cvar10, q90, uncertainty
    except (TypeError, ValueError):
        pass
    return default


def _record_horizon_values(record: StockRecord, horizon: str) -> tuple[float, float, float]:
    hdata = record.horizon_data or {}
    if horizon in hdata:
        p, cvar10, q90, _ = _coerce_horizon_values(hdata[horizon])
        return float(p), float(cvar10), float(q90)
    return float(record.forecast_pct), _record_aggregate_cvar10(record), float(record.upside_pct)


def _record_like_horizon_values(record, horizon: str = DISPLAY_METRIC_HORIZON) -> tuple[float, float, float]:
    if isinstance(record, StockRecord):
        return _record_horizon_values(record, horizon)

    forecast = float(getattr(record, "forecast_pct", 0.0) or 0.0)
    cvar10 = getattr(record, "cvar10_pct", None)
    if cvar10 is None or not _is_finite_number(cvar10):
        downside = max(float(getattr(record, "std_pct", 0.0) or 0.0), 1e-6)
        cvar10 = min(forecast - 1e-8, -downside)
    upside = float(getattr(record, "upside_pct", forecast) or forecast)

    hdata = getattr(record, "horizon_data", None) or {}
    if isinstance(hdata, dict) and horizon in hdata:
        p, q10, q90, _ = _coerce_horizon_values(hdata[horizon], (forecast, float(cvar10), upside, 0.0))
        return float(p), float(q10), float(q90)
    return forecast, float(cvar10), upside


def _format_score(value) -> str:
    return f"{float(value):.4f}" if _is_finite_number(value) else "-"


def _format_return_pct(value) -> str:
    return f"{float(value) * 100:.2f}" if _is_finite_number(value) else "-"


def _record_aggregate_cvar10(record: StockRecord) -> float:
    if record.cvar10_pct is not None and _is_finite_number(record.cvar10_pct):
        return float(record.cvar10_pct)
    mean_return = float(record.forecast_pct)
    downside = max(float(record.std_pct), 1e-6)
    return min(mean_return - 1e-8, -downside)


def _aggregate_horizon_values(horizon_data: dict) -> tuple[float, float, float]:
    agg_pred = 1.0
    agg_cvar10 = 1.0
    agg_q90 = 1.0
    scoring_weight_sum = 0.0

    for horizon in ("1w", "4w", "13w", "26w", "52w"):
        if horizon not in horizon_data:
            continue
        p, cvar10, q90, _ = _coerce_horizon_values(horizon_data[horizon])
        if not (_is_valid_return(p) and _is_valid_return(cvar10) and _is_valid_return(q90)):
            continue
        score_weight = float(HORIZON_SCORE_WEIGHTS.get(horizon, 0.0))
        if score_weight <= 0.0:
            continue
        agg_pred *= math.pow(max(0.0, 1.0 + p), score_weight)
        agg_cvar10 *= math.pow(max(0.0, 1.0 + cvar10), score_weight)
        agg_q90 *= math.pow(max(0.0, 1.0 + q90), score_weight)
        scoring_weight_sum += score_weight

    if scoring_weight_sum <= 0.0:
        return 0.0, 0.0, 0.0
    inv_scoring_weight = 1.0 / scoring_weight_sum
    return (
        math.pow(max(0.0, agg_pred), inv_scoring_weight) - 1.0,
        math.pow(max(0.0, agg_cvar10), inv_scoring_weight) - 1.0,
        math.pow(max(0.0, agg_q90), inv_scoring_weight) - 1.0,
    )


def _power_starr_score(expected_return: float, downside_risk: float, alpha: float, negative_penalty: float = 1.0) -> float:
    """Generalized (Power) STARR score.
    
    For positive returns: expected_return / downside_risk^alpha.
    For negative returns: includes downside penalty raised to the power of alpha.
    """
    tail = max(float(downside_risk), 1e-6)
    exp_ret = float(expected_return)
    denom = math.pow(tail, max(0.0, float(alpha)))
    if exp_ret >= 0:
        return exp_ret / denom
    return (exp_ret / denom) - negative_penalty * denom


def _downside_adjusted_score(mean_return: float, cvar10_return: float, risk_exponent: float) -> float:
    downside_risk = _downside_tail_risk(cvar10_return)
    return _power_starr_score(mean_return, downside_risk, risk_exponent)


def _record_objective_score(record: StockRecord, config: ObjectiveConfig) -> float:
    return _downside_adjusted_score(
        float(record.forecast_pct),
        _record_aggregate_cvar10(record),
        config.risk_exponent,
    )


def _diversified_portfolio_quantiles(
    tickers: list[str],
    weights: list[float],
    preds: list[float],
    cvar10s: list[float],
    q90s: list[float],
    corr_matrix,
) -> tuple[float, float, float]:
    """Compute portfolio-level mean, Q10, Q90 with diversification.

    Instead of naively averaging individual quantiles (which assumes ρ=1),
    we treat each stock's downside/upside spread as a "semi-volatility" and
    combine them through the correlation matrix:

        portfolio_spread = sqrt(w' * Σ_spread * w)

    where Σ_spread[i,j] = spread_i * spread_j * ρ[i,j].

    This correctly produces narrower fans for diversified portfolios.
    """
    np = _load_numpy()
    n = len(tickers)
    w = np.array(weights, dtype=float)
    p = np.array(preds, dtype=float)
    c10 = np.array(cvar10s, dtype=float)
    q90 = np.array(q90s, dtype=float)
    corr = np.asarray(corr_matrix, dtype=float)
    if corr.shape != (n, n):
        corr = np.eye(n)
    else:
        corr = np.nan_to_num(corr, nan=0.0, posinf=0.0, neginf=0.0)
        corr = (corr + corr.T) / 2.0
        corr = np.clip(corr, -0.99, 0.99)
        np.fill_diagonal(corr, 1.0)

    # Portfolio mean = weighted average of individual means
    port_mean = float(w @ p)

    # Downside spread per stock: pred_i - cvar10_i (positive if cvar10 < pred)
    down_spread = np.maximum(p - c10, 1e-8)
    # Upside spread per stock: q90_i - pred_i (positive if q90 > pred)
    up_spread = np.maximum(q90 - p, 1e-8)

    # Build spread covariance matrices and compute portfolio-level spread
    # Σ_down[i,j] = down_spread_i * down_spread_j * ρ[i,j]
    cov_down = np.outer(down_spread, down_spread) * corr
    cov_up = np.outer(up_spread, up_spread) * corr

    port_down_var = max(float(w @ cov_down @ w), 0.0)
    port_up_var = max(float(w @ cov_up @ w), 0.0)

    port_down_spread = port_down_var ** 0.5
    port_up_spread = port_up_var ** 0.5

    port_q10 = port_mean - port_down_spread
    port_q90 = port_mean + port_up_spread

    return port_mean, port_q10, port_q90


def _project_to_simplex(weights: np.ndarray) -> np.ndarray:
    np = _load_numpy()
    vector = np.asarray(weights, dtype=float)
    if vector.size == 0:
        return vector
    vector = np.nan_to_num(vector, nan=0.0, posinf=0.0, neginf=0.0)
    vector = np.maximum(vector, 0.0)
    total = vector.sum()
    if total <= 0:
        return np.ones_like(vector) / len(vector)
    if abs(total - 1.0) < 1e-12:
        return vector

    u = np.sort(vector)[::-1]
    cssv = np.cumsum(u)
    rho = np.nonzero(u * np.arange(1, len(u) + 1) > (cssv - 1.0))[0]
    if len(rho) == 0:
        return np.ones_like(vector) / len(vector)
    rho_idx = rho[-1]
    theta = (cssv[rho_idx] - 1.0) / (rho_idx + 1)
    projected = np.maximum(vector - theta, 0.0)
    projected_sum = projected.sum()
    if projected_sum <= 0:
        return np.ones_like(vector) / len(vector)
    return projected / projected_sum


def _project_to_capped_simplex(weights: np.ndarray, max_weight: float | None = None) -> np.ndarray:
    np = _load_numpy()
    if max_weight is None or max_weight >= 1.0:
        return _project_to_simplex(weights)

    vector = np.asarray(weights, dtype=float)
    if vector.size == 0:
        return vector

    cap = max(1e-12, float(max_weight))
    if cap * len(vector) < 1.0 - 1e-12:
        return _project_to_simplex(vector)

    vector = np.nan_to_num(vector, nan=0.0, posinf=0.0, neginf=0.0)
    low = float(np.min(vector) - cap)
    high = float(np.max(vector))
    for _ in range(80):
        mid = (low + high) / 2.0
        projected = np.clip(vector - mid, 0.0, cap)
        if float(projected.sum()) > 1.0:
            low = mid
        else:
            high = mid

    projected = np.clip(vector - high, 0.0, cap)
    total = float(projected.sum())
    if total <= 1e-12:
        return np.ones_like(vector) / len(vector)
    if abs(total - 1.0) > 1e-10:
        projected = np.clip(projected / total, 0.0, cap)
        remainder = 1.0 - float(projected.sum())
        if abs(remainder) > 1e-10:
            room = np.maximum(cap - projected, 0.0)
            room_sum = float(room.sum())
            if room_sum > 1e-12:
                projected += remainder * room / room_sum
    return projected


def _safe_covariance_matrix(stds: np.ndarray, corr: np.ndarray) -> np.ndarray:
    np = _load_numpy()
    corr = np.asarray(corr, dtype=float)
    if corr.size == 0:
        return corr
    corr = np.nan_to_num(corr, nan=0.0, posinf=0.0, neginf=0.0)
    np.fill_diagonal(corr, 1.0)
    shrink = 0.15
    corr = (1.0 - shrink) * corr + shrink * np.eye(len(corr))
    stds = np.asarray(stds, dtype=float).reshape(-1, 1)
    cov = corr * (stds @ stds.T)
    cov += np.eye(len(stds)) * 1e-8
    return cov


def _portfolio_horizon_quantiles(
    records: list[StockRecord],
    weights: np.ndarray,
    corr: np.ndarray,
    horizon: str,
) -> tuple[float, float, float]:
    np = _load_numpy()
    tickers = [record.ticker for record in records]
    preds = []
    cvar10s = []
    q90s = []
    for record in records:
        p, cvar10, q90 = _record_horizon_values(record, horizon)
        preds.append(p)
        cvar10s.append(cvar10)
        q90s.append(q90)
    return _diversified_portfolio_quantiles(
        tickers,
        list(np.asarray(weights, dtype=float)),
        preds,
        cvar10s,
        q90s,
        corr,
    )


def _portfolio_aggregate_quantiles(
    records: list[StockRecord],
    weights: np.ndarray,
    corr: np.ndarray,
) -> tuple[float, float, float]:
    np = _load_numpy()
    tickers = [record.ticker for record in records]
    preds = [float(record.forecast_pct) for record in records]
    cvar10s = [_record_aggregate_cvar10(record) for record in records]
    q90s = [
        max(float(record.upside_pct), float(record.forecast_pct) + 1e-8)
        for record in records
    ]
    return _diversified_portfolio_quantiles(
        tickers,
        list(np.asarray(weights, dtype=float)),
        preds,
        cvar10s,
        q90s,
        corr,
    )


def _portfolio_downside_utility(
    records: list[StockRecord],
    weights: np.ndarray,
    corr: np.ndarray,
    config: ObjectiveConfig,
) -> tuple[float, float, float, dict[str, tuple[float, float, float]]]:
    if not records:
        return -1e9, 0.0, 0.0, {}
    mean_return, cvar10_return, q90_return = _portfolio_aggregate_quantiles(records, weights, corr)
    downside_risk = _downside_tail_risk(cvar10_return)
    objective = _downside_adjusted_score(mean_return, cvar10_return, config.risk_exponent)
    return (
        float(objective),
        float(mean_return),
        float(downside_risk),
        {"aggregate": (float(mean_return), float(cvar10_return), float(q90_return))},
    )


def _objective_from_target_amounts_downside(
    target_amounts: np.ndarray,
    records: list[StockRecord],
    corr: np.ndarray,
    current_amounts: np.ndarray,
    total_capital: float,
    sale_penalty_rate: float,
    dropped_czk: float,
    config: ObjectiveConfig,
) -> tuple[float, float, float, float, float]:
    np = _load_numpy()
    target_amounts = np.asarray(target_amounts, dtype=float)
    current_amounts = np.asarray(current_amounts, dtype=float)
    weights = target_amounts / max(float(total_capital), 1e-12)
    sold_amount = float(dropped_czk) + float(np.maximum(current_amounts - target_amounts, 0.0).sum())
    penalty_czk = sold_amount * sale_penalty_rate
    penalty_return = penalty_czk / max(float(total_capital), 1e-12)
    raw_objective, expected_return, downside_risk, _ = _portfolio_downside_utility(records, weights, corr, config)
    objective = raw_objective - penalty_return
    expected_profit_czk = (expected_return - penalty_return) * float(total_capital)
    risk_czk = downside_risk * float(total_capital)
    return objective, expected_profit_czk, risk_czk, sold_amount, penalty_czk


def _resolve_target_amounts_after_penalty(
    weights: np.ndarray,
    current_amounts: np.ndarray,
    total_capital: float,
    sale_penalty_rate: float,
    max_iter: int = 100,
) -> tuple[np.ndarray, float, float, float]:
    np = _load_numpy()
    weights = _project_to_simplex(weights)
    current_amounts = np.asarray(current_amounts, dtype=float)

    invested_capital = float(total_capital)
    for _ in range(max_iter):
        target_amounts = weights * invested_capital
        sold_amount = float(np.maximum(current_amounts - target_amounts, 0.0).sum())
        penalty_czk = sold_amount * sale_penalty_rate
        next_invested_capital = max(0.0, float(total_capital) - penalty_czk)
        if abs(next_invested_capital - invested_capital) < 1e-8:
            invested_capital = next_invested_capital
            break
        invested_capital = next_invested_capital

    target_amounts = weights * invested_capital
    sold_amount = float(np.maximum(current_amounts - target_amounts, 0.0).sum())
    penalty_czk = sold_amount * sale_penalty_rate
    invested_capital = max(0.0, float(total_capital) - penalty_czk)
    target_amounts = weights * invested_capital
    sold_amount = float(np.maximum(current_amounts - target_amounts, 0.0).sum())
    penalty_czk = sold_amount * sale_penalty_rate
    return target_amounts, invested_capital, sold_amount, penalty_czk


def _objective_from_target_amounts(
    target_amounts: np.ndarray,
    expected_returns: np.ndarray,
    cov: np.ndarray,
    current_amounts: np.ndarray,
    sale_penalty_rate: float,
    dropped_czk: float = 0.0,
) -> tuple[float, float, float, float, float]:
    np = _load_numpy()
    target_amounts = np.asarray(target_amounts, dtype=float)
    current_amounts = np.asarray(current_amounts, dtype=float)
    expected_returns = np.asarray(expected_returns, dtype=float)
    sold_amount = dropped_czk + float(np.maximum(current_amounts - target_amounts, 0.0).sum())
    penalty_czk = sold_amount * sale_penalty_rate
    expected_profit_czk = float(target_amounts @ expected_returns) - penalty_czk
    risk_czk = float(np.sqrt(max(float(target_amounts @ cov @ target_amounts), 0.0)))
    if risk_czk <= 1e-12:
        return -1e9, expected_profit_czk, risk_czk, sold_amount, penalty_czk
    objective = expected_profit_czk / risk_czk
    return objective, expected_profit_czk, risk_czk, sold_amount, penalty_czk


def _portfolio_objective(
    weights: np.ndarray,
    expected_returns: np.ndarray,
    cov: np.ndarray,
    current_amounts: np.ndarray,
    total_capital: float,
    sale_penalty_rate: float,
    dropped_czk: float = 0.0,
) -> tuple[float, float, float, float]:
    target_amounts = _project_to_simplex(weights) * total_capital
    objective, expected_profit_czk, risk_czk, sold_amount, _ = _objective_from_target_amounts(
        target_amounts,
        expected_returns,
        cov,
        current_amounts,
        sale_penalty_rate,
        dropped_czk,
    )
    effective_return = expected_profit_czk / max(total_capital, 1e-12)
    effective_volatility = risk_czk / max(total_capital, 1e-12)
    return objective, effective_return, effective_volatility, sold_amount


def _turnover_sold_for_weights(
    weights: np.ndarray,
    current_amounts: np.ndarray,
    total_capital: float,
    dropped_czk: float = 0.0,
) -> float:
    np = _load_numpy()
    target_amounts = np.asarray(weights, dtype=float) * float(total_capital)
    current_amounts = np.asarray(current_amounts, dtype=float)
    return float(dropped_czk) + float(np.maximum(current_amounts - target_amounts, 0.0).sum())


def _turnover_safe_seed(
    expected_returns: np.ndarray,
    current_amounts: np.ndarray,
    total_capital: float,
    synthetic_flags: np.ndarray | None = None,
) -> np.ndarray:
    np = _load_numpy()
    n = len(expected_returns)
    if n == 0:
        return np.array([])
    current_amounts = np.asarray(current_amounts, dtype=float)
    weights = np.maximum(current_amounts / max(total_capital, 1e-12), 0.0)
    extra_weight = max(0.0, 1.0 - float(weights.sum()))
    if extra_weight <= 1e-12:
        return _project_to_simplex(weights)

    allowed = np.ones(n, dtype=bool)
    if synthetic_flags is not None:
        allowed = ~np.asarray(synthetic_flags, dtype=bool)
    scores = np.where(allowed, np.maximum(expected_returns, 0.0), 0.0)
    if float(scores.sum()) <= 1e-12:
        scores = np.where(allowed, 1.0, 0.0)
    if float(scores.sum()) <= 1e-12:
        scores = np.ones(n, dtype=float)
    weights = weights + extra_weight * scores / float(scores.sum())
    return _project_to_simplex(weights)


def _optimize_weights_with_sale_penalty(
    expected_returns: np.ndarray,
    corr_matrix: np.ndarray,
    stds: np.ndarray,
    current_amounts: np.ndarray,
    total_capital: float,
    sale_penalty_rate: float,
    synthetic_flags: np.ndarray | None = None,
    stop_event: threading.Event | None = None,
    max_seconds: float = 15.0,
    dropped_czk: float = 0.0,
    max_turnover_czk: float | None = None,
    records: list[StockRecord] | None = None,
    objective_config: ObjectiveConfig | None = None,
    max_position_weight: float | None = None,
) -> tuple[np.ndarray, float, float, float]:
    np = _load_numpy()
    import scipy.optimize as sco
    n = len(expected_returns)
    if n == 0:
        return np.array([]), 0.0, 0.0, 0.0

    corr = np.asarray(corr_matrix, dtype=float)
    cov = _safe_covariance_matrix(stds, corr_matrix.copy())
    current_weights = np.asarray(current_amounts, dtype=float) / max(total_capital, 1e-12)
    weight_cap = None
    if max_position_weight is not None:
        weight_cap = min(1.0, max(1e-6, float(max_position_weight)))
    project_weights = lambda weights: _project_to_capped_simplex(weights, weight_cap)
    if max_turnover_czk is not None and dropped_czk > max_turnover_czk + 1e-6:
        weights = project_weights(_turnover_safe_seed(expected_returns, current_amounts, total_capital, synthetic_flags))
        return weights, -1e9, 0.0, 0.0

    use_downside_objective = (
        objective_config is not None
        and records is not None
        and len(records) == n
    )

    def evaluate_weights(w):
        w_projected = project_weights(w)
        if use_downside_objective:
            target_amounts = w_projected * total_capital
            obj, expected_profit_czk, risk_czk, sold_amount, _ = _objective_from_target_amounts_downside(
                target_amounts,
                records,
                corr,
                current_amounts,
                total_capital,
                sale_penalty_rate,
                dropped_czk,
                objective_config,
            )
            return (
                obj,
                expected_profit_czk / max(total_capital, 1e-12),
                risk_czk / max(total_capital, 1e-12),
                sold_amount,
            )
        return _portfolio_objective(
            w_projected, expected_returns, cov, current_amounts, total_capital, sale_penalty_rate, dropped_czk
        )

    def obj_func(w):
        obj, _, _, _ = evaluate_weights(w)
        return -obj

    seeds = []
    seeds.append(project_weights(_turnover_safe_seed(expected_returns, current_amounts, total_capital, synthetic_flags)))
    seeds.append(project_weights(np.maximum(expected_returns, 0.0)))
    seeds.append(project_weights(np.ones(n) / n))
    if current_weights.sum() > 0:
        boosted_current = current_weights + np.maximum(expected_returns, 0.0) * max(0.0, 1.0 - current_weights.sum())
        seeds.append(project_weights(boosted_current))
    else:
        seeds.append(project_weights(np.maximum(expected_returns, 0.0) + 1.0))

    try:
        inv_cov = np.linalg.pinv(cov)
        seeds.append(project_weights(inv_cov @ expected_returns))
    except np.linalg.LinAlgError:
        pass

    best_weights = seeds[0]
    best_objective, best_ret, best_vol, best_sold = evaluate_weights(best_weights)
    if max_turnover_czk is not None and best_sold > max_turnover_czk + 1e-6:
        best_objective, best_ret, best_vol = -1e9, 0.0, 0.0

    deadline = time.monotonic() + max_seconds
    constraints = [{"type": "eq", "fun": lambda w: np.sum(w) - 1.0}]
    if max_turnover_czk is not None:
        constraints.append({
            "type": "ineq",
            "fun": lambda w: float(max_turnover_czk) - _turnover_sold_for_weights(
                project_weights(w), current_amounts, total_capital, dropped_czk
            ),
        })
    bounds_list = []
    for i in range(n):
        upper = weight_cap if weight_cap is not None else 1.0
        if synthetic_flags is not None and synthetic_flags[i]:
            upper = min(upper, max(1e-6, float(current_weights[i])))
            bounds_list.append([0.0, upper])
        else:
            bounds_list.append([0.0, upper])
            
    max_sum = sum(b[1] for b in bounds_list)
    if max_sum < 1.0:
        if weight_cap is None:
            scale = 1.0 / max_sum
            for b in bounds_list:
                b[1] = min(1.0, b[1] * scale)
        else:
            return best_weights, -1e9, 0.0, 0.0
            
    bounds = tuple(tuple(b) for b in bounds_list)

    for seed in seeds:
        if stop_event is not None and stop_event.is_set():
            break
        if time.monotonic() >= deadline:
            break

        res = sco.minimize(
            obj_func,
            seed,
            method="SLSQP",
            bounds=bounds,
            constraints=constraints,
            options={"maxiter": 200, "ftol": 1e-6, "disp": False}
        )
        
        w_res = project_weights(res.x)
        obj, ret, vol, _ = evaluate_weights(w_res)
        if max_turnover_czk is not None:
            sold = _turnover_sold_for_weights(w_res, current_amounts, total_capital, dropped_czk)
            if sold > max_turnover_czk + 1e-4:
                continue
        if obj > best_objective:
            best_objective = obj
            best_ret = ret
            best_vol = vol
            best_weights = w_res.copy()

    return best_weights, best_objective, best_ret, best_vol


def _optimize_extra_cash_without_sells(
    expected_returns: np.ndarray,
    corr_matrix: np.ndarray,
    stds: np.ndarray,
    current_amounts: np.ndarray,
    total_capital: float,
    stop_event: threading.Event | None = None,
    max_seconds: float = 12.0,
) -> tuple[np.ndarray, float, float, float]:
    np = _load_numpy()
    n = len(expected_returns)
    if n == 0:
        return np.array([]), 0.0, 0.0, 0.0

    current_amounts = np.asarray(current_amounts, dtype=float)
    base_weights = current_amounts / max(total_capital, 1e-12)
    extra_cash = max(0.0, total_capital - float(current_amounts.sum()))
    if extra_cash <= 1e-12:
        weights = _project_to_simplex(base_weights)
        cov = _safe_covariance_matrix(stds, corr_matrix.copy())
        obj, gross_ret, vol, _ = _portfolio_objective(weights, expected_returns, cov, current_amounts, total_capital, 0.0)
        return weights, obj, gross_ret, vol

    alpha = extra_cash / max(total_capital, 1e-12)
    cov = _safe_covariance_matrix(stds, corr_matrix.copy())

    def weights_from_extra(extra_weights: np.ndarray) -> np.ndarray:
        return base_weights + alpha * extra_weights

    seeds = [
        np.ones(n) / n,
        _project_to_simplex(np.maximum(expected_returns, 0.0)),
    ]
    try:
        seeds.append(_project_to_simplex(np.linalg.pinv(cov) @ expected_returns))
    except np.linalg.LinAlgError:
        pass

    rng = np.random.default_rng(42)
    for _ in range(4):
        seeds.append(_project_to_simplex(rng.random(n)))

    deadline = time.monotonic() + max_seconds
    best_extra = seeds[0]
    best_weights = weights_from_extra(best_extra)
    best_obj, best_ret, best_vol, _ = _portfolio_objective(best_weights, expected_returns, cov, current_amounts, total_capital, 0.0)

    for seed in seeds:
        if stop_event is not None and stop_event.is_set():
            break
        if time.monotonic() >= deadline:
            break

        extra_alloc = _project_to_simplex(seed)
        local_step = 0.12

        for _ in range(350):
            if stop_event is not None and stop_event.is_set():
                break
            if time.monotonic() >= deadline:
                break

            weights = weights_from_extra(extra_alloc)
            obj, gross_ret, vol, _ = _portfolio_objective(weights, expected_returns, cov, current_amounts, total_capital, 0.0)
            if obj > best_obj + 1e-10:
                best_obj = obj
                best_ret = gross_ret
                best_vol = vol
                best_extra = extra_alloc.copy()
                best_weights = weights.copy()

            cov_w = cov @ weights
            grad_w = expected_returns / max(vol, 1e-12) - (gross_ret / max(vol**3, 1e-12)) * cov_w
            grad_extra = alpha * grad_w
            extra_alloc = _project_to_simplex(extra_alloc + local_step * grad_extra)
            local_step *= 0.997

    return best_weights, best_obj, best_ret, best_vol


def _objective_gradient(
    weights: np.ndarray,
    expected_returns: np.ndarray,
    cov: np.ndarray,
    current_amounts: np.ndarray,
    total_capital: float,
    sale_penalty_rate: float,
) -> np.ndarray:
    np = _load_numpy()
    weights = _project_to_simplex(weights)
    target_amounts = weights * total_capital
    sold_mask = (current_amounts - target_amounts) > 1e-9
    penalty_gradient_czk = np.where(sold_mask, sale_penalty_rate * total_capital, 0.0)
    net_returns_czk = total_capital * expected_returns - penalty_gradient_czk
    risk_czk = float(np.sqrt(max(float(target_amounts @ cov @ target_amounts), 1e-18)))
    if risk_czk <= 1e-12:
        return np.zeros_like(weights)
    net_profit_czk = float(target_amounts @ expected_returns) - float(
        np.maximum(current_amounts - target_amounts, 0.0).sum() * sale_penalty_rate
    )
    risk_grad = (total_capital * (cov @ weights)) / risk_czk
    return net_returns_czk / risk_czk - (net_profit_czk / max(risk_czk**2, 1e-18)) * risk_grad


class ReturnCache:
    def __init__(self):
        self.cache: dict[str, object] = {}

    def get_returns(self, ticker: str, log_fn):
        pd = _load_pandas()
        if ticker in self.cache:
            return self.cache[ticker]

        yf = _load_yfinance()
        if yf is None:
            log_fn(f"yfinance není dostupné, korelace pro {ticker} nastavena na 0.")
            self.cache[ticker] = pd.Series(dtype=float)
            return self.cache[ticker]

        try:
            data = yf.download(ticker, period="6mo", interval="1wk", progress=False, auto_adjust=True)
            if data.empty or "Close" not in data.columns:
                log_fn(f"Žádná data pro {ticker}, použiji nulovou korelaci.")
                returns = pd.Series(dtype=float)
            else:
                close = data["Close"]
                series = close.squeeze() if isinstance(close, pd.DataFrame) else close
                returns = pd.Series(series).pct_change().dropna()
        except Exception as exc:  # pragma: no cover - depends on runtime/network
            log_fn(f"Nepodařilo se stáhnout korelace pro {ticker}: {exc}. Použiji nulovou korelaci.")
            returns = pd.Series(dtype=float)

        self.cache[ticker] = returns
        return returns


def build_correlation_matrix(tickers: list[str], cache: ReturnCache, log_fn) -> np.ndarray:
    np = _load_numpy()
    pd = _load_pandas()
    if not tickers:
        return np.zeros((0, 0))

    series_map: dict[str, object] = {}
    for ticker in tickers:
        series_map[ticker] = cache.get_returns(ticker, log_fn)

    aligned = pd.concat(series_map, axis=1).dropna(how="all") if series_map else pd.DataFrame()
    if aligned.empty:
        log_fn("Není dost dat pro odhad korelací, použiji nulovou korelační matici.")
        return np.zeros((len(tickers), len(tickers)))

    aligned.columns = tickers
    corr = aligned.corr().reindex(index=tickers, columns=tickers).fillna(0.0)
    return corr.values


def find_all_meta_workbooks_since(base_dir: Path, start_date: dt.date) -> list[tuple[dt.date, Path]]:
    candidates = list(base_dir.glob(f"*{META_FILE_SUFFIX}"))
    candidates.extend(base_dir.joinpath("other").glob(f"*{META_FILE_SUFFIX}"))
    unique_candidates = {path.resolve(): path for path in candidates}.values()
    dated: list[tuple[dt.date, Path]] = []

    for path in unique_candidates:
        match = META_DATE_RE.search(path.name)
        if not match:
            continue
        file_date = dt.datetime.strptime(match.group("date"), "%d.%m.%Y").date()
        if file_date >= start_date:
            dated.append((file_date, path))

    dated.sort(key=lambda item: item[0])
    return dated


def _xlsx_column_index(column_letters: str) -> int:
    value = 0
    for char in column_letters.upper():
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value


def _parse_cell_reference(reference: str) -> tuple[int, int]:
    match = re.match(r"([A-Z]+)(\d+)$", reference.upper())
    if not match:
        raise ValueError(f"Neplatná reference buňky: {reference}")
    return int(match.group(2)), _xlsx_column_index(match.group(1))


def _load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    shared_path = "xl/sharedStrings.xml"
    if shared_path not in archive.namelist():
        return []

    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    strings: list[str] = []
    with archive.open(shared_path) as handle:
        for _, elem in ET.iterparse(handle, events=("end",)):
            if elem.tag == f"{namespace}si":
                parts = []
                for node in elem.iter():
                    if node.tag == f"{namespace}t" and node.text is not None:
                        parts.append(node.text)
                strings.append("".join(parts))
                elem.clear()
    return strings


def _resolve_sheet_path(archive: zipfile.ZipFile, sheet_name: str) -> str | None:
    workbook_ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    rel_id = None
    for sheet in workbook_root.findall("main:sheets/main:sheet", workbook_ns):
        if sheet.attrib.get("name") == sheet_name:
            rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            break

    if rel_id is None:
        return None

    target = None
    for rel in rel_root.findall("pkg:Relationship", workbook_ns):
        if rel.attrib.get("Id") == rel_id:
            target = rel.attrib.get("Target")
            break

    if not target:
        return None

    target = target.lstrip("/")
    if target.startswith("xl/"):
        return target
    return f"xl/{target}"


def _copy_to_temp_for_read(path: Path) -> Path:
    suffix = path.suffix or ".tmp"
    temp_file = tempfile.NamedTemporaryFile(prefix="portfolio_manager_", suffix=suffix, delete=False)
    temp_path = Path(temp_file.name)
    temp_file.close()
    try:
        shutil.copyfile(path, temp_path)
    except Exception:
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise
    return temp_path


def _fast_load_raw_data_from_meta(path: Path) -> dict[str, dict[str, float]]:
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    col_indices = {key: _xlsx_column_index(val) for key, val in META_COLS_MULTI.items()}
    wanted_cols = set(col_indices.values())
    ticker_col = col_indices["ticker"]

    with zipfile.ZipFile(path) as archive:
        sheet_path = _resolve_sheet_path(archive, META_SHEET_NAME)
        if sheet_path is None:
            raise RuntimeError(f"List '{META_SHEET_NAME}' v souboru {path.name} neexistuje.")

        shared_strings = _load_shared_strings(archive)
        results: dict[str, dict[str, float]] = {}

        with archive.open(sheet_path) as handle:
            current_row = None
            row_values: dict[int, object] = {}
            cell_type = None
            cell_ref = None
            inline_parts: list[str] = []
            value_text = None

            def finalize_row():
                nonlocal row_values, current_row
                if current_row is None or current_row < META_START_ROW:
                    row_values = {}
                    return

                ticker_val = row_values.get(ticker_col)
                if not ticker_val:
                    row_values = {}
                    return
                raw_ticker = str(ticker_val).strip().upper()
                ticker_text = normalize_ticker_symbol(raw_ticker)
                if not ticker_text:
                    row_values = {}
                    return

                metrics = {}
                for key, c_idx in col_indices.items():
                    if key == "ticker": continue
                    val = row_values.get(c_idx)
                    try:
                        metrics[key] = float(val) if val is not None else 0.0
                    except Exception:
                        metrics[key] = 0.0

                if raw_ticker != ticker_text and ticker_text in results:
                    row_values = {}
                    return
                results[ticker_text] = metrics
                row_values = {}

            for event, elem in ET.iterparse(handle, events=("start", "end")):
                if event == "start":
                    if elem.tag == f"{namespace}row":
                        current_row = int(elem.attrib.get("r", "0"))
                        row_values = {}
                    elif elem.tag == f"{namespace}c":
                        cell_type = elem.attrib.get("t")
                        cell_ref = elem.attrib.get("r")
                        inline_parts = []
                        value_text = None
                    continue

                if elem.tag == f"{namespace}v":
                    value_text = elem.text
                elif elem.tag == f"{namespace}t":
                    if cell_type == "inlineStr" and elem.text is not None:
                        inline_parts.append(elem.text)
                elif elem.tag == f"{namespace}c":
                    if cell_ref:
                        row_idx, col_idx = _parse_cell_reference(cell_ref)
                        if row_idx >= META_START_ROW and col_idx in wanted_cols:
                            cell_value: object = None
                            if cell_type == "s" and value_text is not None:
                                try:
                                    cell_value = shared_strings[int(value_text)]
                                except Exception:
                                    cell_value = value_text
                            elif cell_type == "inlineStr":
                                cell_value = "".join(inline_parts)
                            else:
                                cell_value = value_text
                            row_values[col_idx] = cell_value
                    elem.clear()
                elif elem.tag == f"{namespace}row":
                    finalize_row()
                    elem.clear()

        return results


def _load_prediction_csv_paired_with_sim(csv_path: Path, base_dir: Path, file_date: dt.date, log_fn) -> dict[str, dict[str, float]]:
    pd = _load_pandas()
    try:
        df_pred = pd.read_csv(csv_path)
    except Exception as e:
        log_fn(f"Chyba při čtení CSV {csv_path.name}: {e}")
        return {}
        
    tickers = None
    # Podpora pro nové CSV ze sítě (nn108n_v3) s nativním sloupcem Ticker
    if 'ticker' in df_pred.columns:
        tickers = df_pred['ticker'].astype(str).tolist()
    elif 'Ticker' in df_pred.columns:
        tickers = df_pred['Ticker'].astype(str).tolist()
        
    # Původní fallback logika pro starší predikce bez Tickeru
    if tickers is None:
        num_rows = len(df_pred)
        sim_dir = base_dir / "processed"
        date_str = file_date.strftime("%d.%m.%Y")
        expected_sim_path = sim_dir / f"{date_str} sim.xlsx"
        
        sim_path_to_use = None
        if expected_sim_path.exists():
            try:
                df_test = pd.read_excel(expected_sim_path)
                if len(df_test) == num_rows:
                    sim_path_to_use = expected_sim_path
            except Exception:
                pass
                
        if sim_path_to_use is None:
            candidates = []
            for p in sim_dir.glob("* sim.xlsx"):
                try:
                    df_test = pd.read_excel(p)
                    if len(df_test) == num_rows:
                        candidates.append((p.stat().st_mtime, p, df_test))
                except Exception:
                    continue
            if candidates:
                candidates.sort(key=lambda x: x[0], reverse=True)
                sim_path_to_use = candidates[0][1]
                df_sim = candidates[0][2]
            else:
                log_fn(f"Varování: Nenašel se vhodný sim.xlsx pro {csv_path.name} (očekáváno {num_rows} řádků).")
                return {}
        else:
            df_sim = pd.read_excel(sim_path_to_use)
            
        if 'Ticker' not in df_sim.columns:
            log_fn(f"Varování: {sim_path_to_use.name} neobsahuje sloupec 'Ticker'.")
            return {}
            
        tickers = df_sim['Ticker'].astype(str).tolist()
        log_fn(f"Použito párování {csv_path.name} <-> {sim_path_to_use.name}")
    else:
        log_fn(f"Použit nativní Ticker sloupec z {csv_path.name}")
        
    results: dict[str, dict[str, float]] = {}
    source_tickers: dict[str, str] = {}
    invalid_horizons = 0
    
    for idx, (_, row) in enumerate(df_pred.iterrows()):
        raw_ticker = str(tickers[idx]).strip().upper()
        ticker = normalize_ticker_symbol(raw_ticker)
        if not ticker:
            continue
            
        metrics = {}
        for horizon in ['1w', '4w', '13w', '26w', '52w']:
            # Podpora pro nový formát procent (nn108n_v3)
            stred_col = f'stred_{horizon}_%'
            q10_col = f'q10_{horizon}_%'
            q90_col = f'q90_{horizon}_%'
            
            # Starý formát reziduí
            mean_col = f'mean_{horizon}'
            q10_old = f'q10_{horizon}'
            q90_old = f'q90_{horizon}'
            
            if stred_col in df_pred.columns:
                val = row[stred_col]
                if not pd.notna(val):
                    continue
                mean_val = float(val) / 100.0

                # Downside riziko: absolutní CVaR10 hodnota (skutečný výnos v nejhorších 10 %)
                # Ukládáme přímo absolutní hodnotu, ne offset od predikce
                cvar10_col = f'cvar10_{horizon}_%'
                lower_tail = None
                if cvar10_col in df_pred.columns:
                    lower_tail = float(row[cvar10_col]) / 100.0 if pd.notna(row[cvar10_col]) else None
                elif q10_col in df_pred.columns:
                    lower_tail = float(row[q10_col]) / 100.0 if pd.notna(row[q10_col]) else None
                    
                # Upside potenciál: absolutní Q90 hodnota
                q90_val = None
                if q90_col in df_pred.columns:
                    q90_val = float(row[q90_col]) / 100.0 if pd.notna(row[q90_col]) else None

                if not (_is_valid_return(mean_val) and _is_valid_return(lower_tail) and _is_valid_return(q90_val)):
                    invalid_horizons += 1
                    continue

                metrics[f'{horizon}_pred'] = mean_val
                metrics[f'{horizon}_std'] = float(lower_tail)  # absolutní CVaR10/Q10
                metrics[f'{horizon}_up'] = float(q90_val)  # absolutní Q90
                uncertainty_col = f'nejistota_{horizon}_%'
                if uncertainty_col in df_pred.columns and pd.notna(row[uncertainty_col]):
                    uncertainty_val = float(row[uncertainty_col]) / 100.0
                    if math.isfinite(uncertainty_val):
                        metrics[f'{horizon}_unc'] = uncertainty_val
                    
            elif mean_col in df_pred.columns:
                val = row[mean_col]
                if not pd.notna(val):
                    continue
                mean_val = float(val)
                lower_tail = None
                if q10_old in df_pred.columns:
                    lower_tail = float(row[q10_old]) if pd.notna(row[q10_old]) else None
                q90_val = None
                if q90_old in df_pred.columns:
                    q90_val = float(row[q90_old]) if pd.notna(row[q90_old]) else None

                if not (_is_valid_return(mean_val) and _is_valid_return(lower_tail) and _is_valid_return(q90_val)):
                    invalid_horizons += 1
                    continue

                metrics[f'{horizon}_pred'] = mean_val
                metrics[f'{horizon}_std'] = float(lower_tail)
                metrics[f'{horizon}_up'] = float(q90_val)
                
        if metrics:
            previous_raw = source_tickers.get(ticker)
            if previous_raw is not None:
                previous_is_alias = previous_raw != ticker
                current_is_alias = raw_ticker != ticker
                if current_is_alias:
                    log_fn(
                        f"Ticker {raw_ticker} je alias pro {ticker}; řádek přeskočen, "
                        f"protože {ticker} už je v {csv_path.name} z {previous_raw}."
                    )
                    continue
                if previous_is_alias:
                    log_fn(
                        f"Ticker {raw_ticker} nahrazuje starý alias {previous_raw} "
                        f"pro {ticker} v {csv_path.name}."
                    )
            results[ticker] = metrics
            source_tickers[ticker] = raw_ticker
            
    log_fn(f"Úspěšně zpracováno {len(results)} tickerů z {csv_path.name}")
    if invalid_horizons:
        log_fn(
            f"Přeskočeno {invalid_horizons} neplatných horizontů z {csv_path.name} "
            "(chybějící metrika nebo návrat <= -100 %)."
        )
    return results

def find_all_prediction_csvs(base_dir: Path, start_date: dt.date) -> list[tuple[dt.date, Path]]:
    candidates = list(base_dir.glob("predictions_*.csv"))
    dated: list[tuple[dt.date, Path]] = []
    
    import re
    date_re = re.compile(r"predictions_(\d{4})(\d{2})(\d{2})_")
    
    for path in candidates:
        match = date_re.search(path.name)
        if match:
            y, m, d = int(match.group(1)), int(match.group(2)), int(match.group(3))
            try:
                file_date = dt.date(y, m, d)
                if file_date >= start_date:
                    dated.append((file_date, path))
            except ValueError:
                pass
                
    dated.sort(key=lambda item: item[0])
    return dated


def find_latest_prediction_csv(base_dir: Path, start_date: dt.date) -> tuple[dt.date, Path] | None:
    dated = find_all_prediction_csvs(base_dir, start_date)
    if not dated:
        return None
    return max(dated, key=lambda item: (item[0], item[1].stat().st_mtime, item[1].name))


def find_ema_prediction_csvs(base_dir: Path, start_date: dt.date) -> list[tuple[dt.date, Path]]:
    latest_by_date: dict[dt.date, tuple[dt.date, Path]] = {}
    for file_date, path in find_all_prediction_csvs(base_dir, start_date):
        current = latest_by_date.get(file_date)
        if current is None or (path.stat().st_mtime, path.name) > (current[1].stat().st_mtime, current[1].name):
            latest_by_date[file_date] = (file_date, path)
    return [latest_by_date[file_date] for file_date in sorted(latest_by_date)]


def _records_from_prediction_metrics(
    raw_data: dict[str, dict[str, float]],
    objective_config: ObjectiveConfig | None = None,
) -> list[StockRecord]:
    np = _load_numpy()
    records: list[StockRecord] = []

    for ticker, metrics in raw_data.items():
        try:
            horizon_data = {}
            for win in ["1w", "4w", "13w", "26w", "52w"]:
                if f"{win}_pred" not in metrics:
                    continue
                p = metrics.get(f"{win}_pred", 0.0)
                s = metrics.get(f"{win}_std", 0.0)
                u = metrics.get(f"{win}_up", s)
                uncertainty = metrics.get(f"{win}_unc", 0.0)
                if u == s and p != 0.0:
                    u = 2.0 * p - s
                horizon_data[win] = (p, s, u, uncertainty)

            agg_pred, agg_cvar10, agg_q90 = _aggregate_horizon_values(horizon_data)
            agg_std = _downside_tail_risk(agg_cvar10)
            sharpe = _starr_score(agg_pred, agg_std)

            if np.isfinite(sharpe) and np.isfinite(agg_pred) and np.isfinite(agg_std):
                record = StockRecord(
                    ticker=ticker,
                    sharpe=sharpe,
                    forecast_pct=agg_pred,
                    std_pct=agg_std,
                    upside_pct=agg_q90,
                    is_synthetic=False,
                    horizon_data=horizon_data,
                    cvar10_pct=agg_cvar10,
                )
                if objective_config is not None:
                    record.score = _record_objective_score(record, objective_config)
                records.append(record)
        except Exception:
            pass

    records.sort(key=lambda record: record.score if objective_config else record.sharpe, reverse=True)
    return records


def load_holding_ema_fallback_records(
    base_dir: Path,
    tickers: list[str],
    objective_config: ObjectiveConfig | None,
    log_fn,
) -> list[StockRecord]:
    tickers = [normalize_ticker_symbol(ticker) for ticker in tickers if normalize_ticker_symbol(ticker)]
    if not tickers:
        return []

    cache_path = base_dir / "ema_cache.pkl"
    if not cache_path.exists():
        return []

    try:
        import pickle
        with open(cache_path, "rb") as f:
            c_data = pickle.load(f)
    except Exception as exc:
        log_fn(f"EMA fallback cache se nepodařilo načíst: {exc}")
        return []

    if c_data.get("model_version") != AGGREGATION_MODEL_VERSION:
        return []

    emas = c_data.get("emas", {})
    if not isinstance(emas, dict):
        return []

    raw_data = {
        ticker: dict(emas[ticker])
        for ticker in tickers
        if ticker in emas and isinstance(emas.get(ticker), dict)
    }
    records = _records_from_prediction_metrics(raw_data, objective_config)
    for record in records:
        record.is_synthetic = True
    if records:
        log_fn(
            "Používám EMA fallback pro držené tickery mimo nákupní universe: "
            + ", ".join(sorted(record.ticker for record in records))
        )
    return records


def load_latest_prediction_universe(base_dir: Path, log_fn) -> tuple[Path, list[StockRecord]]:
    latest = find_latest_prediction_csv(base_dir, EMA_MODEL_PREDICTION_START_DATE)
    if latest is None:
        raise FileNotFoundError(
            f"Nenalezen žádný predictions_*.csv od {EMA_MODEL_PREDICTION_START_DATE.strftime('%d.%m.%Y')}."
        )

    file_date, path = latest
    log_fn(f"Načítám aktuální predikční universe z {path.name} ({file_date.strftime('%d.%m.%Y')}).")
    raw_data = _load_prediction_csv_paired_with_sim(path, base_dir, file_date, log_fn)
    if not raw_data:
        raise RuntimeError(f"{path.name} neobsahuje žádné použitelné predikce.")

    records = _records_from_prediction_metrics(raw_data)
    if not records:
        raise RuntimeError(f"{path.name} nevyrobil žádné použitelné záznamy.")
    return path, records


def load_historical_meta_and_calculate_ema(base_dir: Path, alphas: dict[str, float], log_fn) -> list[StockRecord]:
    np = _load_numpy()

    file_list = find_all_meta_workbooks_since(base_dir, EMA_MODEL_PREDICTION_START_DATE)
    pred_list = find_ema_prediction_csvs(base_dir, EMA_MODEL_PREDICTION_START_DATE)

    if not file_list and not pred_list:
        raise FileNotFoundError(
            f"Nenalezen žádný meta soubor ani csv predikce od "
            f"{EMA_MODEL_PREDICTION_START_DATE.strftime('%d.%m.%Y')}."
        )

    import pickle
    cache_path = base_dir / "ema_cache.pkl"

    emas: dict[str, dict[str, float]] = {}
    prev_actuals: dict[str, dict[str, float]] = {}
    last_seen_date: dict[str, dt.date] = {}
    latest_metric_keys: dict[str, set[str]] = {}
    latest_source_path: Path | None = None
    latest_event_key: tuple[dt.date, int, str] | None = None
    prev_date = None

    def event_sort_key(event: tuple[str, dt.date, Path]) -> tuple[dt.date, int, str]:
        source_kind, file_date, path = event
        return (file_date, 0 if source_kind == "Excel" else 1, path.name)

    loaded_from_cache = False
    if cache_path.exists():
        try:
            with open(cache_path, "rb") as f:
                c_data = pickle.load(f)
            if c_data.get("alphas") == alphas and c_data.get("model_version") == AGGREGATION_MODEL_VERSION:
                emas = c_data["emas"]
                prev_actuals = c_data.get("prev_actuals", {})
                last_seen_date = c_data.get("last_seen_date", {})
                prev_date = c_data.get("prev_date")
                latest_metric_keys = {
                    str(ticker): set(keys)
                    for ticker, keys in c_data.get("latest_metric_keys", {}).items()
                }
                cached_source = c_data.get("latest_source_path")
                latest_source_path = Path(cached_source) if cached_source else None
                raw_event_key = c_data.get("latest_event_key")
                if isinstance(raw_event_key, (list, tuple)) and len(raw_event_key) == 3:
                    latest_event_key = (
                        dt.date.fromisoformat(str(raw_event_key[0])),
                        int(raw_event_key[1]),
                        str(raw_event_key[2]),
                    )
                elif latest_source_path is not None and prev_date is not None:
                    source_order = 1 if latest_source_path.name.lower().startswith("predictions_") else 0
                    latest_event_key = (prev_date, source_order, latest_source_path.name)
                loaded_from_cache = True
                log_fn("Úspěšně načtena vyrovnávací paměť EMA (shodné alfa hodnoty).")
        except Exception as e:
            log_fn(f"Nelze načíst cache: {e}. Počítám od nuly.")

    events = [("Excel", d, p) for d, p in file_list] + [("CSV", d, p) for d, p in pred_list]
    events.sort(key=event_sort_key)

    if loaded_from_cache and (latest_event_key is not None or prev_date is not None):
        if latest_event_key is not None:
            events = [event for event in events if event_sort_key(event) > latest_event_key]
        else:
            events = [event for event in events if event[1] > prev_date]
        if not events:
            log_fn("Nevidím novější vstupy. Používám EMA cache bez cenového fallbacku.")
        else:
            log_fn(f"EMA navazuje na cache a zpracuje {len(events)} nových vstupů.")
    else:
        log_fn(
            f"Zjištěno {len(file_list)} Excelů a {len(pred_list)} denních CSV souborů "
            f"pro EMA analýzu od {EMA_MODEL_PREDICTION_START_DATE.strftime('%d.%m.%Y')}."
        )

    def update_ema(file_date: dt.date, path: Path, raw_data: dict[str, dict[str, float]], source_kind: str) -> None:
        nonlocal latest_metric_keys, latest_source_path, latest_event_key, prev_date

        cleaned_today: dict[str, dict[str, float]] = {}
        for ticker, raw_metrics in raw_data.items():
            if not raw_metrics:
                continue
            metrics = dict(raw_metrics)

            # Přizpůsobení staré směrodatné odchylky na úroveň 10% kvantilu (1.28 násobek).
            if source_kind == "Excel" and file_date < dt.date(2026, 4, 25):
                for mk in list(metrics.keys()):
                    if "std" in mk:
                        metrics[mk] *= 1.28

            cleaned_today[ticker] = metrics

        if not cleaned_today:
            return

        latest_metric_keys = {ticker: set(metrics.keys()) for ticker, metrics in cleaned_today.items()}
        latest_source_path = path
        latest_event_key = event_sort_key((source_kind, file_date, path))

        for ticker, metrics in cleaned_today.items():
            last_seen_date[ticker] = file_date
            prev_actuals[ticker] = metrics.copy()

            if ticker not in emas:
                emas[ticker] = metrics.copy()
                continue

            for metric_key, val in metrics.items():
                a = 1.0
                for win in ["1w", "4w", "13w", "26w", "52w"]:
                    if metric_key.startswith(win):
                        a = alphas.get(win, 1.0)
                        break
                emas[ticker][metric_key] = a * val + (1.0 - a) * emas[ticker].get(metric_key, val)

        prev_date = file_date

    for source_kind, file_date, path in events:
        log_fn(f"Zpracovávám ({source_kind}): {path.name} ({file_date.strftime('%d.%m.%Y')})")
        try:
            if source_kind == "Excel":
                raw_data = _fast_load_raw_data_from_meta(path)
            else:
                raw_data = _load_prediction_csv_paired_with_sim(path, base_dir, file_date, log_fn)
            if not raw_data:
                continue
            update_ema(file_date, path, raw_data, source_kind)
        except Exception as e:
            log_fn(f"Varování: Přeskakuji {path.name} kvůli chybě: {e}")
            continue

    if not latest_metric_keys:
        latest_metric_keys = {ticker: set(metrics.keys()) for ticker, metrics in emas.items()}

    records = []
    for ticker, keys in latest_metric_keys.items():
        metrics_all = emas.get(ticker)
        if not metrics_all:
            continue
        metrics = {key: metrics_all[key] for key in keys if key in metrics_all}
        try:
            horizon_data = {}
            for win in ["1w", "4w", "13w", "26w", "52w"]:
                if f"{win}_pred" not in metrics or f"{win}_std" not in metrics:
                    continue
                p = metrics.get(f"{win}_pred")
                s = metrics.get(f"{win}_std")  # absolutní CVaR10/Q10
                u = metrics.get(f"{win}_up", s)
                uncertainty = metrics.get(f"{win}_unc", 0.0)
                if not (_is_valid_return(p) and _is_valid_return(s) and _is_valid_return(u)):
                    continue
                if u == s and p != 0.0:
                    # Staré Excel vstupy nemají horní kvantil; odhad držíme jen pro jejich kompatibilitu.
                    u = 2.0 * p - s
                horizon_data[win] = (float(p), float(s), float(u), float(uncertainty or 0.0))

            if not horizon_data:
                continue

            agg_pred, agg_cvar10, agg_q90 = _aggregate_horizon_values(horizon_data)
            agg_std = _downside_tail_risk(agg_cvar10)
            sharpe = _starr_score(agg_pred, agg_std)

            if np.isfinite(sharpe) and np.isfinite(agg_pred) and np.isfinite(agg_std):
                records.append(StockRecord(
                    ticker=ticker,
                    sharpe=sharpe,
                    forecast_pct=agg_pred,
                    std_pct=agg_std,
                    upside_pct=agg_q90,
                    is_synthetic=False,
                    horizon_data=horizon_data,
                    cvar10_pct=agg_cvar10,
                ))
        except Exception:
            pass

    if events:
        try:
            c_data = {
                "alphas": alphas,
                "model_version": AGGREGATION_MODEL_VERSION,
                "emas": emas,
                "prev_actuals": prev_actuals,
                "last_seen_date": last_seen_date,
                "prev_date": prev_date,
                "latest_metric_keys": latest_metric_keys,
                "latest_source_path": str(latest_source_path) if latest_source_path else None,
                "latest_event_key": [
                    latest_event_key[0].isoformat(),
                    latest_event_key[1],
                    latest_event_key[2],
                ] if latest_event_key else None,
            }
            with open(cache_path, "wb") as f:
                pickle.dump(c_data, f)
            log_fn("Stav EMA analýzy uložen do cache.")
        except Exception as e:
            log_fn(f"Nepodařilo se uložit cache: {e}")

    if not records:
        raise RuntimeError("Vygenerováno 0 validních záznamů po spojení EMA.")
    records.sort(key=lambda record: record.sharpe, reverse=True)
    return records


def parse_holdings_text(text: str) -> dict[str, float]:
    holdings: dict[str, float] = {}
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    for line in lines:
        parts = [part for part in re.split(r"[;\t, ]+", line) if part]
        if len(parts) < 2:
            raise ValueError(f"Řádek '{line}' nemá formát TICKER ČÁSTKA.")

        ticker = normalize_ticker_symbol(parts[0])
        amount_text = "".join(parts[1:]).replace("Kc", "").replace("CZK", "").replace("czk", "")
        amount = _safe_float(amount_text)
        if amount < 0:
            raise ValueError(f"Řádek '{line}' obsahuje zápornou částku.")
        holdings[ticker] = holdings.get(ticker, 0.0) + amount

    return holdings


class PortfolioManager:
    def __init__(
        self,
        records: list[StockRecord],
        current_holdings_czk: dict[str, float],
        extra_cash_czk: float,
        target_positions: int,
        sale_penalty_rate: float,
        max_turnover_czk: float | None,
        max_position_replacements: int | None,
        min_trade_czk: float,
        min_position_czk: float = 0.0,
        max_position_weight: float | None = None,
        no_sell_tickers: set[str] | list[str] | None = None,
        objective_config: ObjectiveConfig | None = None,
        log_fn=print,
        stop_event: threading.Event | None = None,
    ):
        self.objective_config = objective_config or ObjectiveConfig()
        for record in records:
            record.score = _record_objective_score(record, self.objective_config)
        self.record_map = {record.ticker: record for record in records}
        self.records = sorted(records, key=lambda rec: rec.score, reverse=True)
        self.current_holdings_czk = {normalize_ticker_symbol(ticker): float(amount) for ticker, amount in current_holdings_czk.items() if amount > 0}
        self.extra_cash_czk = float(extra_cash_czk)
        self.target_positions = max(1, int(target_positions))
        self.sale_penalty_rate = max(0.0, float(sale_penalty_rate))
        self.max_turnover_czk = None if max_turnover_czk is None else max(0.0, float(max_turnover_czk))
        self.max_position_replacements = None if max_position_replacements is None else max(0, int(max_position_replacements))
        self.max_sell_trades = self.max_position_replacements
        self.min_trade_czk = max(0.0, float(min_trade_czk))
        self.min_position_czk = max(0.0, float(min_position_czk))
        self.max_position_weight = None if max_position_weight is None else min(1.0, max(1e-6, float(max_position_weight)))
        self.no_sell_tickers = {
            normalize_ticker_symbol(ticker)
            for ticker in (no_sell_tickers or [])
            if normalize_ticker_symbol(ticker) in self.current_holdings_czk
        }
        self.log = log_fn
        self.stop_event = stop_event
        self.cache = ReturnCache()

    def _should_stop(self) -> bool:
        return self.stop_event is not None and self.stop_event.is_set()

    def _eligible_current_holdings(self) -> dict[str, float]:
        eligible: dict[str, float] = {}
        for ticker, amount in self.current_holdings_czk.items():
            if ticker in self.record_map:
                eligible[ticker] = amount
            else:
                self.log(f"Ticker {ticker} není v posledním meta souboru, proto ho nebudu optimalizovat.")
        return eligible

    def _build_candidate_pool(self, current_holdings: dict[str, float]) -> list[StockRecord]:
        positive = [record for record in self.records if record.score > 0 and record.forecast_pct > 0]
        if not positive:
            positive = [record for record in self.records if record.score > 0]
        if not positive:
            positive = self.records[:]

        current_tickers = set(current_holdings)
        pool: list[StockRecord] = []
        added: set[str] = set()
        pool_target_positions = self.target_positions
        if self.max_position_weight is not None:
            pool_target_positions = max(
                pool_target_positions,
                int(math.ceil((1.0 - 1e-9) / max(self.max_position_weight, 1e-9))),
            )

        for ticker in current_tickers:
            record = self.record_map.get(ticker)
            if record is not None and ticker not in added:
                pool.append(record)
                added.add(ticker)

        for record in positive[: max(pool_target_positions * 6, 180)]:
            if record.ticker not in added:
                pool.append(record)
                added.add(record.ticker)

        if len(pool) < pool_target_positions:
            for record in self.records:
                if record.ticker not in added:
                    pool.append(record)
                    added.add(record.ticker)
                if len(pool) >= pool_target_positions:
                    break

        return pool

    def _no_sell_mode(self) -> bool:
        return False

    def _evaluate_subset(
        self,
        subset: list[StockRecord],
        current_holdings: dict[str, float],
        total_capital: float,
    ) -> tuple[float, np.ndarray, float, float, float]:
        np = _load_numpy()
        tickers = [record.ticker for record in subset]
        corr = build_correlation_matrix(tickers, self.cache, self.log)
        expected_returns = np.array([record.forecast_pct for record in subset], dtype=float)
        stds = np.array([record.std_pct for record in subset], dtype=float)
        current_amounts = np.array([current_holdings.get(ticker, 0.0) for ticker in tickers], dtype=float)

        if any(ticker in self.no_sell_tickers for ticker in current_holdings if ticker not in tickers):
            weights = _turnover_safe_seed(expected_returns, current_amounts, total_capital, None)
            return -1e9, weights, 0.0, 0.0, total_capital

        dropped_czk = sum(amt for t, amt in current_holdings.items() if t not in tickers)
        if self.max_turnover_czk is not None and dropped_czk > self.max_turnover_czk + 1e-6:
            weights = _turnover_safe_seed(expected_returns, current_amounts, total_capital, None)
            return -1e9, weights, 0.0, 0.0, dropped_czk
        synth_flags = np.array([record.is_synthetic for record in subset], dtype=bool)

        weights, objective, gross_return, volatility = _optimize_weights_with_sale_penalty(
            expected_returns,
            corr,
            stds,
            current_amounts,
            total_capital,
            self.sale_penalty_rate,
            dropped_czk=dropped_czk,
            synthetic_flags=synth_flags,
            stop_event=self.stop_event,
            max_turnover_czk=self.max_turnover_czk,
            records=subset,
            objective_config=self.objective_config,
            max_position_weight=self.max_position_weight,
        )

        sold_czk = dropped_czk + float(np.maximum(current_amounts - weights * total_capital, 0.0).sum())
        return objective, weights, gross_return, volatility, sold_czk

    def _is_feasible_subset(self, subset_tickers: set[str], current_tickers: set[str]) -> bool:
        if self.max_position_weight is not None and len(subset_tickers) * self.max_position_weight < 1.0 - 1e-9:
            return False
        if any(ticker in current_tickers and ticker not in subset_tickers for ticker in self.no_sell_tickers):
            return False
        if self.max_sell_trades is None:
            return True
        sold_out = len(current_tickers - subset_tickers)
        return sold_out <= self.max_sell_trades

    def _initial_subset(
        self,
        pool: list[StockRecord],
        current_holdings: dict[str, float],
        total_capital: float,
    ) -> list[StockRecord]:
        current_sorted = sorted(
            (self.record_map[ticker] for ticker in current_holdings if ticker in self.record_map),
            key=lambda rec: (current_holdings.get(rec.ticker, 0.0), rec.score),
            reverse=True,
        )
        subset: list[StockRecord] = []
        added: set[str] = set()
        position_plan = _position_count_plan_for_constraints(
            self.target_positions,
            total_capital,
            self.min_position_czk,
            self.max_position_weight,
        )
        effective_target = position_plan.planned
        protected_sorted = [record for record in current_sorted if record.ticker in self.no_sell_tickers]
        if protected_sorted and len(protected_sorted) > effective_target:
            self.log(
                f"Ochrana před prodejem vyžaduje ponechat {len(protected_sorted)} pozic; "
                f"cílový počet {self.target_positions} proto beru jako měkký."
            )
            effective_target = len(protected_sorted)
        if self.max_sell_trades is not None and current_sorted:
            required_keep = max(0, len(current_sorted) - self.max_sell_trades)
            if required_keep > effective_target:
                self.log(
                    f"Limit {self.max_sell_trades} prodejů vyžaduje ponechat alespoň "
                    f"{required_keep} současných pozic; cílový počet {self.target_positions} proto beru jako měkký."
                )
                effective_target = required_keep
        if self.max_turnover_czk is not None and current_sorted:
            dropped_if_omitted = sum(float(current_holdings.get(rec.ticker, 0.0)) for rec in current_sorted)
            required_keep = 0
            for record in current_sorted:
                if dropped_if_omitted <= self.max_turnover_czk + 1e-6:
                    break
                dropped_if_omitted -= float(current_holdings.get(record.ticker, 0.0))
                required_keep += 1
            if required_keep > effective_target:
                self.log(
                    f"Limit protočení {self.max_turnover_czk:,.0f} CZK vyžaduje ponechat "
                    f"alespoň {required_keep} současných pozic; cílový počet {self.target_positions} proto beru jako měkký."
                )
                effective_target = required_keep

        for record in protected_sorted:
            if len(subset) >= effective_target:
                break
            subset.append(record)
            added.add(record.ticker)

        for record in current_sorted:
            if len(subset) >= effective_target:
                break
            if record.ticker in added:
                continue
            subset.append(record)
            added.add(record.ticker)

        for record in pool:
            if len(subset) >= effective_target:
                break
            if record.ticker in added:
                continue
            subset.append(record)
            added.add(record.ticker)

        return subset[: effective_target]

    def _enforce_turnover_limit_on_targets(
        self,
        subset: list[StockRecord],
        target_vector: np.ndarray,
        current_holdings: dict[str, float],
        total_capital: float,
    ) -> tuple[list[StockRecord], np.ndarray]:
        if self.max_turnover_czk is None:
            return subset, target_vector
        np = _load_numpy()
        target_map = {record.ticker: float(target) for record, target in zip(subset, target_vector)}

        def sold_total() -> float:
            return sum(max(float(amount) - target_map.get(ticker, 0.0), 0.0) for ticker, amount in current_holdings.items())

        sold = sold_total()
        if sold <= self.max_turnover_czk + 1e-6:
            return subset, target_vector

        excess = sold - self.max_turnover_czk
        sold_positions = sorted(
            (
                (ticker, max(float(amount) - target_map.get(ticker, 0.0), 0.0))
                for ticker, amount in current_holdings.items()
            ),
            key=lambda item: item[1],
            reverse=True,
        )
        restored = 0.0
        for ticker, sale_amount in sold_positions:
            if excess <= 1e-6:
                break
            if sale_amount <= 0:
                continue
            add_back = min(sale_amount, excess)
            target_map[ticker] = target_map.get(ticker, 0.0) + add_back
            excess -= add_back
            restored += add_back

        if restored <= 1e-6:
            return subset, target_vector

        need_reduce = restored
        buy_positions = sorted(
            (
                (ticker, max(target - float(current_holdings.get(ticker, 0.0)), 0.0))
                for ticker, target in target_map.items()
            ),
            key=lambda item: item[1],
            reverse=True,
        )
        for ticker, buy_amount in buy_positions:
            if need_reduce <= 1e-6:
                break
            if buy_amount <= 0:
                continue
            reduction = min(buy_amount, need_reduce)
            target_map[ticker] = max(0.0, target_map.get(ticker, 0.0) - reduction)
            need_reduce -= reduction

        if need_reduce > 1e-4:
            self.log(
                "Limit protočení je tak nízký, že jsem musel výrazně omezit nákupy; "
                "výsledek bude velmi blízko současnému portfoliu."
            )

        final_records = [
            self.record_map[ticker]
            for ticker, target in target_map.items()
            if target > 1e-6 and ticker in self.record_map
        ]
        final_records.sort(key=lambda rec: target_map.get(rec.ticker, 0.0), reverse=True)
        final_vector = np.array([target_map.get(record.ticker, 0.0) for record in final_records], dtype=float)
        total_target = float(final_vector.sum())
        if total_target > 0 and abs(total_target - total_capital) > 1e-4:
            delta = total_capital - total_target
            candidates = [
                idx for idx, record in enumerate(final_records)
                if target_map.get(record.ticker, 0.0) >= current_holdings.get(record.ticker, 0.0)
            ] or list(range(len(final_records)))
            if candidates:
                piece = delta / len(candidates)
                for idx in candidates:
                    final_vector[idx] = max(0.0, final_vector[idx] + piece)
                for record, target in zip(final_records, final_vector):
                    target_map[record.ticker] = float(target)

        final_sold = sum(
            max(float(amount) - target_map.get(ticker, 0.0), 0.0)
            for ticker, amount in current_holdings.items()
        )
        if final_sold > self.max_turnover_czk + 1e-4:
            self.log(
                f"Bezpečnostní korekce: návrh by prodal {final_sold:,.0f} CZK, "
                f"limit je {self.max_turnover_czk:,.0f} CZK."
            )
        return final_records, final_vector

    def _enforce_no_sell_tickers_on_targets(
        self,
        subset: list[StockRecord],
        target_vector: np.ndarray,
        current_holdings: dict[str, float],
        total_capital: float,
    ) -> tuple[list[StockRecord], np.ndarray]:
        protected = {ticker for ticker in self.no_sell_tickers if ticker in current_holdings}
        if not protected:
            return subset, target_vector
        np = _load_numpy()
        target_map = {record.ticker: float(target) for record, target in zip(subset, target_vector)}
        restored = 0.0
        protected_sales = 0
        for ticker in protected:
            current_amount = float(current_holdings.get(ticker, 0.0))
            target_amount = float(target_map.get(ticker, 0.0))
            if target_amount < current_amount - 1.0:
                restored += current_amount - target_amount
                protected_sales += 1
                target_map[ticker] = current_amount

        if restored <= 1e-6:
            return subset, target_vector

        need_reduce = restored
        buy_positions = sorted(
            (
                (ticker, max(float(target) - float(current_holdings.get(ticker, 0.0)), 0.0))
                for ticker, target in target_map.items()
            ),
            key=lambda item: item[1],
            reverse=True,
        )
        for ticker, buy_amount in buy_positions:
            if need_reduce <= 1e-6:
                break
            if buy_amount <= 0:
                continue
            reduction = min(buy_amount, need_reduce)
            target_map[ticker] = max(0.0, target_map.get(ticker, 0.0) - reduction)
            need_reduce -= reduction

        if need_reduce > 1e-4:
            sellable_positions = sorted(
                (
                    (ticker, float(target_map.get(ticker, 0.0)))
                    for ticker in target_map
                    if ticker not in protected and float(target_map.get(ticker, 0.0)) > 0
                ),
                key=lambda item: item[1],
                reverse=True,
            )
            for ticker, target_amount in sellable_positions:
                if need_reduce <= 1e-6:
                    break
                reduction = min(target_amount, need_reduce)
                target_map[ticker] = max(0.0, target_map.get(ticker, 0.0) - reduction)
                need_reduce -= reduction

        if need_reduce > 1e-4:
            self.log(
                "Ochrana před prodejem je tak přísná, že nebylo možné plně přerozdělit kapitál; "
                "výsledek bude velmi blízko současnému portfoliu."
            )

        final_records = [
            self.record_map[ticker]
            for ticker, target in target_map.items()
            if target > 1e-6 and ticker in self.record_map
        ]
        final_records.sort(key=lambda rec: target_map.get(rec.ticker, 0.0), reverse=True)
        final_vector = np.array([target_map.get(record.ticker, 0.0) for record in final_records], dtype=float)

        total_target = float(final_vector.sum())
        if total_target > 0 and abs(total_target - total_capital) > 1e-4:
            delta = total_capital - total_target
            if delta > 0:
                candidates = [
                    idx for idx, record in enumerate(final_records)
                    if record.ticker not in protected
                    and target_map.get(record.ticker, 0.0) >= current_holdings.get(record.ticker, 0.0)
                ] or [
                    idx for idx, record in enumerate(final_records)
                    if record.ticker not in protected
                ] or list(range(len(final_records)))
                if candidates:
                    piece = delta / len(candidates)
                    for idx in candidates:
                        final_vector[idx] = max(0.0, final_vector[idx] + piece)
            else:
                need_reduce = -delta
                candidates = sorted(
                    (
                        (
                            idx,
                            max(
                                float(final_vector[idx])
                                - (float(current_holdings.get(record.ticker, 0.0)) if record.ticker in protected else 0.0),
                                0.0,
                            ),
                        )
                        for idx, record in enumerate(final_records)
                    ),
                    key=lambda item: item[1],
                    reverse=True,
                )
                for idx, room in candidates:
                    if need_reduce <= 1e-6:
                        break
                    if room <= 0:
                        continue
                    reduction = min(room, need_reduce)
                    final_vector[idx] = max(0.0, final_vector[idx] - reduction)
                    need_reduce -= reduction
                if need_reduce > 1e-4:
                    self.log(
                        "Ochrana před prodejem nechala návrh bez dostatečného volného prostoru; "
                        "hotovostní dorovnání proto nemuselo být úplné."
                    )

        if protected_sales:
            self.log(f"Ochrana před prodejem: ponecháno {protected_sales} chráněných pozic beze snížení.")
        return final_records, final_vector

    def _enforce_max_position_weight_on_targets(
        self,
        subset: list[StockRecord],
        target_vector: np.ndarray,
        current_holdings: dict[str, float],
        total_capital: float,
    ) -> tuple[list[StockRecord], np.ndarray]:
        if self.max_position_weight is None:
            return subset, target_vector

        np = _load_numpy()
        cap_czk = float(total_capital) * self.max_position_weight
        if cap_czk <= 0:
            return subset, target_vector

        target_map = {record.ticker: float(target) for record, target in zip(subset, target_vector)}
        protected = {ticker for ticker in self.no_sell_tickers if ticker in current_holdings}
        protected_over_cap = 0
        capped_positions = 0
        excess = 0.0

        for ticker, target in list(target_map.items()):
            current_amount = float(current_holdings.get(ticker, 0.0))
            if ticker in protected and current_amount > cap_czk + 1.0:
                target_map[ticker] = max(target, current_amount)
                protected_over_cap += 1
                continue
            if target > cap_czk + 1e-6:
                target_map[ticker] = cap_czk
                excess += target - cap_czk
                capped_positions += 1

        if excess <= 1e-6:
            if protected_over_cap:
                self.log(
                    f"Max váha akcie: {protected_over_cap} chráněných pozic je nad limitem, "
                    "protože jsou označené jako Neprodat."
                )
            return subset, target_vector

        candidates = sorted(
            (
                (
                    record.ticker,
                    max(cap_czk - target_map.get(record.ticker, 0.0), 0.0),
                    record.score,
                )
                for record in subset
                if not (
                    record.ticker in protected
                    and float(current_holdings.get(record.ticker, 0.0)) > cap_czk + 1.0
                )
            ),
            key=lambda item: item[2],
            reverse=True,
        )
        for ticker, room, _ in candidates:
            if excess <= 1e-6:
                break
            if room <= 1e-6:
                continue
            addition = min(room, excess)
            target_map[ticker] = target_map.get(ticker, 0.0) + addition
            excess -= addition

        if excess > 1e-4:
            raise RuntimeError(
                "Maximální váha akcie je v konfliktu s počtem pozic nebo s chráněnými pozicemi. "
                "Zvyšte limit, cílový počet akcií, nebo uvolněte některé pozice z Neprodat."
            )

        final_records = [
            self.record_map[ticker]
            for ticker, target in target_map.items()
            if target > 1e-6 and ticker in self.record_map
        ]
        final_records.sort(key=lambda rec: target_map.get(rec.ticker, 0.0), reverse=True)
        final_vector = np.array([target_map.get(record.ticker, 0.0) for record in final_records], dtype=float)

        if capped_positions:
            self.log(
                f"Max váha akcie: sníženo {capped_positions} pozic na limit "
                f"{self.max_position_weight * 100:.2f} %."
            )
        if protected_over_cap:
            self.log(
                f"Max váha akcie: {protected_over_cap} chráněných pozic zůstává nad limitem, "
                "protože jsou označené jako Neprodat."
            )
        return final_records, final_vector

    def _enforce_min_position_size_on_targets(
        self,
        subset: list[StockRecord],
        target_vector: np.ndarray,
        current_holdings: dict[str, float],
        total_capital: float,
    ) -> tuple[list[StockRecord], np.ndarray]:
        floor_czk = float(self.min_position_czk)
        if floor_czk <= 0.0:
            return subset, target_vector

        np = _load_numpy()
        target_map = {record.ticker: float(target) for record, target in zip(subset, target_vector)}
        protected = {ticker for ticker in self.no_sell_tickers if ticker in current_holdings}
        protected_below = 0
        small_records: list[StockRecord] = []

        for record in subset:
            target = float(target_map.get(record.ticker, 0.0))
            if target <= 1e-6 or target >= floor_czk - 1e-6:
                continue
            if record.ticker in protected and float(current_holdings.get(record.ticker, 0.0)) > 0.0:
                protected_below += 1
                continue
            small_records.append(record)

        if not small_records:
            if protected_below:
                self.log(
                    f"Min velikost pozice: {protected_below} chráněných pozic zůstává pod "
                    f"{floor_czk:,.0f} CZK, protože jsou označené jako Neprodat."
                )
            return subset, target_vector

        small_records.sort(key=lambda rec: (target_map.get(rec.ticker, 0.0), rec.score))
        released = 0.0
        for record in small_records:
            released += max(0.0, float(target_map.get(record.ticker, 0.0)))
            target_map[record.ticker] = 0.0

        cap_czk = None
        if self.max_position_weight is not None:
            cap_czk = float(total_capital) * float(self.max_position_weight)

        def remaining_room(record: StockRecord) -> float:
            if cap_czk is None:
                return float("inf")
            if record.ticker in protected and float(current_holdings.get(record.ticker, 0.0)) > cap_czk:
                return 0.0
            return max(0.0, cap_czk - float(target_map.get(record.ticker, 0.0)))

        recipients = sorted(
            (
                record
                for record in subset
                if float(target_map.get(record.ticker, 0.0)) > 0.0
                and record not in small_records
                and remaining_room(record) > 1e-6
            ),
            key=lambda rec: rec.score,
            reverse=True,
        )

        if not recipients and released > 0.0:
            fallback = max(small_records, key=lambda rec: rec.score)
            target_map[fallback.ticker] = released
            recipients = [fallback]
            released = 0.0

        for record in recipients:
            if released <= 1e-6:
                break
            room = remaining_room(record)
            addition = released if math.isinf(room) else min(released, room)
            if addition <= 1e-6:
                continue
            target_map[record.ticker] = float(target_map.get(record.ticker, 0.0)) + addition
            released -= addition

        if released > 1e-6 and recipients:
            target_map[recipients[0].ticker] = float(target_map.get(recipients[0].ticker, 0.0)) + released
            released = 0.0

        final_records = [
            self.record_map[ticker]
            for ticker, target in target_map.items()
            if target > 1e-6 and ticker in self.record_map
        ]
        final_records.sort(key=lambda rec: target_map.get(rec.ticker, 0.0), reverse=True)
        final_vector = np.array([target_map.get(record.ticker, 0.0) for record in final_records], dtype=float)

        self.log(
            f"Min velikost pozice: sloučeno {len(small_records)} pozic pod "
            f"{floor_czk:,.0f} CZK."
        )
        if protected_below:
            self.log(
                f"Min velikost pozice: {protected_below} chráněných pozic zůstává pod limitem, "
                "protože jsou označené jako Neprodat."
            )
        return final_records, final_vector

    def _trim_targets_to_available_capital(
        self,
        subset: list[StockRecord],
        target_vector: np.ndarray,
        current_holdings: dict[str, float],
        total_capital: float,
    ) -> tuple[list[StockRecord], np.ndarray]:
        np = _load_numpy()
        target_map = {record.ticker: float(target) for record, target in zip(subset, target_vector)}
        excess = float(sum(target_map.values())) - float(total_capital)
        if excess <= 1e-4:
            return subset, target_vector

        protected = {ticker for ticker in self.no_sell_tickers if ticker in current_holdings}
        candidates = sorted(
            (
                (
                    ticker,
                    max(
                        float(target) - (float(current_holdings.get(ticker, 0.0)) if ticker in protected else 0.0),
                        0.0,
                    ),
                )
                for ticker, target in target_map.items()
            ),
            key=lambda item: item[1],
            reverse=True,
        )
        for ticker, room in candidates:
            if excess <= 1e-6:
                break
            if room <= 0:
                continue
            reduction = min(room, excess)
            target_map[ticker] = max(0.0, target_map.get(ticker, 0.0) - reduction)
            excess -= reduction

        if excess > 1e-4:
            raise RuntimeError(
                "Nepodařilo se sestavit hotovostně vyrovnaný návrh bez prodeje chráněných pozic. "
                "Zkontroluj prosím nastavení Neprodat, cílový počet pozic a dostupnou hotovost."
            )

        final_records = [
            self.record_map[ticker]
            for ticker, target in target_map.items()
            if target > 1e-6 and ticker in self.record_map
        ]
        final_records.sort(key=lambda rec: target_map.get(rec.ticker, 0.0), reverse=True)
        final_vector = np.array([target_map.get(record.ticker, 0.0) for record in final_records], dtype=float)
        self.log("Bezpečnostní kontrola hotovosti: cílové nákupy byly dorovnány na dostupný kapitál.")
        return final_records, final_vector

    def _enforce_sell_count_limit_on_targets(
        self,
        subset: list[StockRecord],
        target_vector: np.ndarray,
        current_holdings: dict[str, float],
        total_capital: float,
    ) -> tuple[list[StockRecord], np.ndarray]:
        if self.max_sell_trades is None:
            return subset, target_vector
        np = _load_numpy()
        target_map = {record.ticker: float(target) for record, target in zip(subset, target_vector)}
        sales = []
        for ticker, amount in current_holdings.items():
            sale_amount = max(float(amount) - target_map.get(ticker, 0.0), 0.0)
            if sale_amount > 1.0:
                record = self.record_map.get(ticker)
                score = record.score if record is not None else 0.0
                sales.append((ticker, sale_amount, score))

        if len(sales) <= self.max_sell_trades:
            return subset, target_vector

        sales.sort(key=lambda item: (item[1], -item[2]), reverse=True)
        allowed = {ticker for ticker, _, _ in sales[: self.max_sell_trades]}
        blocked = [(ticker, sale_amount) for ticker, sale_amount, _ in sales if ticker not in allowed]
        restored = 0.0
        for ticker, sale_amount in blocked:
            current_amount = float(current_holdings.get(ticker, 0.0))
            target_map[ticker] = current_amount
            restored += sale_amount

        need_reduce = restored
        buy_positions = sorted(
            (
                (ticker, max(float(target) - float(current_holdings.get(ticker, 0.0)), 0.0))
                for ticker, target in target_map.items()
            ),
            key=lambda item: item[1],
            reverse=True,
        )
        for ticker, buy_amount in buy_positions:
            if need_reduce <= 1e-6:
                break
            if buy_amount <= 0:
                continue
            reduction = min(buy_amount, need_reduce)
            target_map[ticker] = max(0.0, target_map.get(ticker, 0.0) - reduction)
            need_reduce -= reduction

        if need_reduce > 1e-4:
            self.log(
                "Limit počtu prodejů je tak nízký, že jsem nenašel dost nákupů k omezení; "
                "výsledek může být velmi blízko současnému portfoliu."
            )

        final_sales = [
            ticker
            for ticker, amount in current_holdings.items()
            if max(float(amount) - target_map.get(ticker, 0.0), 0.0) > 1.0
        ]
        if len(final_sales) > self.max_sell_trades:
            for ticker in final_sales[self.max_sell_trades:]:
                target_map[ticker] = float(current_holdings.get(ticker, 0.0))

        final_records = [
            self.record_map[ticker]
            for ticker, target in target_map.items()
            if target > 1e-6 and ticker in self.record_map
        ]
        final_records.sort(key=lambda rec: target_map.get(rec.ticker, 0.0), reverse=True)
        final_vector = np.array([target_map.get(record.ticker, 0.0) for record in final_records], dtype=float)
        total_target = float(final_vector.sum())
        if total_target > 0 and abs(total_target - total_capital) > 1e-4:
            delta = total_capital - total_target
            if delta > 0:
                candidates = [
                    idx for idx, record in enumerate(final_records)
                    if target_map.get(record.ticker, 0.0) >= current_holdings.get(record.ticker, 0.0)
                ] or list(range(len(final_records)))
                if candidates:
                    piece = delta / len(candidates)
                    for idx in candidates:
                        final_vector[idx] = max(0.0, final_vector[idx] + piece)
            else:
                need_reduce = -delta
                candidates = sorted(
                    (
                        (idx, max(float(final_vector[idx]) - float(current_holdings.get(record.ticker, 0.0)), 0.0))
                        for idx, record in enumerate(final_records)
                    ),
                    key=lambda item: item[1],
                    reverse=True,
                )
                for idx, room in candidates:
                    if need_reduce <= 1e-6:
                        break
                    if room <= 0:
                        continue
                    reduction = min(room, need_reduce)
                    final_vector[idx] = max(0.0, final_vector[idx] - reduction)
                    need_reduce -= reduction
            for record, target in zip(final_records, final_vector):
                target_map[record.ticker] = float(target)

        final_sell_count = sum(
            1
            for ticker, amount in current_holdings.items()
            if max(float(amount) - target_map.get(ticker, 0.0), 0.0) > 1.0
        )
        self.log(
            f"Limit počtu prodejů: návrh snížen z {len(sales)} na {final_sell_count} "
            f"prodejních obchodů."
        )
        return final_records, final_vector

    def _optimize_single_run(self, seed_noise: float = 0.0) -> dict[str, float]:
        np = _load_numpy()
        
        # Save original records and state
        orig_records = self.records
        orig_record_map = self.record_map
        
        if seed_noise > 0.0:
            # Create perturbed copies of records
            perturbed_records = []
            for rec in orig_records:
                import copy
                rec_copy = copy.copy(rec)
                noise = np.random.normal(0.0, seed_noise)
                rec_copy.forecast_pct = rec_copy.forecast_pct + noise
                rec_copy.score = _record_objective_score(rec_copy, self.objective_config)
                rec_copy.sharpe = _starr_score(rec_copy.forecast_pct, rec_copy.std_pct)
                perturbed_records.append(rec_copy)
                
            self.records = sorted(perturbed_records, key=lambda rec: rec.score, reverse=True)
            self.record_map = {record.ticker: record for record in self.records}
            
        try:
            current_holdings = self._eligible_current_holdings()
            current_total = sum(current_holdings.values())
            total_capital = current_total + self.extra_cash_czk
            if total_capital <= 0:
                raise ValueError("Celkový kapitál musí být kladný.")

            pool = self._build_candidate_pool(current_holdings)
            if not pool:
                raise RuntimeError("Nejsou dostupní žádní kandidáti pro optimalizaci.")

            current_tickers = set(current_holdings)
            subset = self._initial_subset(pool, current_holdings, total_capital)
            if not subset:
                raise RuntimeError("Nepodařilo se sestavit počáteční portfolio.")

            best_subset = subset
            best_objective, best_weights, best_return, best_vol, best_sold = self._evaluate_subset(best_subset, current_holdings, total_capital)

            ordered_pool = sorted(pool, key=lambda record: record.score, reverse=True)
            improvement = True
            iteration = 0
            global_deadline = time.monotonic() + 15.0
            while improvement and iteration < 6:
                if self._should_stop() or time.monotonic() >= global_deadline:
                    break
                improvement = False
                iteration += 1
                ranked_subset = sorted(zip(best_subset, best_weights), key=lambda item: (item[1], item[0].score))
                outside = [record for record in ordered_pool if record.ticker not in {item.ticker for item in best_subset}]
                outside = outside[: max(self.target_positions * 3, 40)]

                for replace_record, _ in ranked_subset[: max(3, min(len(ranked_subset), 8))]:
                    if self._should_stop() or time.monotonic() >= global_deadline:
                        break
                    for candidate in outside:
                        if self._should_stop() or time.monotonic() >= global_deadline:
                            break
                        if replace_record.ticker == candidate.ticker:
                            continue
                        trial_subset = [record for record in best_subset if record.ticker != replace_record.ticker] + [candidate]
                        trial_tickers = {record.ticker for record in trial_subset}
                        if len(trial_tickers) != len(trial_subset):
                            continue
                        if not self._is_feasible_subset(trial_tickers, current_tickers):
                            continue

                        trial_objective, trial_weights, trial_return, trial_vol, trial_sold = self._evaluate_subset(
                            trial_subset, current_holdings, total_capital
                        )
                        if self.max_turnover_czk is not None and trial_sold > self.max_turnover_czk + 1e-6:
                            continue
                        if trial_objective > best_objective + 1e-5:
                            best_subset = trial_subset
                            best_objective = trial_objective
                            best_weights = trial_weights
                            best_return = trial_return
                            best_vol = trial_vol
                            best_sold = trial_sold
                            improvement = True
                            break
                    if improvement:
                        break

            if self.max_turnover_czk is not None and best_sold > self.max_turnover_czk + 1e-6:
                best_subset = self._initial_subset(pool, current_holdings, total_capital)
                best_objective, best_weights, best_return, best_vol, best_sold = self._evaluate_subset(
                    best_subset, current_holdings, total_capital
                )

            target_vector = best_weights * total_capital
            target_amount_map = {record.ticker: target_czk for record, target_czk in zip(best_subset, target_vector)}
            return target_amount_map
            
        finally:
            self.records = orig_records
            self.record_map = orig_record_map

    def optimize(self) -> OptimizationResult:
        np = _load_numpy()
        current_holdings = self._eligible_current_holdings()
        current_total = sum(current_holdings.values())
        total_capital = current_total + self.extra_cash_czk
        if total_capital <= 0:
            raise ValueError("Celkový kapitál musí být kladný.")

        position_plan = _position_count_plan_for_constraints(
            self.target_positions,
            total_capital,
            self.min_position_czk,
            self.max_position_weight,
        )
        if position_plan.infeasible_reason:
            raise ValueError(position_plan.infeasible_reason)
        if self.min_position_czk > 0.0 and position_plan.planned < self.target_positions:
            self.log(
                f"Min velikost pozice {self.min_position_czk:,.0f} CZK snižuje efektivní "
                f"cílový počet z {self.target_positions} na {position_plan.planned}."
            )
        if position_plan.planned > self.target_positions:
            self.log(
                f"Max váha akcie {self.max_position_weight * 100:.2f} % vyžaduje alespoň "
                f"{position_plan.min_required_by_max_weight} pozic; cílový počet "
                f"{self.target_positions} proto beru jako měkký a použiju {position_plan.planned}."
            )
            
        pool = self._build_candidate_pool(current_holdings)
        if not pool:
            raise RuntimeError("Nejsou dostupní žádní kandidáti pro optimalizaci.")
            
        num_runs = 10
        perturbation_std = 0.005
        
        self.log(f"Spouštím Resampled (Ensemble) optimalizaci s {num_runs} běhy pro zvýšení diverzifikace.")
        self.log("Tento proces eliminuje citlivost na drobné výkyvy predikcí a přirozeně navrhne robustnější portfolio.")
        
        results = []
        for run in range(num_runs):
            if self._should_stop():
                break
            self.log(f"== Optimalizační krok {run + 1} / {num_runs} (výpočet spuštěn...) ==")
            
            t0 = time.monotonic()
            noise = 0.0 if run == 0 else perturbation_std
            res = self._optimize_single_run(seed_noise=noise)
            results.append(res)
            
            elapsed = time.monotonic() - t0
            self.log(f"   Krok {run + 1} / {num_runs} dokončen za {elapsed:.2f}s (nalezeno {len(res)} pozic).")
            
        if self._should_stop() or not results:
            raise RuntimeError("Optimalizace byla přerušena.")
            
        self.log("== Konsolidace a zprůměrování výsledných portfolií ==")
        
        # Average target amounts across all successful runs
        sum_target_amounts = {}
        all_tickers_seen = set()
        
        for res in results:
            for ticker, amount in res.items():
                all_tickers_seen.add(ticker)
                sum_target_amounts[ticker] = sum_target_amounts.get(ticker, 0.0) + amount
                
        # Average amounts
        avg_target_amounts = {
            ticker: sum_target_amounts[ticker] / len(results)
            for ticker in all_tickers_seen
        }
        
        best_subset = [rec for rec in pool if avg_target_amounts.get(rec.ticker, 0.0) > 0.0]
        if not best_subset:
            best_subset = pool[:self.target_positions]
            
        target_amount_map = avg_target_amounts.copy()
        
        for ticker, current_czk in current_holdings.items():
            if ticker not in target_amount_map:
                target_amount_map[ticker] = 0.0

        for ticker, current_czk in current_holdings.items():
            target_czk = target_amount_map.get(ticker, 0.0)
            trade_size = abs(target_czk - current_czk)
            is_full_exit = target_czk <= 1.0
            if 0 < trade_size < self.min_trade_czk and not is_full_exit:
                target_amount_map[ticker] = current_czk

        for ticker, target_czk in list(target_amount_map.items()):
            current_czk = current_holdings.get(ticker, 0.0)
            trade_size = abs(target_czk - current_czk)
            if current_czk <= 1.0 and 0 < trade_size < self.min_trade_czk:
                target_amount_map[ticker] = 0.0

        target_vector = np.array([target_amount_map.get(record.ticker, 0.0) for record in best_subset], dtype=float)
        target_sum = float(target_vector.sum())
        if target_sum > 0:
            adjustment_budget = total_capital - target_sum
            if abs(adjustment_budget) >= 1e-6:
                adjustable = [
                    idx for idx, record in enumerate(best_subset)
                    if target_amount_map.get(record.ticker, 0.0) > 0
                    and abs(target_amount_map.get(record.ticker, 0.0) - current_holdings.get(record.ticker, 0.0)) >= self.min_trade_czk
                ]
                if not adjustable:
                    adjustable = [idx for idx, record in enumerate(best_subset) if target_amount_map.get(record.ticker, 0.0) > 0]
                if adjustable:
                    extra_piece = adjustment_budget / len(adjustable)
                    for idx in adjustable:
                        target_vector[idx] = max(0.0, target_vector[idx] + extra_piece)

        best_subset, target_vector = self._enforce_min_position_size_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._enforce_no_sell_tickers_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._enforce_turnover_limit_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._enforce_sell_count_limit_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._enforce_no_sell_tickers_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._enforce_max_position_weight_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._enforce_min_position_size_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._enforce_turnover_limit_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._enforce_sell_count_limit_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._enforce_no_sell_tickers_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._enforce_max_position_weight_on_targets(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )
        best_subset, target_vector = self._trim_targets_to_available_capital(
            best_subset,
            target_vector,
            current_holdings,
            total_capital,
        )

        final_tickers = [record.ticker for record in best_subset]
        final_current = np.array([current_holdings.get(ticker, 0.0) for ticker in final_tickers], dtype=float)
        final_corr = build_correlation_matrix(final_tickers, self.cache, self.log)
        
        final_dropped_czk = sum(amt for t, amt in current_holdings.items() if t not in final_tickers)
        
        final_objective, final_profit_czk, final_risk_czk, _, _ = _objective_from_target_amounts_downside(
            target_vector,
            best_subset,
            final_corr,
            final_current,
            total_capital,
            self.sale_penalty_rate,
            final_dropped_czk,
            self.objective_config,
        )

        best_return = final_profit_czk / max(total_capital, 1e-12)
        best_vol = final_risk_czk / max(total_capital, 1e-12)
        best_objective = final_objective

        trade_lines: list[TradeLine] = []
        removed_positions = 0

        for record, target_czk in sorted(zip(best_subset, target_vector), key=lambda item: item[1], reverse=True):
            current_czk = current_holdings.get(record.ticker, 0.0)
            delta = target_czk - current_czk
            if delta > 1.0:
                action = "BUY"
            elif delta < -1.0:
                action = "SELL"
            else:
                action = "KEEP"
            trade_lines.append(
                TradeLine(
                    ticker=record.ticker,
                    current_czk=current_czk,
                    target_czk=target_czk,
                    delta_czk=delta,
                    action=action,
                    sharpe=record.sharpe,
                    forecast_pct=record.forecast_pct,
                    std_pct=record.std_pct,
                    upside_pct=record.upside_pct,
                    horizon_data=record.horizon_data,
                    score=record.score,
                )
            )

        for ticker, amount in current_holdings.items():
            if ticker in final_tickers:
                continue
            removed_positions += 1
            record = self.record_map[ticker]
            trade_lines.append(
                TradeLine(
                    ticker=ticker,
                    current_czk=amount,
                    target_czk=0.0,
                    delta_czk=-amount,
                    action="SELL_ALL",
                    sharpe=record.sharpe,
                    forecast_pct=record.forecast_pct,
                    std_pct=record.std_pct,
                    upside_pct=record.upside_pct,
                    horizon_data=record.horizon_data,
                    score=record.score,
                )
            )

        sold_czk = sum(max(line.current_czk - line.target_czk, 0.0) for line in trade_lines)
        penalty_czk = sold_czk * self.sale_penalty_rate
        invested_czk = max(0.0, total_capital - penalty_czk)
        final_weights = target_vector / max(total_capital, 1e-12)

        return OptimizationResult(
            records=best_subset,
            weights=final_weights,
            sharpe=best_objective,
            expected_return=best_return,
            volatility=best_vol,
            sold_czk=sold_czk,
            penalty_czk=penalty_czk,
            invested_czk=invested_czk,
            trades=trade_lines,
            target_total_czk=total_capital,
            current_total_czk=current_total,
            extra_cash_czk=self.extra_cash_czk,
            position_changes=removed_positions,
            min_trade_czk=self.min_trade_czk,
        )



import json
import threading

class TickerMetadataFetcher:
    def __init__(self, log_fn):
        self.log_fn = log_fn
        from pathlib import Path
        self.cache_file = Path("portfolios/portfolio_metadata_cache.json")
        self.cache = {}
        self._load()

    def _load(self):
        if self.cache_file.exists():
            try:
                with open(self.cache_file, "r", encoding="utf-8") as f:
                    self.cache = json.load(f)
            except: pass

    def _save(self):
        try:
            self.cache_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.cache_file, "w", encoding="utf-8") as f:
                json.dump(self.cache, f, indent=4)
        except: pass

    @staticmethod
    def _positive_float(value):
        try:
            value = float(value)
        except (TypeError, ValueError):
            return None
        if not math.isfinite(value) or value <= 0:
            return None
        return value

    def _latest_prices(self, yf, tickers: list[str]) -> dict[str, float]:
        prices: dict[str, float] = {}
        if not tickers:
            return prices
        try:
            data = yf.download(
                tickers,
                period="5d",
                interval="1d",
                auto_adjust=False,
                progress=False,
                threads=True,
            )
            if data is None or data.empty:
                return prices
            close = data["Close"] if "Close" in data else data.get("Adj Close")
            if close is None:
                return prices
            if hasattr(close, "columns"):
                for ticker in tickers:
                    if ticker not in close.columns:
                        continue
                    series = close[ticker].dropna()
                    value = self._positive_float(series.iloc[-1] if not series.empty else None)
                    if value:
                        prices[ticker] = value
            else:
                series = close.dropna()
                value = self._positive_float(series.iloc[-1] if not series.empty else None)
                if value:
                    prices[tickers[0]] = value
        except Exception as exc:
            try:
                self.log_fn(f"Nepodařilo se obnovit aktuální ceny: {exc}")
            except Exception:
                pass
        for ticker in tickers:
            if ticker in prices:
                continue
            value = self._latest_price_individual(yf, ticker)
            if value:
                prices[ticker] = value
        return prices

    def _latest_price_individual(self, yf, ticker: str):
        try:
            tkr = yf.Ticker(ticker)
            try:
                fast_info = getattr(tkr, "fast_info", {}) or {}
                value = self._positive_float(fast_info.get("last_price") or fast_info.get("lastPrice"))
                if value:
                    return value
            except Exception:
                pass
            try:
                hist = tkr.history(period="5d", interval="1d", auto_adjust=False)
                if hist is not None and not hist.empty and "Close" in hist:
                    series = hist["Close"].dropna()
                    value = self._positive_float(series.iloc[-1] if not series.empty else None)
                    if value:
                        return value
            except Exception:
                pass
            try:
                info = tkr.info
                return self._positive_float(info.get("currentPrice", info.get("regularMarketPrice")))
            except Exception:
                return None
        except Exception:
            return None

    def _fetch_info(self, yf, ticker: str, fallback_price=None) -> dict:
        try:
            tkr = yf.Ticker(ticker)
            info = tkr.info
            name = info.get("shortName", info.get("longName", ticker))
            sector = info.get("sector", "N/A")
            price = self._positive_float(
                info.get("currentPrice", info.get("regularMarketPrice"))
            ) or self._positive_float(fallback_price) or 0.0
            mcap = info.get("marketCap", 0)
            if mcap > 1e9: mcap_str = f"{mcap/1e9:.1f}B"
            elif mcap > 1e6: mcap_str = f"{mcap/1e6:.1f}M"
            else: mcap_str = str(mcap)
            return {"name": name, "sector": sector, "price": price, "mcap": mcap_str}
        except Exception:
            return {"name": ticker, "sector": "N/A", "price": self._positive_float(fallback_price) or 0.0, "mcap": "0"}

    def fetch_async(self, tickers: list[str], callback):
        def worker():
            yf = _load_yfinance()
            if not yf:
                callback({})
                return
            updated = False
            results = {}
            normalized = []
            seen = set()
            for raw in tickers:
                t = normalize_ticker_symbol(str(raw))
                if t and t not in seen:
                    normalized.append(t)
                    seen.add(t)
            latest_prices = self._latest_prices(yf, normalized)
            for t in normalized:
                cached = dict(self.cache.get(t, {}))
                previous_price = self._positive_float(cached.get("price"))
                data = cached if cached else None
                if data is None:
                    data = self._fetch_info(yf, t, fallback_price=latest_prices.get(t))

                fresh_price = latest_prices.get(t)
                if fresh_price:
                    data["price"] = fresh_price
                    if data.get("name") in {t, "", None} or data.get("sector") in {"N/A", "", None}:
                        data = self._fetch_info(yf, t, fallback_price=fresh_price)
                elif not data.get("price"):
                    data["price"] = previous_price or 0.0

                cache_data = {
                    "name": data.get("name", t),
                    "sector": data.get("sector", "N/A"),
                    "price": data.get("price", 0.0),
                    "mcap": data.get("mcap", "0"),
                }
                if self.cache.get(t) != cache_data:
                    self.cache[t] = cache_data
                    updated = True

                result_data = dict(cache_data)
                if previous_price:
                    result_data["_previous_price"] = previous_price
                result_data["_price_refreshed"] = bool(fresh_price)
                results[t] = result_data
            if updated: self._save()
            callback(results)
        threading.Thread(target=worker, daemon=True).start()


class Trading212ApiError(Exception):
    pass


class Trading212Client:
    BASE_URLS = {
        "live": "https://live.trading212.com/api/v0",
        "demo": "https://demo.trading212.com/api/v0",
        "paper": "https://demo.trading212.com/api/v0",
    }

    def __init__(self, environment: str, api_key: str, api_secret: str):
        env = (environment or "live").strip().lower()
        if env in {"demo", "paper trading", "paper_trading", "demo/paper", "paper/demo"}:
            env = "paper"
        if env not in self.BASE_URLS:
            env = "live"
        self.environment = env
        self.base_url = self.BASE_URLS[env]
        self.api_key = api_key.strip()
        self.api_secret = api_secret.strip()
        if not self.api_key or not self.api_secret:
            raise Trading212ApiError("Chybí API key nebo API secret.")

    def _headers(self) -> dict[str, str]:
        token = base64.b64encode(f"{self.api_key}:{self.api_secret}".encode("utf-8")).decode("ascii")
        return {
            "Authorization": f"Basic {token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
            "User-Agent": "portfolio-manager-ui/1.0",
        }

    def _url_for_path(self, path: str, query: dict | None = None) -> str:
        if path.startswith("http://") or path.startswith("https://"):
            url = path
        else:
            if path.startswith("/api/v0/"):
                path = path[len("/api/v0"):]
            if not path.startswith("/"):
                path = "/" + path
            url = self.base_url + path
        if query:
            separator = "&" if "?" in url else "?"
            url += separator + urllib.parse.urlencode(query)
        return url

    def get(self, path: str, query: dict | None = None, retries: int = 2, retry_wait_seconds: float = 2.5):
        url = self._url_for_path(path, query=query)
        last_error = None
        for attempt in range(retries + 1):
            request = urllib.request.Request(url, headers=self._headers(), method="GET")
            try:
                with urllib.request.urlopen(request, timeout=30) as response:
                    raw = response.read().decode("utf-8")
                    return json.loads(raw) if raw else None
            except urllib.error.HTTPError as exc:
                body = ""
                try:
                    body = exc.read().decode("utf-8", errors="replace")
                except Exception:
                    pass
                detail = body.strip() or exc.reason
                if exc.code in {408, 429} and attempt < retries:
                    reset = exc.headers.get("x-ratelimit-reset") if exc.headers else None
                    wait_seconds = retry_wait_seconds
                    try:
                        if reset:
                            wait_seconds = max(1.0, min(60.0, float(reset) - time.time() + 0.5))
                    except Exception:
                        pass
                    time.sleep(wait_seconds)
                    continue
                if exc.code == 401:
                    detail = (
                        f"{detail} | Ověř, že API key patří ke zvolenému prostředí "
                        f"({self.environment}). Live a paper/demo klíče nejsou zaměnitelné."
                    )
                last_error = Trading212ApiError(f"Trading 212 API vrátilo HTTP {exc.code}: {detail}")
                break
            except urllib.error.URLError as exc:
                last_error = Trading212ApiError(f"Nepodařilo se spojit s Trading 212 API: {exc.reason}")
                break
            except TimeoutError as exc:
                last_error = Trading212ApiError("Trading 212 API neodpovědělo v časovém limitu.")
                break
        raise last_error or Trading212ApiError("Trading 212 API volání selhalo.")

    def post(self, path: str, payload: dict | None = None, retries: int = 1, retry_wait_seconds: float = 2.5):
        url = self._url_for_path(path)
        body = json.dumps(payload or {}).encode("utf-8")
        last_error = None
        for attempt in range(retries + 1):
            request = urllib.request.Request(url, data=body, headers=self._headers(), method="POST")
            try:
                with urllib.request.urlopen(request, timeout=30) as response:
                    raw = response.read().decode("utf-8")
                    return json.loads(raw) if raw else None
            except urllib.error.HTTPError as exc:
                detail = ""
                try:
                    detail = exc.read().decode("utf-8", errors="replace").strip()
                except Exception:
                    pass
                if exc.code == 429 and attempt < retries:
                    wait_seconds = retry_wait_seconds
                    reset = exc.headers.get("x-ratelimit-reset") if exc.headers else None
                    try:
                        if reset:
                            wait_seconds = max(1.0, min(60.0, float(reset) - time.time() + 0.5))
                    except Exception:
                        pass
                    time.sleep(wait_seconds)
                    continue
                if exc.code == 401:
                    detail = (
                        f"{detail or exc.reason} | Ověř, že API key patří ke zvolenému prostředí "
                        f"({self.environment}). Live a paper/demo klíče nejsou zaměnitelné."
                    )
                last_error = Trading212ApiError(f"Trading 212 API vrátilo HTTP {exc.code}: {detail or exc.reason}")
                break
            except urllib.error.URLError as exc:
                last_error = Trading212ApiError(f"Nepodařilo se spojit s Trading 212 API: {exc.reason}")
                break
            except TimeoutError:
                last_error = Trading212ApiError("Trading 212 API neodpovědělo v časovém limitu.")
                break
        raise last_error or Trading212ApiError("Trading 212 API POST volání selhalo.")

    def get_all(self, path: str, query: dict | None = None) -> list[dict]:
        request_query = dict(query or {})
        request_query.setdefault("limit", 50)
        items: list[dict] = []
        data = self.get(path, query=request_query)
        while True:
            items.extend(self._as_list(data))
            next_page = data.get("nextPagePath") if isinstance(data, dict) else None
            if not next_page:
                break
            data = self.get(str(next_page), query=None)
        return items

    @staticmethod
    def _as_list(data) -> list[dict]:
        if isinstance(data, list):
            return [item for item in data if isinstance(item, dict)]
        if isinstance(data, dict):
            for key in ("items", "data", "results", "content", "pies", "positions"):
                value = data.get(key)
                if isinstance(value, list):
                    return [item for item in value if isinstance(item, dict)]
            return [data]
        return []

    @staticmethod
    def _pie_id(pie: dict):
        for key in ("id", "pieId", "pie_id"):
            value = pie.get(key)
            if value is not None:
                return value
        settings = pie.get("settings") if isinstance(pie.get("settings"), dict) else {}
        for key in ("id", "pieId", "pie_id"):
            value = settings.get(key)
            if value is not None:
                return value
        return None

    def account_summary(self) -> dict:
        try:
            summary = self.get("/equity/account/summary")
            return summary if isinstance(summary, dict) else {}
        except Trading212ApiError:
            cash = self.get("/equity/account/cash")
            info = self.get("/equity/account/info")
            result = {}
            if isinstance(info, dict):
                result.update(info)
            if isinstance(cash, dict):
                result["cash"] = cash
            return result

    def positions(self) -> list[dict]:
        try:
            data = self.get("/equity/positions")
        except Trading212ApiError:
            data = self.get("/equity/portfolio")
        return self._as_list(data)

    def pies(self) -> list[dict]:
        try:
            return self.get_all("/equity/pies")
        except Trading212ApiError:
            data = self.get("/equity/pies")
            return self._as_list(data)

    def pie_detail(self, pie_id, retries: int = 6, retry_wait_seconds: float = 6.0) -> dict:
        data = self.get(f"/equity/pies/{pie_id}", retries=retries, retry_wait_seconds=retry_wait_seconds)
        return data if isinstance(data, dict) else {}

    def instruments(self) -> list[dict]:
        try:
            return self.get_all("/equity/metadata/instruments")
        except Trading212ApiError:
            data = self.get("/equity/metadata/instruments")
            return self._as_list(data)

    def transactions(self, since_iso: str | None = None) -> list[dict]:
        query = {"limit": 50}
        if since_iso:
            query["time"] = since_iso
        return self.get_all("/equity/history/transactions", query=query)

    def update_pie(self, pie_id, payload: dict):
        return self.post(f"/equity/pies/{pie_id}", payload, retries=2, retry_wait_seconds=4.0)

    def snapshot(self) -> dict:
        summary = self.account_summary()
        positions = self.positions()
        pie_summaries = []
        pie_details = []
        pie_errors = []
        try:
            pie_summaries = self.pies()
            for pie in pie_summaries:
                pie_id = self._pie_id(pie)
                if pie_id is None:
                    continue
                try:
                    if pie_details:
                        time.sleep(6.0)
                    detail = self.pie_detail(pie_id)
                    if detail:
                        detail.setdefault("_summary", pie)
                        pie_details.append(detail)
                except Trading212ApiError as exc:
                    pie_errors.append({"id": pie_id, "error": str(exc), "summary": pie})
                    continue
        except Trading212ApiError:
            pie_summaries = []
        return {
            "summary": summary,
            "positions": positions,
            "pies": pie_summaries,
            "pie_details": pie_details,
            "pie_errors": pie_errors,
        }


class PortfolioManagerApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Portfolio Manager")
        self.root.geometry("1420x940")
        self.root.minsize(1320, 880)

        self.base_dir = application_base_dir()
        prediction_source_config = self._load_prediction_source_config()
        self.records: list[StockRecord] = []
        self.fallback_prediction_cache: dict[str, dict] = {}
        self.fallback_prediction_inflight: set[str] = set()
        self.fallback_uncertainty_by_ticker: dict[str, float] = {}
        self.fallback_prediction_failures: dict[str, int] = {}
        self.fallback_max_retries = 3
        self.meta_path: Path | None = None
        self.worker_thread: threading.Thread | None = None
        self.stop_event: threading.Event | None = None
        self.ui_queue: Queue = Queue()
        self.is_loading_meta = False
        self.last_chart_data: dict | None = None
        self.loaded_historical_projections: dict | None = None
        self.is_optimizing = False
        self.last_result = None
        self._portfolio_refresh_seq = 0
        self._portfolio_state_generation = 0
        self._stock_table_refresh_seq = 0
        self._busy_active = False
        self._closing = False
        self._portfolio_action_widgets = []
        self._settings_entries = []

        main = ttk.Frame(root, padding=12)
        main.pack(fill=tk.BOTH, expand=True)

        self.meta_label_var = tk.StringVar(value="EMA predikce zatím nenačteny.")
        meta_row = ttk.Frame(main)
        meta_row.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(meta_row, textvariable=self.meta_label_var).pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.start_button = ttk.Button(meta_row, text="Optimalizovat portfolio", command=self.start_optimization)
        self.start_button.pack(side=tk.RIGHT)
        self.stop_button = ttk.Button(meta_row, text="Zastavit", command=self.request_stop, state=tk.DISABLED)
        self.stop_button.pack(side=tk.RIGHT, padx=(0, 8))

        self.status_var = tk.StringVar(value="UI připraveno. Načítám data...")
        ttk.Label(main, textvariable=self.status_var).pack(fill=tk.X, pady=(0, 6))
        self.progress = ttk.Progressbar(main, mode="indeterminate")
        self.progress.pack(fill=tk.X, pady=(0, 10))

        input_frame = ttk.Frame(main)
        input_frame.pack(fill=tk.X, pady=(0, 10))

        left = ttk.LabelFrame(input_frame, text="Aktuální portfolio", padding=10)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))

        # Add Ticker input
        add_frame = ttk.Frame(left)
        add_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(add_frame, text="Ticker:").pack(side=tk.LEFT)
        self.add_ticker_var = tk.StringVar()
        self.add_ticker_entry = ttk.Entry(add_frame, textvariable=self.add_ticker_var, width=8)
        self.add_ticker_entry.pack(side=tk.LEFT, padx=5)
        ttk.Label(add_frame, text="CZK:").pack(side=tk.LEFT)
        self.add_amount_var = tk.StringVar()
        self.add_amount_entry = ttk.Entry(add_frame, textvariable=self.add_amount_var, width=10)
        self.add_amount_entry.pack(side=tk.LEFT, padx=5)
        self.add_holding_button = ttk.Button(add_frame, text="Přidat/Upravit", command=self._add_holding)
        self.add_holding_button.pack(side=tk.LEFT, padx=2)
        self.remove_holding_button = ttk.Button(add_frame, text="Odebrat", command=self._remove_holding)
        self.remove_holding_button.pack(side=tk.LEFT, padx=2)
        self.no_sell_button = ttk.Button(add_frame, text="Neprodávat", command=self._toggle_selected_no_sell)
        self.no_sell_button.pack(side=tk.LEFT, padx=2)
        self._portfolio_action_widgets.extend([
            self.add_ticker_entry,
            self.add_amount_entry,
            self.add_holding_button,
            self.remove_holding_button,
            self.no_sell_button,
        ])

        # Treeview
        port_tree_frame = ttk.Frame(left)
        port_tree_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 6))
        self.port_tree = ttk.Treeview(
            port_tree_frame,
            columns=("ticker", "name", "sector", "price", "mcap", "czk", "weight", "score", "pred", "q10", "q90", "nosell", "source"),
            show="headings",
            height=10,
        )
        p_headings = {
            "ticker": "Ticker",
            "name": "Název",
            "sector": "Sektor",
            "price": "Cena USD",
            "mcap": "Market Cap",
            "czk": "Hodnota CZK",
            "weight": "Podíl %",
            "score": "Skóre",
            "pred": "Predikce 52w %",
            "q10": "Q10 52w %",
            "q90": "Q90 52w %",
            "nosell": "Neprodat",
            "source": "Zdroj",
        }
        col_widths = {
            "ticker": 70, "name": 145, "sector": 90, "price": 75, "mcap": 85,
            "czk": 100, "weight": 75, "score": 70, "pred": 75, "q10": 70, "q90": 70, "nosell": 70, "source": 95
        }
        for col, lab in p_headings.items():
            self.port_tree.heading(col, text=lab)
            self.port_tree.column(col, width=col_widths.get(col, 80), anchor=tk.CENTER)
        self.port_tree.tag_configure("no_sell", background="#fff7d1")
        self.port_tree_scrollbar = ttk.Scrollbar(port_tree_frame, orient=tk.VERTICAL, command=self.port_tree.yview)
        self.port_tree.configure(yscrollcommand=self.port_tree_scrollbar.set)
        self.port_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.port_tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.port_tree.bind("<<TreeviewSelect>>", self._on_portfolio_row_selected)
        self.port_tree.bind("<Button-1>", self._on_portfolio_tree_click)

        self.portfolio_dict = {}
        self.no_sell_tickers = set()
        self.portfolio_history = []
        self.portfolio_projections = []
        self.portfolio_price_snapshot = {}
        self.portfolio_broker_positions = {}
        self.portfolio_broker_link = {}
        self.portfolio_currency = "CZK"
        self.total_dividends = 0.0
        self.current_portfolio_path = None
        self.t212_config_path = self.base_dir / "portfolios" / "trading212_config.json"
        self.fetcher = TickerMetadataFetcher(self.log)
        self.return_cache = ReturnCache()

        holdings_btn_frame = ttk.Frame(left)
        holdings_btn_frame.pack(fill=tk.X, pady=(6, 0))
        self.new_portfolio_button = ttk.Button(holdings_btn_frame, text="Nové", command=self._new_portfolio)
        self.new_portfolio_button.pack(side=tk.LEFT, padx=(0, 4))
        self.save_portfolio_button = ttk.Button(holdings_btn_frame, text="💾 Uložit", command=self._save_portfolio)
        self.save_portfolio_button.pack(side=tk.LEFT, padx=(0, 4))
        self.load_portfolio_button = ttk.Button(holdings_btn_frame, text="📂 Načíst", command=self._load_portfolio_from_file)
        self.load_portfolio_button.pack(side=tk.LEFT, padx=(0, 4))
        self.trading212_button = ttk.Button(holdings_btn_frame, text="Trading 212", command=self._open_trading212_dialog)
        self.trading212_button.pack(side=tk.LEFT, padx=(0, 4))
        self.degiro_button = ttk.Button(holdings_btn_frame, text="Degiro", command=self._open_degiro_dialog)
        self.degiro_button.pack(side=tk.LEFT, padx=(0, 4))
        self.show_graph_button = ttk.Button(holdings_btn_frame, text="📈 Graf s historií", command=self._on_show_graph_btn_click)
        self.show_graph_button.pack(side=tk.LEFT, padx=(0, 4))

        self._portfolio_action_widgets.extend([
            self.new_portfolio_button,
            self.save_portfolio_button,
            self.load_portfolio_button,
            self.trading212_button,
            self.degiro_button,
            self.show_graph_button,
        ])
        
        self.apply_btn = ttk.Button(holdings_btn_frame, text="Otevřít optimalizované portfolio", command=self._open_apply_trades_dialog, state=tk.DISABLED)
        self.apply_btn.pack(side=tk.LEFT)
        
        # Add Dividends display
        self.div_var = tk.StringVar(value="Statistika | Celkové obdržené dividendy: 0 CZK")
        ttk.Label(left, textvariable=self.div_var, font=("Arial", 9, "bold")).pack(anchor=tk.W, pady=(5, 0))

        right = ttk.LabelFrame(input_frame, text="Nastavení", padding=10)
        right.pack(side=tk.LEFT, fill=tk.Y)

        self.extra_cash_var = tk.StringVar(value="0")
        self.target_positions_var = tk.IntVar(value=50)
        self.max_turnover_var = tk.StringVar(value="")
        self.max_swaps_var = tk.StringVar(value="")
        self.max_weight_var = tk.StringVar(value="8")
        self.sale_penalty_var = tk.DoubleVar(value=0.02)
        self.min_trade_var = tk.StringVar(value="25")
        self.min_position_var = tk.StringVar(value="")
        self.risk_exponent_var = tk.DoubleVar(value=DEFAULT_RISK_EXPONENT)
        self.remote_predictions_enabled_var = tk.BooleanVar(
            value=bool(prediction_source_config.get("enabled", DEFAULT_REMOTE_PREDICTIONS_ENABLED))
        )
        self.remote_predictions_manifest_url_var = tk.StringVar(
            value=str(prediction_source_config.get("manifest_url") or DEFAULT_REMOTE_PREDICTION_MANIFEST_URL)
        )
        
        self.alpha_1w_var = tk.DoubleVar(value=DEFAULT_EMA_ALPHAS["1w"])
        self.alpha_4w_var = tk.DoubleVar(value=DEFAULT_EMA_ALPHAS["4w"])
        self.alpha_13w_var = tk.DoubleVar(value=DEFAULT_EMA_ALPHAS["13w"])
        self.alpha_26w_var = tk.DoubleVar(value=DEFAULT_EMA_ALPHAS["26w"])
        self.alpha_52w_var = tk.DoubleVar(value=DEFAULT_EMA_ALPHAS["52w"])
        self.fallback_uncertainty_var = tk.DoubleVar(value=DEFAULT_FALLBACK_UNCERTAINTY)

        self._add_labeled_entry(right, "Dodatečný kapitál (CZK)", self.extra_cash_var)
        self._add_labeled_entry(right, "Cílový počet akcií", self.target_positions_var)
        self._add_labeled_entry(right, "Max protočit kapitál (CZK)", self.max_turnover_var)
        self._add_labeled_entry(right, "Max počet prodejů", self.max_swaps_var)
        self._add_labeled_entry(right, "Max váha akcie (%)", self.max_weight_var)
        self._add_labeled_entry(right, "Postih za prodej", self.sale_penalty_var)
        self._add_labeled_entry(right, "Minimální obchod (CZK)", self.min_trade_var)
        self._add_labeled_entry(right, "Min velikost pozice (CZK)", self.min_position_var)
        self._add_labeled_entry(right, "Exponent downside rizika (0-1)", self.risk_exponent_var)

        self.remote_predictions_check = ttk.Checkbutton(
            right,
            text="Online beta predikce",
            variable=self.remote_predictions_enabled_var,
            command=self._save_prediction_source_config,
        )
        self.remote_predictions_check.pack(anchor=tk.W, pady=(10, 0))
        self._settings_entries.append(self.remote_predictions_check)
        ttk.Label(right, text="Manifest URL").pack(anchor=tk.W, pady=(4, 0))
        self.remote_manifest_entry = ttk.Entry(
            right,
            textvariable=self.remote_predictions_manifest_url_var,
            width=24,
        )
        self.remote_manifest_entry.pack(anchor=tk.W, fill=tk.X)
        self.remote_manifest_entry.bind("<FocusOut>", lambda _event: self._save_prediction_source_config())
        self.remote_manifest_entry.bind("<Return>", lambda _event: self._save_prediction_source_config())
        self._settings_entries.append(self.remote_manifest_entry)

        data_actions = ttk.Frame(right)
        data_actions.pack(fill=tk.X, pady=(10, 0))
        self.reload_button = ttk.Button(data_actions, text="Obnovit EMA", command=self.reload_meta_async)
        self.reload_button.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
        self.load_ai_button = ttk.Button(data_actions, text="Import AI CSV", command=self._load_ai_csv)
        self.load_ai_button.pack(side=tk.LEFT, fill=tk.X, expand=True)

        table_frame = ttk.LabelFrame(main, text="Načtený vesmír akcií", padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        universe_tree_frame = ttk.Frame(table_frame)
        universe_tree_frame.pack(fill=tk.BOTH, expand=True)
        universe_columns = ("ticker", "name", "sector", "price", "mcap", "score", "pred", "q10", "q90", "source")
        self.tree = ttk.Treeview(universe_tree_frame, columns=universe_columns, show="headings", height=10)
        headings = {
            "ticker": "Ticker",
            "name": "Název",
            "sector": "Sektor",
            "price": "Cena USD",
            "mcap": "Market Cap",
            "score": "Skóre",
            "pred": "Predikce 52w %",
            "q10": "Q10 52w %",
            "q90": "Q90 52w %",
            "source": "Zdroj",
        }
        universe_col_widths = {
            "ticker": 70,
            "name": 180,
            "sector": 110,
            "price": 75,
            "mcap": 85,
            "score": 70,
            "pred": 95,
            "q10": 85,
            "q90": 85,
            "source": 90,
        }
        for column, label in headings.items():
            self.tree.heading(column, text=label)
            self.tree.column(column, width=universe_col_widths.get(column, 90), anchor=tk.CENTER)
        self.stock_tree_scrollbar = ttk.Scrollbar(universe_tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.stock_tree_scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.stock_tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.bind("<<TreeviewSelect>>", self._on_stock_universe_row_selected)

        log_frame = ttk.Frame(main)
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.log_box = tk.Text(log_frame, height=16, state=tk.DISABLED)
        self.log_scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_box.yview)
        self.log_box.configure(yscrollcommand=self.log_scrollbar.set)
        self.log_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self._update_action_states()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.after(100, self._drain_ui_queue)
        self.root.after(150, self.reload_meta_async)

    def _add_labeled_entry(self, parent: ttk.Widget, label: str, variable):
        ttk.Label(parent, text=label).pack(anchor=tk.W, pady=(4, 0))
        entry = ttk.Entry(parent, textvariable=variable, width=24)
        entry.pack(anchor=tk.W, fill=tk.X)
        if hasattr(self, "_settings_entries"):
            self._settings_entries.append(entry)
        return entry

    def _prediction_source_config_path(self) -> Path:
        return self.base_dir / "portfolios" / "prediction_source_config.json"

    def _load_prediction_source_config(self) -> dict:
        path = self._prediction_source_config_path()
        if not path.exists():
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _save_prediction_source_config(self):
        try:
            data = {
                "enabled": bool(self.remote_predictions_enabled_var.get()),
                "manifest_url": str(self.remote_predictions_manifest_url_var.get() or "").strip(),
            }
            path = self._prediction_source_config_path()
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as exc:
            self.log(f"Nastavení online predikcí se nepodařilo uložit: {exc}")

    def _remote_prediction_cache_dir(self) -> Path:
        return self.base_dir / "cache" / "remote_predictions"

    def _download_url_bytes(self, url: str, timeout: float = 12.0) -> bytes:
        request = urllib.request.Request(
            url,
            headers={"User-Agent": "PortfolioManager/1.0 (+remote-predictions)"},
        )
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return response.read()

    @staticmethod
    def _sha256_bytes(data: bytes) -> str:
        return hashlib.sha256(data).hexdigest()

    @staticmethod
    def _sha256_file(path: Path) -> str:
        digest = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _parse_remote_prediction_date(value, fallback_name: str = "") -> dt.date | None:
        if value:
            text = str(value).strip()
            for fmt in ("%Y-%m-%d", "%Y%m%d"):
                try:
                    return dt.datetime.strptime(text, fmt).date()
                except ValueError:
                    pass
        match = re.search(r"predictions_(\d{4})(\d{2})(\d{2})_", str(fallback_name or ""))
        if match:
            try:
                return dt.date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            except ValueError:
                return None
        return None

    def _remote_prediction_target_name(self, prediction_date: dt.date, sha256_hex: str) -> str:
        return f"predictions_{prediction_date.strftime('%Y%m%d')}_remote_{sha256_hex[:8]}.csv"

    def _sync_remote_predictions_if_enabled(self, config: dict | None = None) -> str:
        if config is None:
            if not getattr(self, "remote_predictions_enabled_var", None):
                return ""
            self._save_prediction_source_config()
            config = {
                "enabled": bool(self.remote_predictions_enabled_var.get()),
                "manifest_url": str(self.remote_predictions_manifest_url_var.get() or "").strip(),
            }

        if not bool(config.get("enabled", False)):
            return ""

        manifest_url = str(config.get("manifest_url") or "").strip()
        if not manifest_url:
            message = "Online beta predikce jsou zapnuté, ale chybí Manifest URL."
            self.log(message)
            return message

        try:
            self.log("Kontroluji online beta predikce...")
            manifest_raw = self._download_url_bytes(manifest_url, timeout=10.0)
            manifest = json.loads(manifest_raw.decode("utf-8-sig"))
            if not isinstance(manifest, dict):
                raise ValueError("manifest není JSON objekt")

            schema_version = int(manifest.get("schema_version", 1))
            if schema_version != 1:
                raise ValueError(f"nepodporovaná verze manifestu {schema_version}")

            file_name = str(manifest.get("source_name") or manifest.get("file") or "").strip()
            file_url = str(manifest.get("url") or manifest.get("file_url") or "").strip()
            parts = manifest.get("parts") if isinstance(manifest.get("parts"), list) else []
            if not file_url:
                if not file_name and not parts:
                    raise ValueError("manifest neobsahuje položku file ani url")
                if not parts:
                    file_url = urllib.parse.urljoin(manifest_url, file_name)
            if not file_name:
                file_name = Path(urllib.parse.urlparse(file_url).path).name

            expected_sha = str(manifest.get("sha256") or "").strip().lower()
            if not re.fullmatch(r"[0-9a-f]{64}", expected_sha):
                raise ValueError("manifest neobsahuje platný sha256")

            prediction_date = self._parse_remote_prediction_date(
                manifest.get("prediction_date"),
                fallback_name=file_name,
            )
            if prediction_date is None:
                raise ValueError("manifest neobsahuje platné prediction_date")
            if prediction_date < EMA_MODEL_PREDICTION_START_DATE:
                raise ValueError(
                    f"prediction_date {prediction_date.isoformat()} je před podporovaným startem "
                    f"{EMA_MODEL_PREDICTION_START_DATE.isoformat()}"
                )

            target_name = self._remote_prediction_target_name(prediction_date, expected_sha)
            cache_dir = self._remote_prediction_cache_dir()
            cache_dir.mkdir(parents=True, exist_ok=True)
            cache_path = cache_dir / target_name
            local_path = self.base_dir / target_name

            if cache_path.exists() and self._sha256_file(cache_path) != expected_sha:
                try:
                    cache_path.unlink()
                except Exception:
                    pass
            if not cache_path.exists():
                if parts:
                    chunks = []
                    for index, part in enumerate(parts, start=1):
                        if not isinstance(part, dict):
                            raise ValueError(f"část {index} v manifestu není objekt")
                        part_file = str(part.get("file") or "").strip()
                        part_url = str(part.get("url") or "").strip()
                        if not part_url:
                            if not part_file:
                                raise ValueError(f"část {index} nemá file ani url")
                            part_url = urllib.parse.urljoin(manifest_url, part_file)
                        chunk = self._download_url_bytes(part_url, timeout=20.0)
                        part_sha = str(part.get("sha256") or "").strip().lower()
                        if part_sha and self._sha256_bytes(chunk) != part_sha:
                            raise ValueError(f"sha256 nesedí u části {index}: {part_file}")
                        chunks.append(chunk)
                    payload = b"".join(chunks)
                else:
                    payload = self._download_url_bytes(file_url, timeout=20.0)

                payload_encoding = str(manifest.get("encoding") or "").strip().lower()
                if payload_encoding == "base64":
                    payload = base64.b64decode(payload, validate=True)
                elif payload_encoding:
                    raise ValueError(f"nepodporované encoding={payload_encoding}")

                payload_compression = str(manifest.get("compression") or "").strip().lower()
                if payload_compression == "gzip":
                    payload = gzip.decompress(payload)
                elif payload_compression:
                    raise ValueError(f"nepodporované compression={payload_compression}")

                actual_sha = self._sha256_bytes(payload)
                if actual_sha != expected_sha:
                    raise ValueError(
                        f"sha256 nesedí: manifest {expected_sha[:12]}..., staženo {actual_sha[:12]}..."
                    )
                temp_path = cache_path.with_suffix(cache_path.suffix + ".tmp")
                with open(temp_path, "wb") as f:
                    f.write(payload)
                os.replace(temp_path, cache_path)

            if local_path.exists() and self._sha256_file(local_path) != expected_sha:
                try:
                    local_path.unlink()
                except Exception:
                    pass
            if not local_path.exists():
                shutil.copy2(cache_path, local_path)
            for stale_path in self.base_dir.glob(f"predictions_{prediction_date.strftime('%Y%m%d')}_remote_*.csv"):
                if stale_path.resolve() == local_path.resolve():
                    continue
                try:
                    stale_path.unlink()
                except Exception:
                    pass

            manifest_copy = dict(manifest)
            manifest_copy.update({
                "cached_file": target_name,
                "cached_at": dt.datetime.now().isoformat(timespec="seconds"),
                "manifest_url": manifest_url,
            })
            with open(cache_dir / "last_manifest.json", "w", encoding="utf-8") as f:
                json.dump(manifest_copy, f, indent=2, ensure_ascii=False)

            summary = (
                f"Online beta predikce aktuální: {target_name} "
                f"({prediction_date.strftime('%d.%m.%Y')})."
            )
            self.log(summary)
            return summary
        except Exception as exc:
            message = f"Online beta predikce se nepodařilo aktualizovat: {exc}. Používám lokální predikce."
            self.log(message)
            return message

    def _current_objective_config(self) -> ObjectiveConfig:
        risk_exponent = float(self.risk_exponent_var.get())
        if risk_exponent < 0:
            raise ValueError("Exponent downside rizika nemůže být záporný.")
        return ObjectiveConfig(risk_exponent=risk_exponent)

    def _apply_record_scores(self, records: list[StockRecord], config: ObjectiveConfig | None = None) -> ObjectiveConfig:
        config = config or self._current_objective_config()
        for record in records:
            record.score = _record_objective_score(record, config)
        records.sort(key=lambda record: record.score, reverse=True)
        return config

    def _refresh_stock_table(self):
        if not hasattr(self, "tree"):
            return
        self._stock_table_refresh_seq += 1
        refresh_seq = self._stock_table_refresh_seq
        records = list(self.records[:150])

        def update_rows(metadata=None):
            if refresh_seq != self._stock_table_refresh_seq:
                return
            source_metadata = dict(getattr(self.fetcher, "cache", {}) or {})
            source_metadata.update(metadata or {})
            normalized_metadata = {
                normalize_ticker_symbol(k): v
                for k, v in source_metadata.items()
                if normalize_ticker_symbol(k)
            }
            self.tree.delete(*self.tree.get_children())
            for record in records:
                ticker = normalize_ticker_symbol(record.ticker)
                meta = normalized_metadata.get(ticker, {"name": ticker, "sector": "N/A", "price": 0.0, "mcap": "0"})
                pred, q10, q90 = _record_like_horizon_values(record)
                source = "AI (starší)" if record.is_synthetic else "AI"
                price = _safe_float(meta.get("price"), 0.0)
                self.tree.insert(
                    "",
                    tk.END,
                    values=(
                        ticker,
                        meta.get("name", ticker),
                        meta.get("sector", "N/A"),
                        f"{price:.2f}" if price > 0 else "-",
                        meta.get("mcap", "0"),
                        _format_score(record.score),
                        _format_return_pct(pred),
                        _format_return_pct(q10),
                        _format_return_pct(q90),
                        source,
                    ),
                )

        update_rows(getattr(self.fetcher, "cache", {}))
        if records:
            self.fetcher.fetch_async(
                [record.ticker for record in records],
                lambda m: self.ui_queue.put(("call", update_rows, (m,))),
            )

    def _on_stock_universe_row_selected(self, _event=None):
        selected = self.tree.selection()
        if not selected:
            return
        values = self.tree.item(selected[0], "values")
        if not values:
            return
        columns = list(self.tree["columns"])
        try:
            ticker_index = columns.index("ticker")
        except ValueError:
            ticker_index = 0
        if ticker_index >= len(values):
            return
        ticker = normalize_ticker_symbol(str(values[ticker_index]))
        if not ticker:
            return
        self.add_ticker_var.set(ticker)
        try:
            self.add_amount_entry.focus_set()
        except Exception:
            pass

    def _append_log(self, message: str):
        self.log_box.configure(state=tk.NORMAL)
        timestamp = dt.datetime.now().strftime("%H:%M:%S")
        self.log_box.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_box.see(tk.END)
        self.log_box.configure(state=tk.DISABLED)
        
        # Update the status bar for key optimization milestones so they are immediately visible
        strip_msg = message.strip()
        if (
            "Optimalizační krok" in strip_msg
            or "Krok " in strip_msg
            or "Konsolidace" in strip_msg
            or "Spouštím Resampled" in strip_msg
            or "Optimalizace hotova" in strip_msg
            or "Optimalizace selhala" in strip_msg
        ):
            display_msg = strip_msg.strip("= ")
            self.status_var.set(display_msg)

    def log(self, message: str):
        if getattr(self, "_closing", False):
            return
        if threading.current_thread() is threading.main_thread():
            self._append_log(message)
        else:
            self.ui_queue.put(("log", message))

    def _on_close(self):
        self._closing = True
        if self.stop_event is not None:
            self.stop_event.set()
        try:
            self.root.destroy()
        except tk.TclError:
            pass

    def _is_ui_locked(self) -> bool:
        return bool(self._busy_active or self.is_loading_meta or self.is_optimizing)

    def _blocked_action(self, action: str = "Akce") -> bool:
        if not self._is_ui_locked():
            return False
        self.status_var.set(f"{action} je dočasně zamčená, probíhá zpracování dat.")
        return True

    def _reset_portfolio_session_state(self, clear_portfolio_data: bool = False):
        self._portfolio_state_generation += 1
        self._portfolio_refresh_seq += 1
        self.last_result = None
        self.last_chart_data = None
        self.loaded_historical_projections = None
        self.fallback_prediction_cache.clear()
        self.fallback_prediction_inflight.clear()
        self.fallback_uncertainty_by_ticker.clear()
        self.fallback_prediction_failures.clear()
        self.stop_event = None
        self.worker_thread = None

        if clear_portfolio_data:
            self.portfolio_dict = {}
            self.no_sell_tickers = set()
            self.portfolio_history = []
            self.portfolio_projections = []
            self.portfolio_price_snapshot = {}
            self.portfolio_broker_positions = {}
            self.portfolio_broker_link = {}
            self.portfolio_currency = "CZK"
            self.total_dividends = 0.0
            self.current_portfolio_path = None
            self.extra_cash_var.set("0")
            self.div_var.set("Statistika | Celkové obdržené dividendy: 0 CZK")

        if hasattr(self, "add_ticker_var"):
            self.add_ticker_var.set("")
        if hasattr(self, "add_amount_var"):
            self.add_amount_var.set("")
        if hasattr(self, "apply_btn"):
            self.apply_btn.configure(state=tk.DISABLED)
        self._update_action_states()

    def _sync_fallback_uncertainty_entry(self, ticker: str | None = None, source: str | None = None):
        if not hasattr(self, "fallback_uncertainty_entry"):
            return
        if self._is_ui_locked():
            self.fallback_uncertainty_entry.configure(state=tk.DISABLED)
            return
        if ticker is None or source is None:
            selected = self.port_tree.selection()
            ticker = ""
            source = ""
            if selected:
                values = self.port_tree.item(selected[0], "values")
                if values:
                    ticker = normalize_ticker_symbol(str(values[0]))
                    source = str(values[-1]) if len(values) > 0 else ""
        if str(source or "").startswith("AI"):
            self.fallback_uncertainty_entry.configure(state=tk.DISABLED)
            self.fallback_uncertainty_var.set(1.0)
            return
        self.fallback_uncertainty_entry.configure(state=tk.NORMAL)
        if ticker:
            self.fallback_uncertainty_var.set(float(self.fallback_uncertainty_by_ticker.get(ticker, 1.0)))

    def _refresh_apply_button_state(self):
        if not hasattr(self, "apply_btn"):
            return
        locked = self._is_ui_locked()
        self.apply_btn.configure(state=tk.DISABLED if locked or self.last_result is None else tk.NORMAL)

    def _update_action_states(self):
        locked = self._is_ui_locked()
        standard_state = tk.DISABLED if locked else tk.NORMAL
        for widget in self._portfolio_action_widgets:
            try:
                widget.configure(state=standard_state)
            except Exception:
                pass
        for widget in self._settings_entries:
            try:
                widget.configure(state=standard_state)
            except Exception:
                pass
        if hasattr(self, "load_ai_button"):
            self.load_ai_button.configure(state=tk.DISABLED if locked else tk.NORMAL)
        if hasattr(self, "reload_button"):
            self.reload_button.configure(state=tk.DISABLED if locked else tk.NORMAL)
        self.start_button.configure(state=tk.DISABLED if locked or not self.records else tk.NORMAL)
        self.stop_button.configure(state=tk.NORMAL if self.is_optimizing else tk.DISABLED)
        self._refresh_apply_button_state()
        if hasattr(self, "fallback_uncertainty_entry"):
            self._sync_fallback_uncertainty_entry()

    def _set_busy(self, active: bool, message: str):
        self._busy_active = active
        self.status_var.set(message)
        if active:
            self.progress.start(12)
        else:
            self.progress.stop()
        self._update_action_states()

    def _drain_ui_queue(self):
        if getattr(self, "_closing", False):
            return
        try:
            while True:
                try:
                    event = self.ui_queue.get_nowait()
                except Empty:
                    break
                
                try:
                    if getattr(self, "_closing", False):
                        break
                    kind = event[0]
                    if kind == "log":
                        self._append_log(event[1])
                    elif kind == "call":
                        _, callback, args = event
                        callback(*args)
                    elif kind == "meta_loaded":
                        _, path, records = event[:3]
                        repricing_summary = event[3] if len(event) > 3 else ""
                        remote_sync_summary = event[4] if len(event) > 4 else ""
                        self._finish_meta_load(
                            path,
                            records,
                            repricing_summary=repricing_summary,
                            remote_sync_summary=remote_sync_summary,
                        )
                    elif kind == "meta_error":
                        self._handle_meta_error(event[1])
                    elif kind == "opt_done":
                        self._finish_optimization(event[1])
                    elif kind == "opt_error":
                        self._handle_optimization_error(event[1])
                    elif kind == "opt_stopped":
                        self._finish_stopped_optimization()
                    elif kind == "charts_ready":
                        self._display_charts(event[1])
                    elif kind == "fallback_ready":
                        if not ENABLE_HOLDING_FALLBACK_PREDICTIONS:
                            continue
                        if len(event) == 4 and isinstance(event[1], int):
                            _, generation, data, tickers = event
                            if generation != self._portfolio_state_generation:
                                continue
                        else:
                            _, data, tickers = event
                        self.fallback_prediction_cache.update(data)
                        for t in tickers:
                            ticker = normalize_ticker_symbol(t)
                            self.fallback_prediction_inflight.discard(ticker)
                            if ticker in data:
                                self.fallback_prediction_failures.pop(ticker, None)
                                if isinstance(self.fallback_prediction_cache.get(ticker), dict):
                                    self.fallback_prediction_cache[ticker].pop("not_found", None)
                                    self.fallback_prediction_cache[ticker].pop("no_prediction", None)
                            else:
                                fails = self.fallback_prediction_failures.get(ticker, 0) + 1
                                self.fallback_prediction_failures[ticker] = fails
                                if fails >= self.fallback_max_retries:
                                    if self._ticker_has_market_evidence(ticker):
                                        self.fallback_prediction_cache[ticker] = {"no_prediction": True}
                                    else:
                                        self.fallback_prediction_cache[ticker] = {"not_found": True}
                        if data:
                            self.log(f"Dopočítána fallback týdenní predikce pro {len(data)} tickerů.")
                        if tickers and len(data) < len(tickers):
                            missing = [
                                normalize_ticker_symbol(t)
                                for t in tickers
                                if normalize_ticker_symbol(t) not in data
                            ]
                            no_prediction = sum(
                                1
                                for t in missing
                                if self.fallback_prediction_cache.get(t, {}).get("no_prediction")
                            )
                            not_found = sum(
                                1
                                for t in missing
                                if self.fallback_prediction_cache.get(t, {}).get("not_found")
                            )
                            pending = max(len(missing) - no_prediction - not_found, 0)
                            parts = []
                            if no_prediction:
                                parts.append(f"{no_prediction} validních bez predikční historie")
                            if not_found:
                                parts.append(f"{not_found} nenalezených")
                            if pending:
                                parts.append(f"{pending} čeká na další pokus")
                            detail = ", ".join(parts) if parts else f"{len(missing)} bez výsledku"
                            self.log(f"Fallback predikce zatím nedostupná pro {len(missing)} tickerů ({detail}).")
                        self._refresh_portfolio_tree()
                except Exception as e:
                    if not getattr(self, "_closing", False):
                        self._append_log(f"Chyba při zpracování události {event[0] if event else '?'}: {e}")
        finally:
            if not getattr(self, "_closing", False):
                try:
                    self.root.after(100, self._drain_ui_queue)
                except tk.TclError:
                    self._closing = True

    @staticmethod
    def _prediction_reference_close_date(source_date: dt.date) -> dt.date:
        return source_date - dt.timedelta(days=(source_date.weekday() - 4) % 7)

    @staticmethod
    def _bounded_delta(value: float, cap: float) -> float:
        return max(-float(cap), min(float(cap), float(value)))

    @staticmethod
    def _price_at_or_before(pd, series, day: dt.date):
        try:
            normalized = pd.Series(series).dropna()
            if normalized.empty:
                return None
            target = pd.Timestamp(day).normalize()
            before = normalized[normalized.index <= target]
            if before.empty:
                return None
            value = float(before.iloc[-1])
            return value if math.isfinite(value) and value > 0.0 else None
        except Exception:
            return None

    def _adjust_horizon_for_midweek_move(
        self,
        horizon: str,
        values,
        move_since_reference: float,
    ) -> tuple[float, float, float, float] | None:
        p, cvar10, q90, uncertainty = _coerce_horizon_values(values)
        if not (_is_valid_return(p) and _is_valid_return(cvar10) and _is_valid_return(q90)):
            return None
        if not math.isfinite(move_since_reference) or move_since_reference <= -0.95:
            return None

        config = MIDWEEK_REPRICING_CONFIG.get(horizon, {"damping": 0.7, "cap": 0.06})
        damping = max(0.0, min(1.0, float(config.get("damping", 0.7))))
        cap = max(0.0, float(config.get("cap", 0.06)))
        denom = 1.0 + float(move_since_reference)

        raw_pred = (1.0 + float(p)) / denom - 1.0
        pred_delta = self._bounded_delta(raw_pred - float(p), cap)
        adj_pred = float(p) + damping * pred_delta
        adj_pred = max(adj_pred, -0.99)

        raw_q90 = (1.0 + float(q90)) / denom - 1.0
        q90_delta = self._bounded_delta(raw_q90 - float(q90), cap)
        adj_q90 = float(q90) + damping * q90_delta

        original_down_spread = max(float(p) - float(cvar10), 1e-8)
        raw_cvar = (1.0 + float(cvar10)) / denom - 1.0
        raw_down_spread = max(adj_pred - raw_cvar, 1e-8)
        spread_delta = self._bounded_delta(raw_down_spread - original_down_spread, cap)
        adj_down_spread = original_down_spread + damping * spread_delta

        adverse_move = (float(p) >= 0.0 and move_since_reference < 0.0) or (
            float(p) < 0.0 and move_since_reference > 0.0
        )
        if adverse_move:
            adj_down_spread = max(adj_down_spread, original_down_spread)
        else:
            adj_down_spread = max(adj_down_spread, original_down_spread * 0.80)

        adj_cvar10 = adj_pred - adj_down_spread
        if adverse_move:
            adj_cvar10 = min(adj_cvar10, float(cvar10))
        adj_cvar10 = max(adj_cvar10, -0.99)
        if adj_cvar10 >= adj_pred:
            adj_cvar10 = max(-0.99, adj_pred - 1e-8)

        original_up_spread = max(float(q90) - float(p), 1e-8)
        adj_q90 = max(adj_q90, adj_pred + original_up_spread * 0.50, adj_pred + 1e-8)
        return float(adj_pred), float(adj_cvar10), float(adj_q90), float(uncertainty or 0.0)

    def _recompute_record_aggregate(self, record: StockRecord, objective_config: ObjectiveConfig | None = None) -> None:
        agg_pred, agg_cvar10, agg_q90 = _aggregate_horizon_values(record.horizon_data or {})
        agg_std = _downside_tail_risk(agg_cvar10)
        record.forecast_pct = float(agg_pred)
        record.cvar10_pct = float(agg_cvar10)
        record.std_pct = float(agg_std)
        record.upside_pct = float(agg_q90)
        record.sharpe = _starr_score(record.forecast_pct, record.std_pct)
        if objective_config is not None:
            record.score = _record_objective_score(record, objective_config)

    def _apply_midweek_prediction_repricing(
        self,
        records: list[StockRecord],
        prediction_source_date: dt.date | None,
        objective_config: ObjectiveConfig | None = None,
        log_summary: bool = True,
        allow_download: bool = False,
        max_download_tickers: int | None = None,
    ) -> str:
        if not records or prediction_source_date is None:
            return ""
        reference_date = self._prediction_reference_close_date(prediction_source_date)
        today = dt.datetime.now().date()
        if today <= reference_date:
            return ""

        pd = _load_pandas()
        tickers = list(dict.fromkeys(
            normalize_ticker_symbol(record.ticker)
            for record in records
            if record.horizon_data and str(record.ticker or "").strip()
        ))
        if not tickers:
            return ""

        start = pd.Timestamp(reference_date) - pd.Timedelta(days=10)
        end = pd.Timestamp(dt.datetime.now()).normalize() + pd.Timedelta(days=1)
        data = self._cached_daily_close_prices(pd, tickers, start, end)

        if allow_download:
            missing_tickers = [ticker for ticker in tickers if ticker not in data.columns]
            if max_download_tickers is not None:
                missing_tickers = missing_tickers[:max(0, int(max_download_tickers))]
            if missing_tickers:
                yf = _load_yfinance()
                if yf is None:
                    if log_summary:
                        self.log("Midweek přecenění predikcí bez stažení: yfinance není dostupné.")
                else:
                    try:
                        downloaded = self._download_daily_close_prices_cached(yf, pd, missing_tickers, start, end)
                        if not downloaded.empty:
                            data = downloaded if data.empty else pd.concat([data, downloaded], axis=1)
                            data = data.loc[:, ~data.columns.duplicated(keep="last")].sort_index()
                    except Exception as exc:
                        if log_summary:
                            self.log(f"Midweek přecenění predikcí bez stažení: ceny se nepodařilo načíst ({exc}).")
        if data.empty:
            if log_summary:
                self.log(
                    "Midweek přecenění predikcí přeskočeno: v lokální cache nejsou denní ceny "
                    f"od páteční close {reference_date.strftime('%d.%m.%Y')}."
                )
            return ""

        adjusted = 0
        skipped = 0
        moves = []
        for record in records:
            ticker = normalize_ticker_symbol(record.ticker)
            if ticker not in data.columns or not record.horizon_data:
                skipped += 1
                continue
            series = self._normalize_close_series(pd, data[ticker])
            if series.empty:
                skipped += 1
                continue
            ref_price = self._price_at_or_before(pd, series, reference_date)
            if not ref_price:
                skipped += 1
                continue
            latest_series = series[series.index > pd.Timestamp(reference_date).normalize()]
            if latest_series.empty:
                skipped += 1
                continue
            latest_price = float(latest_series.iloc[-1])
            if not math.isfinite(latest_price) or latest_price <= 0.0:
                skipped += 1
                continue
            move = latest_price / float(ref_price) - 1.0
            if not math.isfinite(move) or abs(move) <= 1e-9:
                continue

            new_horizon_data = {}
            changed = False
            for horizon, values in (record.horizon_data or {}).items():
                adjusted_values = self._adjust_horizon_for_midweek_move(horizon, values, move)
                if adjusted_values is None:
                    new_horizon_data[horizon] = values
                    continue
                new_horizon_data[horizon] = adjusted_values
                changed = True
            if not changed:
                continue

            record.horizon_data = new_horizon_data
            self._recompute_record_aggregate(record, objective_config=objective_config)
            adjusted += 1
            moves.append(move)

        if adjusted:
            records.sort(key=lambda record: record.score if objective_config else record.sharpe, reverse=True)
            avg_move = sum(moves) / len(moves) if moves else 0.0
            summary = (
                f"Predikce přeceněny podle pohybu od páteční close {reference_date.strftime('%d.%m.%Y')} "
                f"pro {adjusted} tickerů (průměrný pohyb {avg_move * 100:+.2f} %)."
            )
            if log_summary:
                self.log(summary)
            return summary
        if log_summary and skipped:
            self.log(
                f"Midweek přecenění predikcí nic neupravilo; pro {skipped} tickerů chyběla vhodná cena "
                f"od {reference_date.strftime('%d.%m.%Y')}."
            )
        return ""

    def reload_meta_async(self):
        if self._is_ui_locked():
            return

        self.is_loading_meta = True
        self._set_busy(True, "Načítám EMA predikce...")
        self.log("Spouštím EMA nad modelovými predikcemi; fallback zůstává jen pro aktuálně držené akcie.")
        
        self.fallback_prediction_cache.clear()
        self.fallback_prediction_inflight.clear()
        self.fallback_prediction_failures.clear()

        alphas = {
            "1w": float(self.alpha_1w_var.get()),
            "4w": float(self.alpha_4w_var.get()),
            "13w": float(self.alpha_13w_var.get()),
            "26w": float(self.alpha_26w_var.get()),
            "52w": float(self.alpha_52w_var.get()),
        }
        objective_config = self._current_objective_config()
        self._save_prediction_source_config()
        remote_prediction_config = {
            "enabled": bool(self.remote_predictions_enabled_var.get()),
            "manifest_url": str(self.remote_predictions_manifest_url_var.get() or "").strip(),
        }

        def worker():
            try:
                remote_sync_summary = self._sync_remote_predictions_if_enabled(remote_prediction_config)
                records = load_historical_meta_and_calculate_ema(self.base_dir, alphas, self.log)
                latest = find_latest_prediction_csv(self.base_dir, EMA_MODEL_PREDICTION_START_DATE)
                path = latest[1] if latest else Path("EMA predikce")
                reference_date = latest[0] if latest else None
                repricing_summary = self._apply_midweek_prediction_repricing(
                    records,
                    reference_date,
                    objective_config=objective_config,
                    log_summary=True,
                )
                self.ui_queue.put(("meta_loaded", path, records, repricing_summary, remote_sync_summary))
            except Exception as exc:
                self.ui_queue.put(("meta_error", str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _source_label_for_meta_path(self, path: Path, remote_sync_summary: str = "") -> str:
        path = Path(path)
        manifest_path = self._remote_prediction_cache_dir() / "last_manifest.json"
        if path.name.startswith("predictions_") and "_remote_" in path.name and manifest_path.exists():
            try:
                with open(manifest_path, "r", encoding="utf-8") as f:
                    manifest = json.load(f)
                if isinstance(manifest, dict) and manifest.get("cached_file") == path.name:
                    source_name = str(manifest.get("source_name") or manifest.get("file") or path.name)
                    return f"EMA predikce z online GitHubu: {source_name} (cache: {path.name})"
            except Exception:
                pass
        if str(remote_sync_summary or "").startswith("Online beta predikce aktuální"):
            return f"EMA predikce z online GitHubu: {path.name}"
        return f"EMA predikce po poslední soubor: {path.name}"

    def _finish_meta_load(
        self,
        path: Path,
        records: list[StockRecord],
        repricing_summary: str = "",
        remote_sync_summary: str = "",
    ):
        self.meta_path = path
        self._apply_record_scores(records)
        self.records = records
        source_label = self._source_label_for_meta_path(self.meta_path, remote_sync_summary)
        self.meta_label_var.set(source_label)
        self._refresh_stock_table()
        self.is_loading_meta = False
        status = f"Načteno {len(self.records)} akcií z {self.meta_path.name}."
        if remote_sync_summary:
            status += " " + remote_sync_summary
        if repricing_summary:
            status += " " + repricing_summary
        self._set_busy(False, status)
        self._append_log(f"Načteno {len(self.records)} akcií z {self.meta_path.name}. {source_label}")
        self._refresh_portfolio_tree()

    def _handle_meta_error(self, message: str):
        self.is_loading_meta = False
        self.records = []
        self.meta_path = None
        self.meta_label_var.set("EMA predikce se nepodařilo načíst.")
        self._set_busy(False, "Načtení EMA predikcí selhalo.")
        self._append_log(f"Chyba při načítání EMA predikcí: {message}")
        messagebox.showerror("Chyba", message)
        
    def _load_ai_csv(self):
        if self._blocked_action("Načtení AI kvantilů"):
            return
        if not self.records:
            messagebox.showwarning("Chyba", "Nejprve načtěte EMA predikce.")
            return
            
        path = filedialog.askopenfilename(
            title="Vyberte CSV s AI Kvantily",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        if not path:
            return
            
        try:
            pd = _load_pandas()
            df = pd.read_csv(path)
            if "Ticker" not in df.columns:
                messagebox.showerror("Chyba", "CSV neobsahuje sloupec 'Ticker'!")
                return
                
            updates = 0
            record_map = {r.ticker: r for r in self.records}
            import math
            
            for _, row in df.iterrows():
                ticker = normalize_ticker_symbol(str(row["Ticker"]))
                if ticker in record_map:
                    rec = record_map[ticker]
                    
                    q10_val = row.get("q10_1w")
                    q90_val = row.get("q90_1w")
                    mean_val = row.get("mean_1w")
                    
                    if pd.notna(q10_val) and pd.notna(mean_val):
                        if not (_is_valid_return(mean_val) and _is_valid_return(q10_val)):
                            continue
                        # CSV hodnoty bereme jako absolutní výnosové metriky.
                        # Pro STARR používáme jen ztrátu tailu pod nulou.
                        downside = _downside_tail_risk(float(q10_val))
                        upside_bonus = max(0.0, float(q90_val)) if pd.notna(q90_val) and _is_valid_return(q90_val) else 0.0
                        
                        rec.std_pct = downside
                        rec.upside_pct = upside_bonus
                        rec.forecast_pct = float(mean_val)
                        rec.cvar10_pct = float(q10_val)
                        rec.sharpe = _starr_score(rec.forecast_pct, downside)
                        rec.score = _record_objective_score(rec, self._current_objective_config())
                        updates += 1
                        
            # Re-sort a update UI
            self._apply_record_scores(self.records)
            self._finish_meta_load(self.meta_path if self.meta_path else Path("AI Data"), self.records)
            self.log(f"Úspěšně nahrazeno/aktualizováno {updates} Tickerů z CSV (q10 Downside, q90 Upside).")
            
        except Exception as e:
            messagebox.showerror("Chyba čtení CSV", str(e))

    def start_optimization(self):
        if self.is_loading_meta:
            messagebox.showinfo("Načítání dat", "Vyčkejte, než se dokončí načtení meta souboru.")
            return

        if self.worker_thread is not None and self.worker_thread.is_alive():
            messagebox.showinfo("Probíhá výpočet", "Vyčkejte na dokončení nebo použijte 'Zastavit'.")
            return

        if self._busy_active:
            messagebox.showinfo("Probíhá zpracování", "Vyčkejte, než se dokončí aktuální zpracování dat.")
            return

        if not self.records or self.meta_path is None:
            messagebox.showwarning("Chybí data", "Nejdříve načtěte meta soubor.")
            return

        try:
            holdings = parse_holdings_text(self.get_holdings_text())
            extra_cash_czk = _safe_float(self.extra_cash_var.get(), default=0.0)
            target_positions = int(self.target_positions_var.get())
            max_turnover = self.max_turnover_var.get().strip()
            max_sells = self.max_swaps_var.get().strip()
            max_weight_pct = _safe_float(self.max_weight_var.get(), default=8.0)
            sale_penalty_rate = float(self.sale_penalty_var.get())
            min_trade_czk = _safe_float(self.min_trade_var.get(), default=25.0)
            min_position_czk = _safe_float(self.min_position_var.get(), default=0.0)
            objective_config = self._current_objective_config()
        except Exception as exc:
            messagebox.showerror("Chyba vstupu", str(exc))
            return

        record_tickers = {r.ticker for r in self.records}
        missing = [t for t in holdings.keys() if t not in record_tickers]
        if missing:
            msg = (
                f"Tyto tickery z portfolia nemají aktuální modelovou predikci:\n\n"
                f"{', '.join(missing)}\n\n"
                "Použiju pro ně fallback jen jako existující držené pozice. "
                "Nebudou zařazeny do vesmíru pro nové nákupy."
            )
            messagebox.showwarning("Chybí modelová predikce", msg)

        if target_positions <= 0:
            messagebox.showerror("Chyba vstupu", "Cílový počet akcií musí být kladný.")
            return
        if extra_cash_czk < 0:
            messagebox.showerror("Chyba vstupu", "Dodatečný kapitál nemůže být záporný.")
            return
        if min_trade_czk < 0:
            messagebox.showerror("Chyba vstupu", "Minimální obchod nemůže být záporný.")
            return
        if min_position_czk < 0:
            messagebox.showerror("Chyba vstupu", "Minimální velikost pozice nemůže být záporná.")
            return
        total_capital_czk = sum(holdings.values()) + extra_cash_czk
        if min_position_czk > 0 and min_position_czk > total_capital_czk:
            messagebox.showerror(
                "Chyba vstupu",
                "Minimální velikost pozice je vyšší než celý dostupný kapitál.",
            )
            return
        if max_weight_pct <= 0 or max_weight_pct > 100:
            messagebox.showerror("Chyba vstupu", "Max váha akcie musí být v rozsahu 0-100 %.")
            return
        max_position_weight = max_weight_pct / 100.0
        position_plan = _position_count_plan_for_constraints(
            target_positions,
            total_capital_czk,
            min_position_czk,
            max_position_weight,
        )
        if position_plan.infeasible_reason:
            messagebox.showerror(
                "Chyba vstupu",
                position_plan.infeasible_reason,
            )
            return
        max_turnover_czk = None if not max_turnover else _safe_float(max_turnover)
        max_sell_trades = None if not max_sells else int(_safe_float(max_sells))
        if max_sell_trades is not None and max_sell_trades < 0:
            messagebox.showerror("Chyba vstupu", "Max počet prodejů nemůže být záporný.")
            return

        self.stop_event = threading.Event()
        self.is_optimizing = True
        self._set_busy(True, "Probíhá optimalizace portfolia...")
        self.log("Výpočet spuštěn. Připravuji optimalizaci portfolia.")
        self._apply_record_scores(self.records, objective_config)
        self._refresh_stock_table()
        no_sell_tickers = self._active_no_sell_tickers()
        if no_sell_tickers:
            self.log("Pozice chráněné před prodejem: " + ", ".join(sorted(no_sell_tickers)))

        def worker():
            try:
                base_records = list(self.records)
                fallback_records = self._build_fallback_records_for_holdings(
                    holdings,
                    DEFAULT_FALLBACK_UNCERTAINTY,
                    objective_config,
                )
                unresolved = sorted(
                    set(missing)
                    - {normalize_ticker_symbol(record.ticker) for record in fallback_records}
                )
                if unresolved:
                    raise RuntimeError(
                        "Nepodařilo se vytvořit fallback predikci pro držené pozice: "
                        + ", ".join(unresolved)
                    )
                if fallback_records:
                    self.log(
                        f"Používám fallback predikce pro {len(fallback_records)} držených tickerů "
                        "mimo nákupní universe."
                    )
                manager = PortfolioManager(
                    records=base_records + fallback_records,
                    current_holdings_czk=holdings,
                    extra_cash_czk=extra_cash_czk,
                    target_positions=target_positions,
                    sale_penalty_rate=sale_penalty_rate,
                    max_turnover_czk=max_turnover_czk,
                    max_position_replacements=max_sell_trades,
                    min_trade_czk=min_trade_czk,
                    min_position_czk=min_position_czk,
                    max_position_weight=max_position_weight,
                    no_sell_tickers=no_sell_tickers,
                    objective_config=objective_config,
                    log_fn=self.log,
                    stop_event=self.stop_event,
                )
                result = manager.optimize()
                if self.stop_event is not None and self.stop_event.is_set():
                    self.ui_queue.put(("opt_stopped",))
                    return
                self.ui_queue.put(("opt_done", result))
            except Exception as exc:  # pragma: no cover - UI path
                self.ui_queue.put(("opt_error", str(exc)))

        self.worker_thread = threading.Thread(target=worker, daemon=True)
        self.worker_thread.start()

    def request_stop(self):
        if self.stop_event is not None:
            self.stop_event.set()
            self.status_var.set("Zastavuji výpočet...")
            self._append_log("Byl odeslán požadavek na zastavení výpočtu.")

    def _finish_optimization(self, result: OptimizationResult):
        self.is_optimizing = False
        self.last_result = result
        self._set_busy(False, "Optimalizace dokončena.")
        self._refresh_apply_button_state()
        self.log(f"Optimalizace hotova. Cílové skóre: {result.sharpe:.3f}")
        self._log_trade_changes(result)
        self._append_current_equity_snapshot(replace_recent=True)
        self._store_original_and_optimized_projections(result)
        self._start_chart_computation(result)

    def _store_original_and_optimized_projections(self, result: OptimizationResult):
        import datetime
        now = datetime.datetime.now().isoformat()
        current_projection = self._compose_projection_from_holdings(self.portfolio_dict, label="Původní portfolio", source="current", created_at=now)
        if current_projection:
            self._append_projection(current_projection, replace_same_day_sources={"current"})
        
        optimized_holdings = {t.ticker: max(0.0, float(t.target_czk)) for t in result.trades if t.target_czk > 0}
        optimized_projection = self._compose_projection_from_holdings(
            optimized_holdings, label="Optimalizované portfolio", source="optimized",
            created_at=now, override_total=current_projection["start_value"] if current_projection else None
        )
        if optimized_projection:
            self._append_projection(optimized_projection)

    @staticmethod
    def _projection_day(proj) -> dt.date | None:
        if not isinstance(proj, dict):
            return None
        try:
            return dt.datetime.fromisoformat(str(proj.get("date", ""))).date()
        except Exception:
            return None

    @staticmethod
    def _projection_source(proj) -> str:
        if not isinstance(proj, dict):
            return ""
        source = str(proj.get("source", "") or "").strip().lower()
        if source:
            return source
        label = str(proj.get("label", "") or "").strip().lower()
        if "původní" in label or "puvodni" in label:
            return "current"
        if "optimaliz" in label:
            return "optimized"
        return ""

    def _projection_identity(self, proj) -> str:
        if not isinstance(proj, dict):
            return ""
        identity = str(proj.get("id") or proj.get("_pm_id") or "").strip()
        if identity:
            return identity
        expected_values = proj.get("expected", [])
        if not isinstance(expected_values, (list, tuple)):
            expected_values = []
        parts = [
            self._projection_source(proj),
            str(proj.get("label", "")),
            str(proj.get("date", "")),
            str(proj.get("start_value", "")),
            "|".join(str(value) for value in expected_values),
        ]
        return "legacy:" + "||".join(parts)

    def _append_projection(self, projection: dict, replace_same_day_sources: set[str] | None = None):
        projection = dict(projection)
        projection.setdefault(
            "_pm_id",
            f"proj-{dt.datetime.now().strftime('%Y%m%d%H%M%S%f')}-{len(self.portfolio_projections)}",
        )
        replace_same_day_sources = replace_same_day_sources or set()
        source = self._projection_source(projection)
        day = self._projection_day(projection)
        if source in replace_same_day_sources and day is not None:
            for idx, existing in enumerate(self.portfolio_projections):
                if self._projection_source(existing) == source and self._projection_day(existing) == day:
                    self.portfolio_projections[idx] = projection
                    return
        self.portfolio_projections.append(projection)

    def _chart_projections(self, projections: list[dict]) -> list[dict]:
        cleaned: list[dict] = []
        current_by_day: dict[dt.date, int] = {}
        seen_exact: set[tuple] = set()
        for proj in projections:
            if not isinstance(proj, dict) or self._is_custom_projection(proj):
                continue
            source = self._projection_source(proj)
            day = self._projection_day(proj)
            if source == "current" and day is not None:
                if day in current_by_day:
                    cleaned[current_by_day[day]] = proj
                else:
                    current_by_day[day] = len(cleaned)
                    cleaned.append(proj)
                continue

            expected_values = proj.get("expected", [])
            if not isinstance(expected_values, (list, tuple)):
                expected_values = []
            try:
                start_signature = round(float(proj.get("start_value", 0.0) or 0.0), 2)
            except (TypeError, ValueError):
                start_signature = 0.0
            signature = (
                source,
                str(proj.get("label", "")),
                day,
                start_signature,
                tuple(round(float(x), 2) for x in expected_values if _is_finite_number(x)),
            )
            if signature in seen_exact:
                continue
            seen_exact.add(signature)
            cleaned.append(proj)
        return cleaned

    def _cash_flow_refs_from_history(self) -> set[str]:
        refs: set[str] = set()
        for snapshot in self.portfolio_history:
            if not isinstance(snapshot, dict):
                continue
            raw_refs = snapshot.get("cash_flow_refs", [])
            if isinstance(raw_refs, list):
                refs.update(str(ref) for ref in raw_refs if str(ref))
        return refs

    def _last_portfolio_history_timestamp(self) -> str | None:
        latest = None
        for snapshot in self.portfolio_history:
            if not isinstance(snapshot, dict) or not snapshot.get("timestamp"):
                continue
            try:
                ts = dt.datetime.fromisoformat(str(snapshot["timestamp"]))
            except Exception:
                continue
            if latest is None or ts > latest:
                latest = ts
        return latest.isoformat() if latest is not None else None

    @staticmethod
    def _snapshot_holdings_signature(raw_holdings) -> dict[str, float] | None:
        if not isinstance(raw_holdings, dict):
            return None
        cleaned: dict[str, float] = {}
        for ticker, value in raw_holdings.items():
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                continue
            if math.isfinite(numeric) and numeric > 0:
                cleaned[normalize_ticker_symbol(ticker)] = numeric
        return cleaned

    @staticmethod
    def _snapshot_cash_value(snapshot: dict) -> float:
        return _safe_float(snapshot.get("extra_cash_czk", snapshot.get("cash_czk", 0.0)), default=0.0)

    def _current_holdings_snapshot(self) -> dict[str, float]:
        return {
            normalize_ticker_symbol(ticker): float(value)
            for ticker, value in self._coerce_holdings_dict(self.portfolio_dict).items()
            if float(value) > 0
        }

    def _enrich_legacy_history_snapshots(self) -> int:
        """Old portfolio files stored only total_equity; add composition anchors for charts."""
        if not isinstance(self.portfolio_history, list):
            self.portfolio_history = []
            return 0
        holdings = self._current_holdings_snapshot()
        if not holdings:
            return 0
        cash = _safe_float(self.extra_cash_var.get(), default=0.0)
        changed = 0
        for snapshot in self.portfolio_history:
            if not isinstance(snapshot, dict) or not snapshot.get("timestamp"):
                continue
            if "total_equity" not in snapshot:
                continue
            record_changed = False
            if not self._snapshot_holdings_signature(snapshot.get("holdings")):
                snapshot["holdings"] = dict(holdings)
                record_changed = True
            if "extra_cash_czk" not in snapshot and "cash_czk" not in snapshot:
                snapshot["extra_cash_czk"] = float(cash)
                record_changed = True
            if record_changed:
                changed += 1
        return changed

    def _snapshot_composition_matches(self, old_snapshot: dict, new_snapshot: dict, tolerance: float = 0.004) -> bool:
        old_holdings = self._snapshot_holdings_signature(old_snapshot.get("holdings"))
        new_holdings = self._snapshot_holdings_signature(new_snapshot.get("holdings"))
        if old_holdings is None or new_holdings is None:
            return False
        old_keys = set(old_holdings)
        new_keys = set(new_holdings)
        if old_keys != new_keys:
            return False

        old_cash = self._snapshot_cash_value(old_snapshot)
        new_cash = self._snapshot_cash_value(new_snapshot)
        old_total = old_cash + sum(old_holdings.values())
        new_total = new_cash + sum(new_holdings.values())
        if old_total <= 0 or new_total <= 0:
            return False

        keys = sorted(old_keys)
        for ticker in keys:
            old_weight = old_holdings[ticker] / old_total
            new_weight = new_holdings[ticker] / new_total
            if abs(old_weight - new_weight) > tolerance:
                return False
        if abs((old_cash / old_total) - (new_cash / new_total)) > tolerance:
            return False
        return True

    def _append_current_equity_snapshot(
        self,
        replace_recent: bool = False,
        external_flow_czk: float = 0.0,
        cash_flow_refs: list[str] | None = None,
        cash_flow_note: str = "",
        replace_same_day: bool = False,
    ):
        import datetime
        total_cash = float(self.extra_cash_var.get() or 0.0)
        total_eq = total_cash + sum(float(v) for v in self.portfolio_dict.values())
        now = datetime.datetime.now()
        new_refs = [str(ref) for ref in (cash_flow_refs or []) if str(ref)]
        snapshot = {
            "timestamp": now.isoformat(),
            "total_equity": float(total_eq),
            "holdings": self._current_holdings_snapshot(),
            "extra_cash_czk": float(total_cash),
        }
        flow_value = float(external_flow_czk or 0.0)
        if abs(flow_value) > 1e-9:
            snapshot["external_flow_czk"] = flow_value
        if new_refs:
            snapshot["cash_flow_refs"] = sorted(set(new_refs))
        if cash_flow_note:
            snapshot["cash_flow_note"] = cash_flow_note
        if replace_recent and self.portfolio_history:
            try:
                last_snapshot = self.portfolio_history[-1] if isinstance(self.portfolio_history[-1], dict) else {}
                last_ts = datetime.datetime.fromisoformat(last_snapshot.get("timestamp", ""))
                if (
                    (last_ts.date() == now.date() or (now - last_ts).total_seconds() < 3600)
                    and (replace_same_day or self._snapshot_composition_matches(last_snapshot, snapshot))
                ):
                    existing_refs = {
                        str(ref)
                        for ref in last_snapshot.get("cash_flow_refs", [])
                        if str(ref)
                    } if isinstance(last_snapshot.get("cash_flow_refs", []), list) else set()
                    refs_to_add = [ref for ref in new_refs if ref not in existing_refs]
                    merged_refs = sorted(existing_refs.union(refs_to_add))
                    existing_flow = _safe_float(last_snapshot.get("external_flow_czk", 0.0), default=0.0)
                    flow_to_add = flow_value if (not new_refs or refs_to_add) else 0.0
                    merged_flow = existing_flow + flow_to_add
                    if abs(merged_flow) > 1e-9:
                        snapshot["external_flow_czk"] = merged_flow
                    if merged_refs:
                        snapshot["cash_flow_refs"] = merged_refs
                    if not cash_flow_note and last_snapshot.get("cash_flow_note"):
                        snapshot["cash_flow_note"] = last_snapshot.get("cash_flow_note")
                    self.portfolio_history[-1] = snapshot
                    return
            except Exception:
                pass
        self.portfolio_history.append(snapshot)

    def _log_trade_changes(self, result: OptimizationResult):
        changes = [t for t in result.trades if abs(t.delta_czk) >= max(1e-9, result.min_trade_czk)]
        if not changes:
            self.log("Změny portfolia: bez významných změn.")
            return
        self.log("Změny portfolia (návrh):")
        for t in sorted(changes, key=lambda x: abs(x.delta_czk), reverse=True)[:40]:
            self.log(
                f" - {t.ticker}: {t.action} {abs(t.delta_czk):.0f} CZK "
                f"(aktuálně {t.current_czk:.0f} -> cíl {t.target_czk:.0f})"
            )

    def _handle_optimization_error(self, message: str):
        self.is_optimizing = False
        self._set_busy(False, "Optimalizace selhala.")
        self._append_log(f"Chyba při optimalizaci: {message}")
        messagebox.showerror("Chyba", message)

    def _finish_stopped_optimization(self):
        self.is_optimizing = False
        self._set_busy(False, "Výpočet byl zastaven.")
        self._append_log("Výpočet byl zastaven uživatelem.")

    def _load_trading212_config(self) -> dict:
        config = {}
        if self.t212_config_path.exists():
            try:
                with open(self.t212_config_path, "r", encoding="utf-8") as f:
                    loaded = json.load(f)
                    if isinstance(loaded, dict):
                        config.update(loaded)
            except Exception:
                pass
        env_key = os.environ.get("TRADING212_API_KEY")
        env_secret = os.environ.get("TRADING212_API_SECRET")
        if env_key:
            config["api_key"] = env_key
        if env_secret:
            config["api_secret"] = env_secret
        return config

    def _save_trading212_config(self, environment: str, api_key: str, api_secret: str):
        self.t212_config_path.parent.mkdir(parents=True, exist_ok=True)
        existing = self._load_trading212_config()
        profiles = existing.get("profiles") if isinstance(existing.get("profiles"), dict) else {}
        profiles[environment] = {"api_key": api_key, "api_secret": api_secret}
        data = {"environment": environment, "profiles": profiles}
        with open(self.t212_config_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4)

    @staticmethod
    def _trading212_env_name(value: str) -> str:
        env = str(value or "live").strip().lower()
        if env in {"demo", "paper trading", "paper_trading", "demo/paper", "paper/demo"}:
            return "paper"
        return env if env in {"live", "paper"} else "live"

    def _trading212_profile_credentials(self, config: dict, environment: str) -> tuple[str, str]:
        profiles = config.get("profiles") if isinstance(config.get("profiles"), dict) else {}
        profile = profiles.get(environment) if isinstance(profiles.get(environment), dict) else {}
        api_key = str(profile.get("api_key") or "")
        api_secret = str(profile.get("api_secret") or "")
        if profiles:
            return api_key, api_secret
        root_environment = self._trading212_env_name(str(config.get("environment", "live")))
        if environment == root_environment and not api_key:
            api_key = str(config.get("api_key") or "")
        if environment == root_environment and not api_secret:
            api_secret = str(config.get("api_secret") or "")
        return api_key, api_secret

    def _trading212_client_for_link(self, link: dict | None = None) -> Trading212Client:
        link = link or self.portfolio_broker_link
        environment = self._trading212_env_name(str(link.get("environment") or self._load_trading212_config().get("environment", "live")))
        config = self._load_trading212_config()
        api_key, api_secret = self._trading212_profile_credentials(config, environment)
        return Trading212Client(environment, api_key, api_secret)

    def _t212_enrich_source_link(self, source: dict, link: dict) -> dict:
        source = dict(source)
        source_link = source.get("broker_link") if isinstance(source.get("broker_link"), dict) else {}
        merged = dict(source_link)
        for key, value in link.items():
            if value not in (None, "", {}):
                merged[key] = value
        source["broker_link"] = merged
        return source

    def _t212_source_from_linked_snapshot(self, snapshot: dict, link: dict) -> dict | None:
        source_id = str(link.get("source_id") or "")
        pie_id = str(link.get("pie_id") or "")
        if not source_id and pie_id:
            source_id = f"pie:{pie_id}"
        sources = self._t212_sources_from_snapshot(snapshot)
        for source in sources:
            if str(source.get("id") or "") == source_id:
                return self._t212_enrich_source_link(source, link)
        return None

    def _t212_snapshot_for_linked_source(self, client: Trading212Client, link: dict) -> dict:
        source_id = str(link.get("source_id") or "")
        pie_id = str(link.get("pie_id") or "")
        if not pie_id and source_id.startswith("pie:"):
            pie_id = source_id.split(":", 1)[1]

        summary = client.account_summary()
        positions = client.positions()
        if source_id.startswith("pie:") and pie_id:
            pie_summaries = []
            pie_summary = None
            try:
                pie_summaries = client.pies()
                for pie in pie_summaries:
                    if str(Trading212Client._pie_id(pie)) == str(pie_id):
                        pie_summary = pie
                        break
            except Trading212ApiError:
                pie_summaries = []

            pie_details = []
            pie_errors = []
            try:
                detail = client.pie_detail(pie_id, retries=2, retry_wait_seconds=4.0)
                if detail:
                    detail.setdefault("_summary", pie_summary or detail)
                    pie_details.append(detail)
            except Trading212ApiError as exc:
                pie_errors.append({"id": pie_id, "error": str(exc), "summary": pie_summary or {}})

            return {
                "summary": summary,
                "positions": positions,
                "pies": [pie_summary] if isinstance(pie_summary, dict) else pie_summaries,
                "pie_details": pie_details,
                "pie_errors": pie_errors,
            }

        return {
            "summary": summary,
            "positions": positions,
            "pies": [],
            "pie_details": [],
            "pie_errors": [],
        }

    def _is_trading212_linked_portfolio(self) -> bool:
        link = self.portfolio_broker_link if isinstance(self.portfolio_broker_link, dict) else {}
        return link.get("broker") == "Trading 212" and bool(link.get("source_id"))

    @staticmethod
    def _cash_flow_amount_from_t212_transaction(tx: dict) -> float:
        tx_type = str(tx.get("type") or "").strip().upper()
        amount = _safe_float(tx.get("amount", 0.0), default=0.0)
        if tx_type == "DEPOSIT":
            return abs(amount)
        if tx_type == "WITHDRAW":
            return -abs(amount)
        return 0.0

    def _trading212_external_flows_since(
        self,
        client: Trading212Client,
        since_iso: str | None,
        known_refs: set[str],
    ) -> tuple[float, list[str]]:
        if not since_iso:
            return 0.0, []
        try:
            since_dt = dt.datetime.fromisoformat(str(since_iso))
        except Exception:
            since_dt = None
        total_flow = 0.0
        refs: list[str] = []
        for tx in client.transactions(since_iso):
            if not isinstance(tx, dict):
                continue
            flow = self._cash_flow_amount_from_t212_transaction(tx)
            if abs(flow) <= 1e-9:
                continue
            raw_ref = str(tx.get("reference") or "").strip()
            ref = raw_ref or f"{tx.get('type')}:{tx.get('dateTime')}:{tx.get('amount')}"
            if ref in known_refs or ref in refs:
                continue
            if since_dt is not None and tx.get("dateTime"):
                try:
                    tx_dt = dt.datetime.fromisoformat(str(tx["dateTime"]).replace("Z", "+00:00"))
                    compare_since = since_dt
                    if tx_dt.tzinfo is not None and compare_since.tzinfo is None:
                        compare_since = compare_since.replace(tzinfo=tx_dt.tzinfo)
                    if tx_dt < compare_since:
                        continue
                except Exception:
                    pass
            total_flow += flow
            refs.append(ref)
        return total_flow, refs

    def _refresh_trading212_linked_portfolio_async(
        self,
        reason: str = "Načítám aktuální stav z Trading 212...",
        notify: bool = False,
        parent=None,
    ):
        if self._blocked_action("Aktualizace Trading 212"):
            return False
        link = self.portfolio_broker_link if isinstance(self.portfolio_broker_link, dict) else {}
        if link.get("broker") != "Trading 212" or not link.get("source_id"):
            return False
        try:
            client = self._trading212_client_for_link(link)
        except Trading212ApiError as exc:
            self.log(f"Trading 212 refresh přeskočen: {exc}")
            if notify:
                messagebox.showerror("Trading 212", str(exc), parent=parent or self.root)
            return False
        since_iso = str(link.get("cash_flow_checked_at") or self._last_portfolio_history_timestamp() or "")
        known_cash_flow_refs = self._cash_flow_refs_from_history()
        self.log(reason)
        self._set_busy(True, reason)

        def dialog_parent():
            if parent is not None:
                try:
                    if parent.winfo_exists():
                        return parent
                except Exception:
                    pass
            return self.root

        def finish(source, error, external_flow_czk=0.0, cash_flow_refs=None, flow_error=None):
            final_status = "Trading 212 aktualizace dokončena."
            try:
                if error:
                    final_status = "Trading 212 aktualizace selhala."
                    self.log(f"Trading 212 refresh selhal: {error}")
                    if notify:
                        messagebox.showerror("Trading 212", f"Aktualizace z Trading 212 selhala:\n{error}", parent=dialog_parent())
                    return
                if not source:
                    final_status = "Trading 212 zdroj nebyl nalezen."
                    self.log("Trading 212 refresh: uložený zdroj už nebyl nalezen.")
                    if notify:
                        messagebox.showwarning(
                            "Trading 212",
                            "Uložený Trading 212 zdroj už nebyl nalezen. Zkus portfolio z Trading 212 znovu importovat.",
                            parent=dialog_parent(),
                        )
                    return
                if source.get("load_error"):
                    final_status = "Trading 212 detail koláče se nepodařilo načíst."
                    self.log(f"Trading 212 refresh přeskočen: {source.get('load_error')}")
                    if notify:
                        messagebox.showwarning(
                            "Trading 212",
                            "Trading 212 detail napojeného koláče se teď nepodařilo načíst. "
                            "Lokální portfolio jsem proto nepřepsal prázdnými hodnotami.",
                            parent=dialog_parent(),
                        )
                    return
                self._apply_trading212_source(source, record_history=False)
                self._append_current_equity_snapshot(
                    replace_recent=True,
                    external_flow_czk=float(external_flow_czk or 0.0),
                    cash_flow_refs=cash_flow_refs or [],
                    cash_flow_note="Trading 212 synchronizace",
                )
                if flow_error:
                    self.log(f"Trading 212 cash-flow historie se nepodařila načíst: {flow_error}")
                elif abs(float(external_flow_czk or 0.0)) > 1e-9:
                    direction = "vklad" if external_flow_czk > 0 else "výběr"
                    self.log(
                        f"Trading 212 cash-flow od minulé kontroly: {direction} "
                        f"{abs(external_flow_czk):,.2f} {self.portfolio_currency}. "
                        "Historie portfolia dostala novou kotvu stavu."
                    )
                if not flow_error:
                    self.portfolio_broker_link = dict(self.portfolio_broker_link or {})
                    self.portfolio_broker_link["cash_flow_checked_at"] = dt.datetime.now().isoformat()
                saved = False
                try:
                    if self.current_portfolio_path:
                        self._write_portfolio_file(quiet=True)
                        saved = True
                except Exception as exc:
                    self.log(f"Aktuální stav z Trading 212 se nepodařilo uložit: {exc}")
                if saved:
                    self.log("Portfolio aktualizováno podle aktuálního stavu Trading 212 a uloženo.")
                else:
                    self.log("Portfolio aktualizováno podle aktuálního stavu Trading 212.")
                if notify:
                    messagebox.showinfo(
                        "Trading 212",
                        "Portfolio bylo aktualizováno podle skutečného stavu v Trading 212.\n"
                        "Historie a uložené predikce zůstaly zachované.",
                        parent=dialog_parent(),
                    )
            finally:
                self._set_busy(False, final_status)

        def worker():
            try:
                snapshot = self._t212_snapshot_for_linked_source(client, link)
                source = self._t212_source_from_linked_snapshot(snapshot, link)
                flow = 0.0
                flow_refs: list[str] = []
                flow_error = None
                try:
                    flow, flow_refs = self._trading212_external_flows_since(client, since_iso, known_cash_flow_refs)
                except Trading212ApiError as exc:
                    flow_error = exc
                self.ui_queue.put(("call", finish, (source, None, flow, flow_refs, flow_error)))
            except Exception as exc:
                self.ui_queue.put(("call", finish, (None, exc, 0.0, [], None)))

        threading.Thread(target=worker, daemon=True).start()
        return True

    def _refresh_current_trading212_portfolio(self):
        if self._blocked_action("Aktualizace Trading 212"):
            return
        if not self._is_trading212_linked_portfolio():
            messagebox.showinfo(
                "Trading 212",
                "Aktuální portfolio není napojené na Trading 212. Nejdřív ho importuj tlačítkem Trading 212.",
                parent=self.root,
            )
            return
        self._refresh_trading212_linked_portfolio_async(
            "Aktualizuji portfolio podle skutečného stavu v Trading 212...",
            notify=True,
            parent=self.root,
        )

    @staticmethod
    def _t212_broker_base(broker_ticker: str) -> str:
        broker = str(broker_ticker or "").strip().upper()
        broker = re.sub(r"_[A-Z]{2,3}_(?:EQ|ETF|STOCK)$", "", broker)
        broker = re.sub(r"_(?:EQ|ETF|STOCK)_[A-Z]{2,3}$", "", broker)
        return broker

    @staticmethod
    def _t212_short_name_ticker(short_name: str) -> str:
        text = str(short_name or "").strip().upper()
        if not text or any(ch.isspace() for ch in text):
            return ""
        normalized = normalize_ticker_symbol(text)
        if re.match(r"^[A-Z0-9]+(?:-[A-Z0-9]+)?$", normalized):
            return normalized
        return ""

    @staticmethod
    def _t212_app_ticker(broker_ticker: str, name: str = "", isin: str = "", short_name: str = "") -> str:
        broker = str(broker_ticker or "").strip()
        market = PortfolioManagerApp._t212_market_code(broker)
        base = PortfolioManagerApp._t212_broker_base(broker)
        short_ticker = PortfolioManagerApp._t212_short_name_ticker(short_name)
        if short_ticker:
            resolved = resolve_instrument_ticker(short_ticker, isin=isin)
        else:
            resolved = resolve_instrument_ticker(base, name=name, isin=isin)
        if market and market != "US":
            return f"{resolved}.{market}"
        return resolved

    @staticmethod
    def _t212_market_code(broker_ticker: str, instrument: dict | None = None) -> str:
        broker = str(broker_ticker or "").strip().upper()
        match = re.search(r"_([A-Z]{2,3})_(?:EQ|ETF|STOCK)$", broker)
        if match:
            return match.group(1)
        match = re.search(r"_(?:EQ|ETF|STOCK)_([A-Z]{2,3})$", broker)
        if match:
            return match.group(1)
        if "_" not in broker and re.match(r"^[A-Z0-9./-]{1,12}$", broker):
            if instrument is None:
                return "US"
            currency = PortfolioManagerApp._t212_instrument_text(
                instrument,
                "currencyCode",
                "currency",
                "tradingCurrency",
            )
            isin = PortfolioManagerApp._t212_instrument_text(instrument, "isin")
            if currency == "USD" or isin.startswith("US"):
                return "US"
        return ""

    @staticmethod
    def _t212_is_us_broker_ticker(broker_ticker: str, instrument: dict | None = None) -> bool:
        return PortfolioManagerApp._t212_market_code(broker_ticker, instrument=instrument) == "US"

    @staticmethod
    def _t212_default_us_broker_ticker(app_ticker: str) -> str:
        ticker = normalize_ticker_symbol(app_ticker)
        # Trading 212 keeps some US class-share tickers with slash notation.
        broker_aliases = {
            "BRK-A": "BRK/A",
            "BRK-B": "BRK_B",
            "BF-B": "BFB",
            "CRD-B": "CRD/B",
            "HEI-A": "HEI/A",
        }
        broker_base = broker_aliases.get(ticker, ticker)
        return f"{broker_base}_US_EQ"

    @staticmethod
    def _t212_instrument_text(instrument: dict | None, *keys: str) -> str:
        if not isinstance(instrument, dict):
            return ""
        for key in keys:
            value = instrument.get(key)
            if value is not None:
                text = str(value).strip().upper()
                if text:
                    return text
        return ""

    @staticmethod
    def _t212_us_preference_score(
        broker_ticker: str,
        app_ticker: str = "",
        instrument: dict | None = None,
    ) -> int:
        broker = str(broker_ticker or "").strip().upper()
        if not broker:
            return -10_000

        score = 0
        market = PortfolioManagerApp._t212_market_code(broker, instrument=instrument)
        if market == "US":
            score += 1_000
        elif market:
            score -= 500

        if app_ticker:
            resolved = PortfolioManagerApp._t212_app_ticker(
                broker,
                name=PortfolioManagerApp._t212_instrument_text(instrument, "name"),
                short_name=PortfolioManagerApp._t212_instrument_text(instrument, "shortName", "short_name"),
                isin=PortfolioManagerApp._t212_instrument_text(instrument, "isin"),
            )
            if resolved == normalize_ticker_symbol(app_ticker):
                score += 50

        currency = PortfolioManagerApp._t212_instrument_text(
            instrument,
            "currencyCode",
            "currency",
            "tradingCurrency",
        )
        if currency == "USD":
            score += 80
        elif currency:
            score -= 20

        country = PortfolioManagerApp._t212_instrument_text(
            instrument,
            "countryCode",
            "country",
            "region",
        )
        if country in {"US", "USA", "UNITED STATES", "UNITED STATES OF AMERICA"}:
            score += 80
        elif country:
            score -= 10

        exchange = PortfolioManagerApp._t212_instrument_text(
            instrument,
            "exchange",
            "exchangeCode",
            "exchangeName",
            "market",
            "marketName",
        )
        us_exchange_markers = {
            "NASDAQ",
            "NYSE",
            "AMEX",
            "ARCA",
            "BATS",
            "CBOE",
            "IEX",
            "NMS",
            "NYQ",
            "ASE",
            "NGS",
        }
        if exchange in us_exchange_markers or any(marker in exchange for marker in ("NASDAQ", "NYSE", "AMEX")):
            score += 80

        return score

    @staticmethod
    def _prefer_t212_broker_ticker(
        current: str | None,
        candidate: str | None,
        app_ticker: str = "",
        current_instrument: dict | None = None,
        candidate_instrument: dict | None = None,
    ) -> str:
        current_ticker = str(current or "").strip()
        candidate_ticker = str(candidate or "").strip()
        if not candidate_ticker:
            return current_ticker
        if not current_ticker:
            return candidate_ticker

        current_score = PortfolioManagerApp._t212_us_preference_score(
            current_ticker,
            app_ticker=app_ticker,
            instrument=current_instrument,
        )
        candidate_score = PortfolioManagerApp._t212_us_preference_score(
            candidate_ticker,
            app_ticker=app_ticker,
            instrument=candidate_instrument,
        )
        if candidate_score > current_score:
            return candidate_ticker
        return current_ticker

    def _remember_preferred_t212_broker_ticker(
        self,
        mapping: dict[str, str],
        app_ticker: str,
        broker_ticker: str,
        instrument: dict | None = None,
    ) -> None:
        key = normalize_ticker_symbol(app_ticker)
        broker = str(broker_ticker or "").strip()
        if not key or not broker:
            return
        mapping[key] = self._prefer_t212_broker_ticker(
            mapping.get(key),
            broker,
            app_ticker=key,
            candidate_instrument=instrument,
        )

    @staticmethod
    def _t212_get_nested(data: dict, *paths):
        for path in paths:
            current = data
            ok = True
            for key in path:
                if isinstance(current, dict) and key in current:
                    current = current[key]
                else:
                    ok = False
                    break
            if ok and current is not None:
                return current
        return None

    def _t212_number(self, value, default: float = 0.0) -> float:
        number = self._positive_float(value)
        if number is None:
            return default
        return number

    def _t212_position_record(self, position: dict) -> dict | None:
        if not isinstance(position, dict):
            return None
        instrument = position.get("instrument") if isinstance(position.get("instrument"), dict) else {}
        wallet = position.get("walletImpact") if isinstance(position.get("walletImpact"), dict) else {}
        broker_ticker = position.get("ticker") or instrument.get("ticker")
        if not broker_ticker:
            return None
        app_ticker = self._t212_app_ticker(
            str(broker_ticker),
            name=str(instrument.get("name", "")),
            short_name=str(instrument.get("shortName", "")),
            isin=str(instrument.get("isin", "")),
        )
        quantity = self._t212_number(position.get("quantity"))
        current_price = self._t212_number(position.get("currentPrice", position.get("price")))
        average_price = self._t212_number(position.get("averagePricePaid", position.get("averagePrice")))
        current_value = self._t212_number(
            wallet.get("currentValue", position.get("currentValue", position.get("value")))
        )
        if current_value <= 0 and quantity > 0 and current_price > 0:
            current_value = quantity * current_price
        if not app_ticker or quantity <= 0 or current_value <= 0:
            return None
        quantity_in_pies = self._t212_number(position.get("quantityInPies", position.get("pieQuantity")))
        quantity_available = self._t212_number(
            position.get("quantityAvailableForTrading"),
            max(0.0, quantity - quantity_in_pies),
        )
        return {
            "app_ticker": app_ticker,
            "broker_ticker": str(broker_ticker),
            "name": instrument.get("name", app_ticker),
            "currency": instrument.get("currency") or wallet.get("currency", ""),
            "quantity": quantity,
            "quantity_in_pies": quantity_in_pies,
            "quantity_available": quantity_available,
            "current_price": current_price,
            "average_price": average_price,
            "current_value": current_value,
            "total_cost": self._t212_number(wallet.get("totalCost")),
            "unrealized_profit_loss": self._t212_number(wallet.get("unrealizedProfitLoss"), 0.0),
        }

    def _t212_cash_from_summary(self, summary: dict, include_pie_cash: bool = True) -> tuple[float, str]:
        cash = summary.get("cash") if isinstance(summary.get("cash"), dict) else {}
        available = self._t212_number(cash.get("availableToTrade", summary.get("free")))
        in_pies = self._t212_number(cash.get("inPies")) if include_pie_cash else 0.0
        currency = str(summary.get("currency") or cash.get("currencyCode") or summary.get("currencyCode") or "CZK")
        return available + in_pies, currency

    def _t212_source_from_records(
        self,
        source_id: str,
        label: str,
        records: list[dict],
        cash: float,
        currency: str,
        broker_link: dict | None = None,
    ) -> dict:
        holdings: dict[str, float] = {}
        details: dict[str, dict] = {}
        price_snapshot: dict[str, dict] = {}
        now = dt.datetime.now().isoformat()
        for record in records:
            ticker = record["app_ticker"]
            value = float(record.get("current_value", 0.0))
            if value <= 0:
                continue
            holdings[ticker] = holdings.get(ticker, 0.0) + value
            detail = details.setdefault(
                ticker,
                {
                    "broker": "Trading 212",
                    "source_id": source_id,
                    "source": label,
                    "currency": currency,
                    "quantity": 0.0,
                    "current_value": 0.0,
                    "positions": [],
                },
            )
            detail["quantity"] += float(record.get("quantity", 0.0))
            detail["current_value"] += value
            detail["positions"].append(record)
            price = self._positive_float(record.get("current_price"))
            if price:
                price_snapshot[ticker] = {
                    "price": float(price),
                    "timestamp": now,
                    "broker": "Trading 212",
                    "broker_ticker": record.get("broker_ticker", ticker),
                }
        return {
            "id": source_id,
            "label": label,
            "holdings": holdings,
            "cash": float(cash or 0.0),
            "currency": currency or "CZK",
            "broker_positions": details,
            "broker_link": broker_link or {
                "broker": "Trading 212",
                "source_id": source_id,
                "source_label": label,
                "type": "portfolio",
            },
            "price_snapshot": price_snapshot,
            "position_count": len(holdings),
            "invested_value": sum(holdings.values()),
        }

    def _t212_sources_from_snapshot(self, snapshot: dict) -> list[dict]:
        summary = snapshot.get("summary", {}) if isinstance(snapshot.get("summary"), dict) else {}
        raw_positions = snapshot.get("positions", []) if isinstance(snapshot.get("positions", []), list) else []
        positions = [p for p in (self._t212_position_record(pos) for pos in raw_positions) if p]
        position_by_broker = {p["broker_ticker"]: p for p in positions}
        all_cash, currency = self._t212_cash_from_summary(summary, include_pie_cash=True)
        sources = [
            self._t212_source_from_records(
                "all",
                "Celé portfolio včetně koláčů",
                positions,
                all_cash,
                currency,
                broker_link={
                    "broker": "Trading 212",
                    "source_id": "all",
                    "source_label": "Celé portfolio včetně koláčů",
                    "type": "portfolio",
                },
            )
        ]

        outside_records = []
        for pos in positions:
            quantity = float(pos.get("quantity", 0.0))
            outside_quantity = min(float(pos.get("quantity_available", 0.0)), quantity)
            if quantity <= 0 or outside_quantity <= 0:
                continue
            ratio = outside_quantity / quantity
            outside = dict(pos)
            outside["quantity"] = outside_quantity
            outside["quantity_in_pies"] = 0.0
            outside["current_value"] = float(pos.get("current_value", 0.0)) * ratio
            outside_records.append(outside)
        outside_cash, _ = self._t212_cash_from_summary(summary, include_pie_cash=False)
        if outside_records or outside_cash > 0:
            sources.append(
                self._t212_source_from_records(
                    "outside_pies",
                    "Volné pozice mimo koláče",
                    outside_records,
                    outside_cash,
                    currency,
                    broker_link={
                        "broker": "Trading 212",
                        "source_id": "outside_pies",
                        "source_label": "Volné pozice mimo koláče",
                        "type": "outside_pies",
                    },
                )
            )

        pie_summaries = {
            str(Trading212Client._pie_id(p)): p
            for p in snapshot.get("pies", [])
            if isinstance(p, dict) and Trading212Client._pie_id(p) is not None
        }
        added_pie_ids = set()
        for detail in snapshot.get("pie_details", []):
            if not isinstance(detail, dict):
                continue
            summary = detail.get("_summary") if isinstance(detail.get("_summary"), dict) else {}
            settings = detail.get("settings") if isinstance(detail.get("settings"), dict) else {}
            pie_id = Trading212Client._pie_id(detail) or Trading212Client._pie_id(summary)
            if pie_id is None:
                continue
            added_pie_ids.add(str(pie_id))
            pie_name = self._clean_temporary_t212_pie_name(
                settings.get("name") or summary.get("name") or detail.get("name") or pie_id
            )
            label = f"Koláč: {pie_name}"
            pie_settings = settings if settings else detail
            pie_summary = pie_summaries.get(str(pie_id), summary)
            pie_instrument_shares = self._extract_t212_pie_shares(detail) or pie_settings.get("instrumentShares")
            pie_records = []
            instruments = []
            for candidate in (
                detail.get("instruments"),
                detail.get("items"),
                detail.get("positions"),
                detail.get("slices"),
            ):
                if isinstance(candidate, list):
                    instruments = candidate
                    break
            for item in instruments:
                if not isinstance(item, dict):
                    continue
                broker_ticker = self._t212_get_nested(
                    item,
                    ("ticker",),
                    ("instrument", "ticker"),
                    ("instrumentTicker",),
                )
                owned_quantity = self._t212_number(
                    self._t212_get_nested(
                        item,
                        ("ownedQuantity",),
                        ("quantity",),
                        ("owned_quantity",),
                        ("instrument", "ownedQuantity"),
                    )
                )
                if not broker_ticker or owned_quantity <= 0:
                    continue
                base_record = position_by_broker.get(str(broker_ticker))
                if not base_record:
                    result = item.get("result") if isinstance(item.get("result"), dict) else {}
                    value = self._t212_number(
                        self._t212_get_nested(
                            item,
                            ("currentValue",),
                            ("value",),
                            ("walletImpact", "currentValue"),
                            ("result", "priceAvgValue"),
                        )
                    )
                    base_record = {
                        "app_ticker": self._t212_app_ticker(
                            str(broker_ticker),
                            name=str(self._t212_get_nested(item, ("name",), ("instrument", "name")) or ""),
                            short_name=str(self._t212_get_nested(item, ("shortName",), ("short_name",), ("instrument", "shortName")) or ""),
                            isin=str(self._t212_get_nested(item, ("isin",), ("instrument", "isin")) or ""),
                        ),
                        "broker_ticker": str(broker_ticker),
                        "name": str(self._t212_get_nested(item, ("name",), ("instrument", "name")) or self._t212_app_ticker(str(broker_ticker))),
                        "currency": currency,
                        "quantity": owned_quantity,
                        "current_price": 0.0,
                        "average_price": 0.0,
                        "current_value": value,
                    }
                    pie_records.append(base_record)
                    continue
                total_quantity = float(base_record.get("quantity", 0.0))
                if total_quantity <= 0:
                    continue
                ratio = min(owned_quantity / total_quantity, 1.0)
                pie_record = dict(base_record)
                pie_record["quantity"] = owned_quantity
                pie_record["quantity_available"] = 0.0
                pie_record["current_value"] = float(base_record.get("current_value", 0.0)) * ratio
                pie_records.append(pie_record)
            sources.append(
                self._t212_source_from_records(
                    f"pie:{pie_id}",
                    label,
                    pie_records,
                    self._t212_number(
                        self._t212_get_nested(
                            pie_summary,
                            ("cash",),
                            ("cashDetails", "availableToInvest"),
                            ("cashDetails", "availableForInvestment"),
                        )
                    ),
                    currency,
                    broker_link={
                        "broker": "Trading 212",
                        "source_id": f"pie:{pie_id}",
                        "source_label": label,
                        "type": "pie",
                        "pie_id": str(pie_id),
                        "pie_settings": {
                        "dividendCashAction": pie_settings.get("dividendCashAction"),
                        "endDate": pie_settings.get("endDate"),
                        "goal": pie_settings.get("goal"),
                        "icon": pie_settings.get("icon"),
                        "instrumentShares": pie_instrument_shares,
                        "name": pie_name,
                    },
                },
                )
            )
        for pie_id, pie_summary in pie_summaries.items():
            if pie_id in added_pie_ids:
                continue
            cash = self._t212_number(
                self._t212_get_nested(
                    pie_summary,
                    ("cash",),
                    ("cashDetails", "availableToInvest"),
                    ("cashDetails", "availableForInvestment"),
                )
            )
            pie_name = self._clean_temporary_t212_pie_name(pie_summary.get("name") or pie_id)
            source = self._t212_source_from_records(
                f"pie:{pie_id}",
                f"Koláč: {pie_name} (detail nenačten)",
                [],
                cash,
                currency,
                broker_link={
                    "broker": "Trading 212",
                    "source_id": f"pie:{pie_id}",
                    "source_label": f"Koláč: {pie_name}",
                    "type": "pie",
                    "pie_id": str(pie_id),
                    "pie_settings": {
                        "dividendCashAction": pie_summary.get("dividendCashAction"),
                        "endDate": pie_summary.get("endDate"),
                        "goal": pie_summary.get("goal"),
                        "icon": pie_summary.get("icon"),
                        "instrumentShares": pie_summary.get("instrumentShares"),
                        "name": pie_name,
                    },
                },
            )
            source["load_error"] = "Detail koláče se nepodařilo načíst z Trading 212 API."
            sources.append(source)
        return sources

    def _apply_trading212_source(self, source: dict, record_history: bool = True):
        self.portfolio_dict = self._coerce_holdings_dict(source.get("holdings", {}))
        self._prune_no_sell_tickers()
        self.extra_cash_var.set(f"{float(source.get('cash', 0.0)):.2f}")
        self.portfolio_broker_positions = source.get("broker_positions", {}) if isinstance(source.get("broker_positions"), dict) else {}
        self.portfolio_price_snapshot = source.get("price_snapshot", {}) if isinstance(source.get("price_snapshot"), dict) else {}
        self.portfolio_broker_link = source.get("broker_link", {}) if isinstance(source.get("broker_link"), dict) else {}
        self._repair_portfolio_symbols_from_broker_positions()
        self.portfolio_currency = str(source.get("currency") or "CZK")
        if record_history:
            self._append_current_equity_snapshot(replace_recent=True)
        self._refresh_portfolio_tree()
        invested = float(source.get("invested_value", 0.0))
        cash = float(source.get("cash", 0.0))
        self.log(
            f"Načteno z Trading 212: {source.get('label')} | "
            f"{len(self.portfolio_dict)} pozic, investováno {invested:,.2f} {self.portfolio_currency}, "
            f"cash {cash:,.2f} {self.portfolio_currency}."
        )
        if self.portfolio_currency.upper() != "CZK":
            self.log(
                f"Pozor: účet Trading 212 je v {self.portfolio_currency}. "
                "Portfolio manager částky zatím bere jako CZK, takže pro přesnost bude potřeba doplnit FX převod."
            )

    def _save_trading212_debug_snapshot(self, snapshot: dict, sources: list[dict]):
        try:
            debug_path = self.base_dir / "portfolios" / "trading212_last_snapshot_debug.json"
            debug_path.parent.mkdir(parents=True, exist_ok=True)
            payload = {
                "saved_at": dt.datetime.now().isoformat(),
                "snapshot": snapshot,
                "sources": sources,
            }
            with open(debug_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, ensure_ascii=False, default=str)
        except Exception as exc:
            self.log(f"Nepodařilo se uložit Trading 212 diagnostiku: {exc}")

    def _open_trading212_dialog(self):
        if self._blocked_action("Trading 212 import"):
            return
        config = self._load_trading212_config()
        top = tk.Toplevel(self.root)
        top.title("Trading 212 import")
        top.geometry("780x520")
        top.configure(padx=12, pady=12)

        configured_env = self._trading212_env_name(str(config.get("environment", "live")))
        configured_key, configured_secret = self._trading212_profile_credentials(config, configured_env)
        environment_var = tk.StringVar(value=configured_env)
        api_key_var = tk.StringVar(value=configured_key)
        api_secret_var = tk.StringVar(value=configured_secret)
        save_var = tk.BooleanVar(value=bool(configured_key and configured_secret))
        status_var = tk.StringVar(value="Zadej API klíče a načti zdroje.")
        source_var = tk.StringVar()
        sources_by_label: dict[str, dict] = {}

        form = ttk.LabelFrame(top, text="Připojení", padding=10)
        form.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(form, text="Prostředí").grid(row=0, column=0, sticky=tk.W, padx=(0, 6), pady=3)
        env_combo = ttk.Combobox(form, textvariable=environment_var, values=["live", "paper"], width=10, state="readonly")
        env_combo.grid(row=0, column=1, sticky=tk.W, pady=3)
        ttk.Label(form, text="paper = Trading 212 demo účet").grid(row=0, column=2, columnspan=2, sticky=tk.W, padx=(10, 0), pady=3)
        ttk.Label(form, text="API key").grid(row=1, column=0, sticky=tk.W, padx=(0, 6), pady=3)
        ttk.Entry(form, textvariable=api_key_var, width=58).grid(row=1, column=1, columnspan=3, sticky=tk.EW, pady=3)
        ttk.Label(form, text="API secret").grid(row=2, column=0, sticky=tk.W, padx=(0, 6), pady=3)
        ttk.Entry(form, textvariable=api_secret_var, width=58, show="*").grid(row=2, column=1, columnspan=3, sticky=tk.EW, pady=3)
        ttk.Checkbutton(form, text="Uložit lokálně do portfolios/trading212_config.json", variable=save_var).grid(row=3, column=1, columnspan=3, sticky=tk.W, pady=(4, 0))
        form.columnconfigure(3, weight=1)

        sources_frame = ttk.LabelFrame(top, text="Dostupné zdroje", padding=10)
        sources_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        columns = ("source", "positions", "invested", "cash", "currency")
        source_tree = ttk.Treeview(sources_frame, columns=columns, show="headings", height=9)
        headings = {
            "source": "Zdroj",
            "positions": "Pozic",
            "invested": "Investováno",
            "cash": "Cash",
            "currency": "Měna",
        }
        widths = {"source": 300, "positions": 70, "invested": 120, "cash": 120, "currency": 70}
        for col, label in headings.items():
            source_tree.heading(col, text=label)
            source_tree.column(col, width=widths[col], anchor=tk.CENTER if col != "source" else tk.W)
        source_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        source_scroll = ttk.Scrollbar(sources_frame, orient=tk.VERTICAL, command=source_tree.yview)
        source_tree.configure(yscrollcommand=source_scroll.set)
        source_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        selector = ttk.Frame(top)
        selector.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(selector, text="Importovat").pack(side=tk.LEFT)
        source_combo = ttk.Combobox(selector, textvariable=source_var, values=[], state="readonly", width=42)
        source_combo.pack(side=tk.LEFT, padx=(8, 0), fill=tk.X, expand=True)
        ttk.Label(top, textvariable=status_var).pack(fill=tk.X, pady=(0, 8))

        def apply_environment_profile(_event=None):
            env = self._trading212_env_name(environment_var.get())
            key, secret = self._trading212_profile_credentials(config, env)
            api_key_var.set(key)
            api_secret_var.set(secret)
            status_var.set(
                "Paper Trading vyžaduje API key vygenerovaný přímo v paper/demo účtu."
                if env == "paper" else
                "Live účet vyžaduje live API key."
            )

        env_combo.bind("<<ComboboxSelected>>", apply_environment_profile)

        def set_busy(active: bool, message: str):
            status_var.set(message)
            load_btn.configure(state=tk.DISABLED if active else tk.NORMAL)
            import_btn.configure(state=tk.DISABLED if active else tk.NORMAL)

        def current_client() -> Trading212Client:
            return Trading212Client(environment_var.get(), api_key_var.get(), api_secret_var.get())

        def load_sources():
            try:
                client = current_client()
            except Trading212ApiError as exc:
                messagebox.showerror("Trading 212", str(exc), parent=top)
                return
            set_busy(True, "Načítám Trading 212 portfolio a koláče...")

            def worker():
                try:
                    snapshot = client.snapshot()
                    sources = self._t212_sources_from_snapshot(snapshot)
                    self._save_trading212_debug_snapshot(snapshot, sources)
                    pie_error_count = len(snapshot.get("pie_errors", []))
                    self.log(
                        f"Trading 212: pozic {len(snapshot.get('positions', []))}, "
                        f"koláčů {len(snapshot.get('pies', []))}, "
                        f"detailů koláčů {len(snapshot.get('pie_details', []))}, "
                        f"chyb detailů {pie_error_count}."
                    )
                    if pie_error_count:
                        for pie_error in snapshot.get("pie_errors", [])[:5]:
                            self.log(f"Trading 212 detail koláče {pie_error.get('id')} selhal: {pie_error.get('error')}")
                    self.ui_queue.put(("call", finish_sources, (sources, None, pie_error_count)))
                except Exception as exc:
                    self.ui_queue.put(("call", finish_sources, ([], exc, 0)))

            threading.Thread(target=worker, daemon=True).start()

        def finish_sources(sources: list[dict], error, pie_error_count: int = 0):
            set_busy(False, "Zdroje načteny." if not error else "Načtení selhalo.")
            if error:
                messagebox.showerror("Trading 212", str(error), parent=top)
                return
            if save_var.get():
                try:
                    self._save_trading212_config(environment_var.get(), api_key_var.get(), api_secret_var.get())
                    profiles = config.setdefault("profiles", {})
                    profiles[environment_var.get()] = {
                        "api_key": api_key_var.get(),
                        "api_secret": api_secret_var.get(),
                    }
                    config["environment"] = environment_var.get()
                except Exception as exc:
                    self.log(f"Nepodařilo se uložit Trading 212 config: {exc}")
            sources_by_label.clear()
            for item in source_tree.get_children():
                source_tree.delete(item)
            labels = []
            for source in sources:
                label = str(source.get("label"))
                labels.append(label)
                sources_by_label[label] = source
                display_positions = source.get("position_count", 0)
                if source.get("load_error"):
                    display_positions = "?"
                source_tree.insert(
                    "",
                    tk.END,
                    values=(
                        label,
                        display_positions,
                        f"{float(source.get('invested_value', 0.0)):,.2f}",
                        f"{float(source.get('cash', 0.0)):,.2f}",
                        source.get("currency", ""),
                    ),
                )
            source_combo.configure(values=labels)
            if labels:
                source_var.set(labels[0])
                first_item = source_tree.get_children()[0]
                source_tree.selection_set(first_item)
                source_tree.focus(first_item)
            suffix = f" Detaily {pie_error_count} koláčů se nepodařilo načíst." if pie_error_count else ""
            status_var.set(f"Načteno {len(labels)} zdrojů. Vyber zdroj a importuj ho do portfolia.{suffix}")

        def select_source_from_tree(_event=None):
            selected = source_tree.selection()
            if not selected:
                return
            values = source_tree.item(selected[0], "values")
            if values:
                source_var.set(str(values[0]))

        def import_selected():
            if self._blocked_action("Trading 212 import"):
                return
            label = source_var.get()
            source = sources_by_label.get(label)
            if not source:
                messagebox.showinfo("Trading 212", "Nejprve načti zdroje a vyber jeden z nich.", parent=top)
                return
            if source.get("load_error"):
                messagebox.showerror(
                    "Trading 212",
                    "Tento koláč se v seznamu objevil, ale jeho detail se nepodařilo načíst. "
                    "Zkus znovu načíst zdroje; Trading 212 API bývá u pies omezené rate-limitem.",
                    parent=top,
                )
                return
            if not source.get("holdings") and float(source.get("cash", 0.0)) <= 0:
                messagebox.showinfo("Trading 212", "Vybraný zdroj neobsahuje žádné pozice ani cash.", parent=top)
                return
            if self.portfolio_dict and not messagebox.askyesno(
                "Nahradit portfolio?",
                "Import z Trading 212 nahradí aktuální portfolio v UI. Pokračovat?",
                parent=top,
            ):
                return
            link = source.get("broker_link") if isinstance(source.get("broker_link"), dict) else {}
            link["environment"] = self._trading212_env_name(environment_var.get())
            link.setdefault("cash_flow_checked_at", dt.datetime.now().isoformat())
            source["broker_link"] = link
            self._reset_portfolio_session_state(clear_portfolio_data=False)
            self._apply_trading212_source(source)
            messagebox.showinfo(
                "Trading 212",
                f"Importováno {len(self.portfolio_dict)} pozic ze zdroje:\n{label}",
                parent=top,
            )
            top.destroy()

        source_tree.bind("<<TreeviewSelect>>", select_source_from_tree)
        buttons = ttk.Frame(top)
        buttons.pack(fill=tk.X)
        ttk.Button(buttons, text="Zavřít", command=top.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        import_btn = ttk.Button(buttons, text="Importovat vybraný zdroj", command=import_selected)
        import_btn.pack(side=tk.RIGHT, padx=(8, 0))
        load_btn = ttk.Button(buttons, text="Načíst zdroje", command=load_sources)
        load_btn.pack(side=tk.RIGHT)

        if api_key_var.get() and api_secret_var.get():
            top.after(250, load_sources)
        
    @staticmethod
    def _positive_float(value):
        try:
            value = float(value)
        except (TypeError, ValueError):
            return None
        if not math.isfinite(value) or value <= 0:
            return None
        return value

    def _coerce_holdings_dict(self, raw_holdings) -> dict[str, float]:
        holdings: dict[str, float] = {}
        if not isinstance(raw_holdings, dict):
            return holdings
        for ticker, amount in raw_holdings.items():
            key = normalize_ticker_symbol(str(ticker))
            try:
                value = float(amount)
            except (TypeError, ValueError):
                continue
            if key and math.isfinite(value) and value > 0:
                holdings[key] = value
        return holdings

    def _repair_portfolio_symbols_from_broker_positions(self):
        if not isinstance(getattr(self, "portfolio_broker_positions", None), dict):
            return
        rename_map: dict[str, str] = {}
        for app_ticker, detail in list(self.portfolio_broker_positions.items()):
            if not isinstance(detail, dict):
                continue
            old_key = normalize_ticker_symbol(app_ticker)
            positions = detail.get("positions") if isinstance(detail.get("positions"), list) else []
            for position in positions:
                if not isinstance(position, dict):
                    continue
                resolved = resolve_instrument_ticker(
                    str(position.get("broker_ticker") or app_ticker),
                    name=str(position.get("name") or detail.get("name") or ""),
                    isin=str(position.get("isin") or ""),
                )
                if resolved and resolved != old_key:
                    rename_map[old_key] = resolved
                    break
        if not rename_map:
            return

        for old_key, new_key in rename_map.items():
            if old_key in self.portfolio_dict:
                self.portfolio_dict[new_key] = self.portfolio_dict.get(new_key, 0.0) + float(self.portfolio_dict.pop(old_key))
            detail = self.portfolio_broker_positions.pop(old_key, None)
            if detail is not None:
                existing = self.portfolio_broker_positions.get(new_key)
                if isinstance(existing, dict) and isinstance(detail, dict):
                    existing_positions = existing.setdefault("positions", [])
                    if isinstance(existing_positions, list):
                        existing_positions.extend(detail.get("positions", []) if isinstance(detail.get("positions"), list) else [])
                    existing["current_value"] = float(existing.get("current_value", 0.0) or 0.0) + float(detail.get("current_value", 0.0) or 0.0)
                    existing["quantity"] = float(existing.get("quantity", 0.0) or 0.0) + float(detail.get("quantity", 0.0) or 0.0)
                else:
                    self.portfolio_broker_positions[new_key] = detail
            snapshot = self.portfolio_price_snapshot.pop(old_key, None) if isinstance(self.portfolio_price_snapshot, dict) else None
            broker_detail = self.portfolio_broker_positions.get(new_key) if isinstance(self.portfolio_broker_positions, dict) else None
            if isinstance(broker_detail, dict):
                positions = broker_detail.get("positions") if isinstance(broker_detail.get("positions"), list) else []
                price = next(
                    (
                        self._positive_float(position.get("current_price"))
                        for position in positions
                        if isinstance(position, dict) and self._positive_float(position.get("current_price"))
                    ),
                    None,
                )
                if price:
                    self.portfolio_price_snapshot[new_key] = {
                        "price": float(price),
                        "timestamp": dt.datetime.now().isoformat(),
                        "broker": broker_detail.get("broker"),
                    }
                elif snapshot is not None:
                    self.portfolio_price_snapshot[new_key] = snapshot
            elif snapshot is not None:
                self.portfolio_price_snapshot[new_key] = snapshot
            if old_key in self.no_sell_tickers:
                self.no_sell_tickers.discard(old_key)
                self.no_sell_tickers.add(new_key)
            self.log(f"Opraven ticker podle broker ISIN/názvu: {old_key} -> {new_key}.")

    def _open_degiro_dialog(self):
        if self._blocked_action("Degiro import"):
            return
        file_path = filedialog.askopenfilename(
            title="Vyber Degiro export portfolio (CSV)",
            filetypes=[("CSV soubory", "*.csv"), ("Všechny soubory", "*.*")],
            parent=self.root
        )
        if not file_path:
            return

        import csv
        import unicodedata

        def clean_key(k: str) -> str:
            cleaned = unicodedata.normalize('NFKD', k).encode('ASCII', 'ignore').decode('ASCII')
            cleaned = re.sub(r'[^a-z0-9]', '', cleaned.lower())
            return cleaned

        def parse_czech_float(val_str: str) -> float:
            cleaned = str(val_str).strip()
            cleaned = cleaned.replace(" ", "").replace("\xa0", "")
            if "," in cleaned and "." in cleaned:
                cleaned = cleaned.replace(",", "")
            elif "," in cleaned:
                cleaned = cleaned.replace(",", ".")
            try:
                return float(cleaned)
            except ValueError:
                return 0.0

        encodings = ["utf-8-sig", "utf-8", "cp1250", "latin-1"]
        header = []
        rows = []
        parsed_ok = False
        error_msg = ""

        for encoding in encodings:
            try:
                with open(file_path, "r", encoding=encoding, errors="ignore") as f:
                    first_line = f.readline()
                if not first_line:
                    continue
                separator = ","
                if ";" in first_line:
                    separator = ";"
                elif "\t" in first_line:
                    separator = "\t"

                with open(file_path, "r", encoding=encoding, errors="ignore") as f:
                    reader = csv.reader(f, delimiter=separator)
                    header = []
                    rows = []
                    for r in reader:
                        if not r:
                            continue
                        if not header:
                            header = [col.strip() for col in r]
                        else:
                            if len(r) < len(header):
                                r = r + [""] * (len(header) - len(r))
                            rows.append(dict(zip(header, r)))
                if header and rows:
                    parsed_ok = True
                    break
            except Exception as e:
                error_msg = str(e)

        if not parsed_ok or not header or not rows:
            messagebox.showerror("Chyba čtení", f"Nepodařilo se správně načíst CSV soubor:\n{error_msg or 'Soubor je prázdný.'}", parent=self.root)
            return

        clean_headers = [clean_key(col) for col in header]

        col_isin = None
        col_symbol = None
        col_name = None
        col_quantity = None
        col_price = None
        col_value_czk = None
        col_value_any = None
        col_currency = None

        for idx, h in enumerate(clean_headers):
            raw_h = header[idx]
            if "isin" in h or "symbol" in h or "ticker" in h or h == "kod" or h == "symbolisin":
                if not col_isin:
                    col_isin = raw_h
                if not col_symbol:
                    col_symbol = raw_h
            elif "produkt" in h or "product" in h or "nazev" in h or "name" in h or "instrument" in h:
                col_name = raw_h
            elif "pocet" in h or "mnozstvi" in h or "kusy" in h or "quantity" in h or "shares" in h or "aantal" in h or "anzahl" in h:
                col_quantity = raw_h
            elif "price" in h or "cena" in h or "koers" in h or "kurs" in h or "uzaviraci" in h or "zaviraci" in h:
                col_price = raw_h
            
            if "czk" in h:
                col_value_czk = raw_h
            elif "eur" in h or "usd" in h or "value" in h or "hodnota" in h or "waarde" in h or "wert" in h or "curval" in h:
                if raw_h and not col_value_any:
                    col_value_any = raw_h
            if "mena" in h or "currency" in h:
                col_currency = raw_h

        # Dynamic currency & value discovery based on actual row values (e.g. Degiro's double Hodnota columns)
        if rows:
            first_row = rows[0]
            for idx, col_name_in_h in enumerate(header):
                val = str(first_row.get(col_name_in_h, "")).strip()
                if len(val) == 3 and val.isupper() and val in ["USD", "EUR", "CZK", "GBP", "CHF"]:
                    col_currency = col_name_in_h
                    if idx + 1 < len(header):
                        col_value_any = header[idx + 1]
                    break

        if not col_name and len(header) > 0:
            col_name = header[0]

        eur_rate = 25.0
        usd_rate = 23.0
        try:
            eur_cache = self.fetcher.cache.get("EURCZK=X", {})
            if eur_cache.get("price"):
                eur_rate = float(eur_cache["price"])
            usd_cache = self.fetcher.cache.get("USDCZK=X", {})
            if usd_cache.get("price"):
                usd_rate = float(usd_cache["price"])
        except Exception:
            pass

        holdings = {}
        broker_positions = {}
        price_snapshot = {}
        total_cash = 0.0
        now = dt.datetime.now().isoformat()

        for row in rows:
            isin = str(row.get(col_isin, "")).strip() if col_isin else ""
            symbol = str(row.get(col_symbol, "")).strip() if col_symbol else ""
            name = str(row.get(col_name, "")).strip() if col_name else ""
            
            quantity = parse_czech_float(row.get(col_quantity, 0)) if col_quantity else 0.0
            price = parse_czech_float(row.get(col_price, 0)) if col_price else 0.0

            # Detect cash row
            is_cash = False
            if not isin and not symbol:
                name_upper = name.upper()
                name_clean = clean_key(name)
                if "penize" in name_clean or "cash" in name_clean or "eur" in name_upper or "czk" in name_upper or "usd" in name_upper:
                    is_cash = True
            elif isin and not symbol and "CASH" in isin.upper():
                is_cash = True
            
            if is_cash:
                cash_val = parse_czech_float(row.get(col_value_czk) or row.get(col_value_any) or row.get(col_price) or 0.0)
                cur = str(row.get(col_currency, "")).upper() if col_currency else ""
                if "EUR" in cur or (col_value_any and "EUR" in str(row.get(col_value_any)).upper()):
                    cash_val *= eur_rate
                elif "USD" in cur or (col_value_any and "USD" in str(row.get(col_value_any)).upper()):
                    cash_val *= usd_rate
                total_cash += cash_val
                continue

            if not isin and not symbol and not name:
                continue

            app_ticker = resolve_instrument_ticker(symbol, name=name, isin=isin)
            if not app_ticker or app_ticker.upper() in ["EUR", "USD", "CZK", "CASH", "PENIZE"]:
                continue

            val_czk = 0.0
            if col_value_czk:
                val_czk = parse_czech_float(row.get(col_value_czk, 0))
            else:
                val = 0.0
                cur = ""
                if col_value_any:
                    val = parse_czech_float(row.get(col_value_any, 0))
                    cur = str(row.get(col_currency, "")).upper() if col_currency else ""
                
                # Secondary check: search for an explicit EUR column in headers if val is empty
                if val <= 0:
                    for idx, h in enumerate(clean_headers):
                        if "eur" in h and header[idx] != col_currency:
                            val = parse_czech_float(row.get(header[idx], 0))
                            cur = "EUR"
                            break

                if val <= 0:
                    val = quantity * price
                    cur = str(row.get(col_currency, "")).upper() if col_currency else ""

                if "EUR" in cur or (col_value_any and "EUR" in str(col_value_any).upper()):
                    val_czk = val * eur_rate
                elif "USD" in cur or (col_value_any and "USD" in str(col_value_any).upper()):
                    val_czk = val * usd_rate
                else:
                    val_czk = val

            if val_czk <= 0 and quantity > 0:
                val_czk = quantity * (price or 1.0)

            if val_czk <= 0:
                continue

            holdings[app_ticker] = holdings.get(app_ticker, 0.0) + val_czk
            
            detail = broker_positions.setdefault(
                app_ticker,
                {
                    "broker": "Degiro",
                    "source_id": "degiro_csv",
                    "source": f"Degiro ({Path(file_path).name})",
                    "currency": "CZK",
                    "quantity": 0.0,
                    "current_value": 0.0,
                    "positions": [],
                },
            )
            detail["quantity"] += quantity
            detail["current_value"] += val_czk
            detail["positions"].append({
                "broker_ticker": symbol or app_ticker,
                "quantity": quantity,
                "current_price": price,
                "current_value": val_czk,
                "isin": isin,
                "name": name,
            })
            if price > 0:
                price_snapshot[app_ticker] = {
                    "price": price,
                    "timestamp": now,
                    "broker": "Degiro",
                    "broker_ticker": symbol or app_ticker,
                }

        if not holdings:
            messagebox.showwarning(
                "Degiro import",
                "V CSV souboru nebyly nalezeny žádné akciové pozice se známým symbolem nebo ISIN kódem.",
                parent=self.root
            )
            return

        if self.portfolio_dict and not messagebox.askyesno(
            "Nahradit portfolio?",
            f"Načteno {len(holdings)} pozic z Degiro exportu (celková hodnota pozic {sum(holdings.values()):,.2f} CZK, hotovost {total_cash:,.2f} CZK).\n\n"
            "Chcete tímto exportem nahradit aktuální portfolio?",
            parent=self.root
        ):
            return

        source = {
            "id": "degiro_csv",
            "label": f"Degiro ({Path(file_path).name})",
            "holdings": holdings,
            "cash": total_cash,
            "currency": "CZK",
            "broker_positions": broker_positions,
            "broker_link": {
                "broker": "Degiro",
                "source_id": "degiro_csv",
                "source_label": f"Degiro ({Path(file_path).name})",
                "type": "portfolio",
            },
            "price_snapshot": price_snapshot,
            "position_count": len(holdings),
            "invested_value": sum(holdings.values()),
        }

        self._reset_portfolio_session_state(clear_portfolio_data=False)
        self._apply_degiro_source(source)
        
        messagebox.showinfo(
            "Degiro import",
            f"Úspěšně importováno {len(holdings)} pozic a {total_cash:,.2f} CZK hotovosti z Degiro exportu.",
            parent=self.root
        )

    def _apply_degiro_source(self, source: dict):
        self.portfolio_dict = self._coerce_holdings_dict(source.get("holdings", {}))
        self._prune_no_sell_tickers()
        self.extra_cash_var.set(f"{float(source.get('cash', 0.0)):.2f}")
        self.portfolio_broker_positions = source.get("broker_positions", {}) if isinstance(source.get("broker_positions"), dict) else {}
        self.portfolio_price_snapshot = source.get("price_snapshot", {}) if isinstance(source.get("price_snapshot"), dict) else {}
        self.portfolio_broker_link = source.get("broker_link", {}) if isinstance(source.get("broker_link"), dict) else {}
        self._repair_portfolio_symbols_from_broker_positions()
        self.portfolio_currency = str(source.get("currency") or "CZK")
        self._append_current_equity_snapshot(replace_recent=True)
        self._refresh_portfolio_tree()
        invested = float(source.get("invested_value", 0.0))
        cash = float(source.get("cash", 0.0))
        self.log(
            f"Načteno z Degiro: {source.get('label')} | "
            f"{len(self.portfolio_dict)} pozic, investováno {invested:,.2f} {self.portfolio_currency}, "
            f"cash {cash:,.2f} {self.portfolio_currency}."
        )

    def _snapshot_price(self, snapshot_entry):
        if isinstance(snapshot_entry, dict):
            return self._positive_float(snapshot_entry.get("price"))
        return self._positive_float(snapshot_entry)

    def _current_price_snapshot(self) -> dict[str, dict[str, float | str]]:
        import datetime
        now = datetime.datetime.now().isoformat()
        snapshot: dict[str, dict[str, float | str]] = {}
        existing = getattr(self, "portfolio_price_snapshot", {}) or {}
        for ticker in self.portfolio_dict.keys():
            key = normalize_ticker_symbol(ticker)
            cache_price = self._positive_float(self.fetcher.cache.get(key, {}).get("price"))
            old_entry = existing.get(key)
            old_price = self._snapshot_price(old_entry)
            price = cache_price or old_price
            if price:
                entry: dict[str, float | str] = {"price": float(price), "timestamp": now}
                if isinstance(old_entry, dict):
                    for field in ("broker", "broker_ticker"):
                        if old_entry.get(field):
                            entry[field] = str(old_entry[field])
                snapshot[key] = entry
        return snapshot

    def _ticker_has_market_evidence(self, ticker: str, metadata: dict | None = None) -> bool:
        key = normalize_ticker_symbol(ticker)
        meta = metadata.get(key, {}) if isinstance(metadata, dict) else {}
        if self._positive_float(meta.get("price")):
            return True
        fetcher = getattr(self, "fetcher", None)
        cache = getattr(fetcher, "cache", {}) if fetcher is not None else {}
        cached = cache.get(key, {}) if isinstance(cache, dict) else {}
        if isinstance(cached, dict) and self._positive_float(cached.get("price")):
            return True
        snapshot = getattr(self, "portfolio_price_snapshot", {}) or {}
        if self._snapshot_price(snapshot.get(key)):
            return True
        broker = getattr(self, "portfolio_broker_positions", {}) or {}
        detail = broker.get(key, {}) if isinstance(broker, dict) else {}
        if isinstance(detail, dict) and detail.get("positions"):
            return True
        return False

    def _write_portfolio_file(self, quiet: bool = False):
        if not self.current_portfolio_path:
            return
        self.current_portfolio_path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.current_portfolio_path, "w", encoding="utf-8") as f:
            json.dump(self._export_portfolio_json(), f, indent=4)
        if not quiet:
            self.log(f"Portfolio uloženo: {self.current_portfolio_path.name}")

    def _is_custom_projection(self, proj) -> bool:
        if not isinstance(proj, dict):
            return True
        source = str(proj.get("source", "")).lower()
        label = str(proj.get("label", "")).lower()
        return source == "custom" or "custom portfolio" in label

    def _export_portfolio_json(self):
        import datetime
        self.portfolio_price_snapshot = self._current_price_snapshot()
        return {
            "version": 3,
            "holdings": self.portfolio_dict,
            "extra_cash": float(self.extra_cash_var.get() or 0),
            "dividends": self.total_dividends,
            "history": self.portfolio_history,
            "projections": self.portfolio_projections,
            "price_snapshot": self.portfolio_price_snapshot,
            "broker_positions": self.portfolio_broker_positions,
            "broker_link": self.portfolio_broker_link,
            "currency": self.portfolio_currency,
            "no_sell_tickers": sorted(self._active_no_sell_tickers()),
            "last_saved": datetime.datetime.now().isoformat()
        }

    def _import_portfolio_json(self, data, refresh_market: bool = False):
        self.portfolio_dict = self._coerce_holdings_dict(data.get("holdings", {}))
        self.extra_cash_var.set(str(data.get("extra_cash", 0.0)))
        self.total_dividends = float(data.get("dividends", 0.0) or 0.0)
        self.portfolio_history = data.get("history", []) if isinstance(data.get("history", []), list) else []
        raw_projections = data.get("projections", [])
        loaded_projections = [
            proj for proj in raw_projections
            if isinstance(proj, dict) and not self._is_custom_projection(proj)
        ] if isinstance(raw_projections, list) else []
        self.portfolio_projections = self._chart_projections(loaded_projections)
        raw_snapshot = data.get("price_snapshot", data.get("portfolio_price_snapshot", {}))
        self.portfolio_price_snapshot = raw_snapshot if isinstance(raw_snapshot, dict) else {}
        raw_broker_positions = data.get("broker_positions", {})
        self.portfolio_broker_positions = raw_broker_positions if isinstance(raw_broker_positions, dict) else {}
        raw_broker_link = data.get("broker_link", {})
        self.portfolio_broker_link = raw_broker_link if isinstance(raw_broker_link, dict) else {}
        if self.portfolio_broker_link.get("broker") == "Trading 212" and self.portfolio_broker_link.get("type") == "pie":
            settings = dict(self.portfolio_broker_link.get("pie_settings") or {})
            raw_name = settings.get("name") or str(self.portfolio_broker_link.get("source_label") or "").replace("Koláč: ", "")
            clean_name = self._clean_temporary_t212_pie_name(raw_name)
            settings["name"] = clean_name
            self.portfolio_broker_link["pie_settings"] = settings
            self.portfolio_broker_link["source_label"] = f"Koláč: {clean_name}"
        raw_no_sell = data.get("no_sell_tickers", data.get("protected_tickers", []))
        if isinstance(raw_no_sell, (list, tuple, set)):
            self.no_sell_tickers = {normalize_ticker_symbol(ticker) for ticker in raw_no_sell if normalize_ticker_symbol(ticker)}
        else:
            self.no_sell_tickers = set()
        self._repair_portfolio_symbols_from_broker_positions()
        self._prune_no_sell_tickers()
        self.portfolio_currency = str(data.get("currency", "CZK") or "CZK")
        enriched_history = self._enrich_legacy_history_snapshots()
        if enriched_history:
            self.log(f"Doplněny historické kotvy portfolia pro {enriched_history} starších záznamů.")
        self.div_var.set(f"Statistika | Celkové obdržené dividendy: {self.total_dividends:.2f} CZK")
        self._refresh_portfolio_tree(revalue_market=refresh_market)
        self._autoset_max_weight_from_holdings()

    def _autoset_max_weight_from_holdings(self):
        if not self.portfolio_dict:
            return
        total_val = sum(self.portfolio_dict.values())
        if total_val > 0:
            max_val = max(self.portfolio_dict.values())
            max_weight_pct = (max_val / total_val) * 100.0
            if abs(max_weight_pct - round(max_weight_pct)) < 0.05:
                self.max_weight_var.set(f"{round(max_weight_pct)}")
            else:
                self.max_weight_var.set(f"{max_weight_pct:.1f}")
            self.log(f"Automaticky nastavuji max váhu akcie na {self.max_weight_var.get()}% (podle největší pozice).")

    def _save_portfolio(self):
        if self._blocked_action("Uložení portfolia"):
            return
        if self.current_portfolio_path:
            from tkinter import messagebox
            try:
                self._write_portfolio_file(quiet=False)
            except Exception as e:
                messagebox.showerror("Chyba", str(e))
        else:
            self._save_portfolio_as()

    def _save_portfolio_as(self):
        if self._blocked_action("Uložení portfolia"):
            return
        portfolios_dir = self.base_dir / "portfolios"
        portfolios_dir.mkdir(parents=True, exist_ok=True)
        from tkinter import filedialog
        filepath = filedialog.asksaveasfilename(
            title="Uložit portfolio jako...",
            initialdir=str(portfolios_dir),
            defaultextension=".json",
            filetypes=[("JSON soubory", "*.json")],
        )
        if not filepath:
            return
        self.current_portfolio_path = Path(filepath)
        self._save_portfolio()

    def _load_portfolio_from_file(self):
        if self._blocked_action("Načtení portfolia"):
            return
        portfolios_dir = self.base_dir / "portfolios"
        portfolios_dir.mkdir(parents=True, exist_ok=True)
        from tkinter import filedialog
        filepath = filedialog.askopenfilename(
            title="Načíst portfolio ze souboru...",
            initialdir=str(portfolios_dir),
            filetypes=[("JSON soubory", "*.json"), ("Textové soubory", "*.txt")],
        )
        if not filepath:
            return
        self._load_portfolio(Path(filepath))

    def _load_portfolio(self, path=None, quiet=False):
        if not path:
            path = self.base_dir / "portfolios" / "default.json"
        
        if not path.exists():
            return

        import json
        from tkinter import messagebox
        previous_path = self.current_portfolio_path
        try:
            with open(path, "r", encoding="utf-8") as f:
                self.current_portfolio_path = path
                if path.suffix == ".json":
                    data = json.load(f)
                    if not isinstance(data, dict):
                        raise ValueError("JSON soubor neobsahuje portfolio.")
                    self._reset_portfolio_session_state(clear_portfolio_data=False)
                    self.current_portfolio_path = path
                    broker_link = data.get("broker_link", {})
                    is_trading212 = isinstance(broker_link, dict) and broker_link.get("broker") == "Trading 212"
                    self._import_portfolio_json(data, refresh_market=not is_trading212)
                else: 
                    p_text = f.read().strip()
                    self._reset_portfolio_session_state(clear_portfolio_data=True)
                    self.current_portfolio_path = path
                    self.portfolio_dict = self._coerce_holdings_dict(parse_holdings_text(p_text))
                    self.portfolio_price_snapshot = {}
                    self.portfolio_broker_positions = {}
                    self.portfolio_broker_link = {}
                    self.no_sell_tickers = set()
                    self._refresh_portfolio_tree(revalue_market=True)
                    self._autoset_max_weight_from_holdings()
            if not quiet:
                self.log(f"Portfolio úspěšně načteno: {path.name}")
            self._refresh_trading212_linked_portfolio_async("Načtené portfolio je napojené na Trading 212, obnovuji aktuální stav...")
        except Exception as e:
            self.current_portfolio_path = previous_path
            if not quiet:
                messagebox.showerror("Chyba čtení", str(e))

    def get_holdings_text(self):
        return "\n".join([f"{k} {v}" for k, v in self.portfolio_dict.items()])

    def _active_no_sell_tickers(self) -> set[str]:
        current = {normalize_ticker_symbol(ticker) for ticker in self.portfolio_dict.keys()}
        return {ticker for ticker in getattr(self, "no_sell_tickers", set()) if ticker in current}

    def _prune_no_sell_tickers(self):
        self.no_sell_tickers = self._active_no_sell_tickers()

    def _update_no_sell_row_marker(self, ticker: str) -> bool:
        ticker = normalize_ticker_symbol(ticker)
        columns = list(self.port_tree["columns"])
        try:
            no_sell_index = columns.index("nosell")
        except ValueError:
            return False
        updated = False
        for item in self.port_tree.get_children():
            values = list(self.port_tree.item(item, "values"))
            if not values or normalize_ticker_symbol(str(values[0])) != ticker:
                continue
            values[no_sell_index] = "☑" if ticker in self.no_sell_tickers else "☐"
            tags = ("no_sell",) if ticker in self.no_sell_tickers else ()
            self.port_tree.item(item, values=values, tags=tags)
            updated = True
        return updated

    def _toggle_no_sell_ticker(self, ticker: str):
        if self._blocked_action("Přepnutí ochrany před prodejem"):
            return
        ticker = normalize_ticker_symbol(ticker)
        if not ticker or ticker not in self.portfolio_dict:
            return
        if ticker in self.no_sell_tickers:
            self.no_sell_tickers.remove(ticker)
            self.log(f"{ticker}: ochrana před prodejem vypnuta.")
        else:
            self.no_sell_tickers.add(ticker)
            self.log(f"{ticker}: ochrana před prodejem zapnuta.")
        if not self._update_no_sell_row_marker(ticker):
            self._refresh_portfolio_tree()
        if self.current_portfolio_path:
            try:
                self._write_portfolio_file(quiet=True)
            except Exception as exc:
                self.log(f"Ochrana byla změněna, ale portfolio se nepodařilo uložit: {exc}")

    def _toggle_selected_no_sell(self):
        if self._blocked_action("Přepnutí ochrany před prodejem"):
            return
        selected = self.port_tree.selection()
        ticker = normalize_ticker_symbol(self.add_ticker_var.get())
        if selected:
            values = self.port_tree.item(selected[0], "values")
            if values:
                ticker = normalize_ticker_symbol(str(values[0]))
        if not ticker:
            return
        self._toggle_no_sell_ticker(ticker)

    def _on_portfolio_tree_click(self, event):
        try:
            if self._is_ui_locked():
                return "break"
            if self.port_tree.identify("region", event.x, event.y) != "cell":
                return None
            item = self.port_tree.identify_row(event.y)
            column = self.port_tree.identify_column(event.x)
            if not item or not column:
                return None
            col_index = int(column.replace("#", "")) - 1
            columns = list(self.port_tree["columns"])
            if col_index < 0 or col_index >= len(columns) or columns[col_index] != "nosell":
                return None
            values = self.port_tree.item(item, "values")
            if values:
                self._toggle_no_sell_ticker(str(values[0]))
                return "break"
        except Exception as exc:
            self.log(f"Nepodařilo se přepnout ochranu před prodejem: {exc}")
        return None

    def _is_degiro_linked_portfolio(self) -> bool:
        link = self.portfolio_broker_link if isinstance(self.portfolio_broker_link, dict) else {}
        if link.get("broker") == "Degiro":
            return True
        broker_positions = self.portfolio_broker_positions if isinstance(self.portfolio_broker_positions, dict) else {}
        return any(
            isinstance(detail, dict) and detail.get("broker") == "Degiro"
            for detail in broker_positions.values()
        )

    def _revalue_detail_from_quantity(
        self,
        detail: dict,
        latest_price: float,
        fallback_old_price: float | None = None,
    ) -> float | None:
        if not isinstance(detail, dict) or detail.get("broker") != "Degiro":
            return None
        latest_price = float(latest_price)
        if latest_price <= 0 or not math.isfinite(latest_price):
            return None

        raw_positions = detail.get("positions", [])
        positions = raw_positions if isinstance(raw_positions, list) and raw_positions else [detail]
        aggregate_value = self._positive_float(detail.get("current_value"))
        aggregate_quantity = self._positive_float(detail.get("quantity"))
        new_total = 0.0
        used_positions = 0

        for position in positions:
            if not isinstance(position, dict):
                continue
            quantity = self._positive_float(position.get("quantity"))
            if not quantity:
                continue
            old_price = self._positive_float(position.get("current_price")) or fallback_old_price
            old_value = self._positive_float(position.get("current_value"))
            if not old_value and aggregate_value and aggregate_quantity:
                old_value = aggregate_value * (float(quantity) / float(aggregate_quantity))
            if not old_price or not old_value:
                continue

            fx_factor = float(old_value) / (float(quantity) * float(old_price))
            if not math.isfinite(fx_factor) or fx_factor <= 0 or fx_factor > 1000:
                continue
            new_value = float(quantity) * latest_price * fx_factor
            if not math.isfinite(new_value) or new_value <= 0:
                continue

            position["current_price"] = latest_price
            position["current_value"] = float(new_value)
            new_total += float(new_value)
            used_positions += 1

        if not used_positions:
            return None
        detail["current_value"] = float(new_total)
        return float(new_total)

    def _revalue_portfolio_from_market(self, metadata: dict[str, dict]) -> bool:
        self.portfolio_dict = self._coerce_holdings_dict(self.portfolio_dict)
        if not self.portfolio_dict:
            self.portfolio_price_snapshot = {}
            return False
        import datetime
        before_total = sum(float(v) for v in self.portfolio_dict.values())
        previous_snapshot = getattr(self, "portfolio_price_snapshot", {}) or {}
        broker_positions = self.portfolio_broker_positions if isinstance(self.portfolio_broker_positions, dict) else {}
        new_snapshot: dict[str, dict[str, float | str]] = {}
        changed_count = 0
        skipped_count = 0
        now = datetime.datetime.now().isoformat()

        for ticker, amount in list(self.portfolio_dict.items()):
            key = normalize_ticker_symbol(ticker)
            meta = metadata.get(key, metadata.get(ticker, {})) or {}
            latest_price = self._positive_float(meta.get("price"))
            old_price = self._snapshot_price(previous_snapshot.get(key, previous_snapshot.get(ticker)))
            if old_price is None:
                old_price = self._positive_float(meta.get("_previous_price"))

            if latest_price:
                snapshot_meta = previous_snapshot.get(key, previous_snapshot.get(ticker))
                broker_ticker = ""
                broker_name = ""
                if isinstance(snapshot_meta, dict):
                    broker_ticker = str(snapshot_meta.get("broker_ticker") or "")
                    broker_name = str(snapshot_meta.get("broker") or "")
                detail = broker_positions.get(key, broker_positions.get(ticker, {}))
                if isinstance(detail, dict):
                    broker_name = broker_name or str(detail.get("broker") or "")
                    positions = detail.get("positions", [])
                    if isinstance(positions, list) and positions and isinstance(positions[0], dict):
                        broker_ticker = broker_ticker or str(positions[0].get("broker_ticker") or "")
                new_snapshot[key] = {
                    "price": float(latest_price),
                    "timestamp": now,
                    **({"broker": broker_name} if broker_name else {}),
                    **({"broker_ticker": broker_ticker} if broker_ticker else {}),
                }

            if not latest_price:
                skipped_count += 1
                continue

            detail = broker_positions.get(key, broker_positions.get(ticker, {}))
            new_amount = self._revalue_detail_from_quantity(detail, latest_price, old_price)
            if new_amount is None:
                if not old_price:
                    skipped_count += 1
                    continue
                ratio = latest_price / old_price
                if ratio <= 0 or ratio < 0.05 or ratio > 20:
                    skipped_count += 1
                    continue
                new_amount = float(amount) * ratio

            if abs(new_amount - float(amount)) >= 0.5:
                if key != ticker:
                    self.portfolio_dict.pop(ticker, None)
                    if ticker in self.no_sell_tickers:
                        self.no_sell_tickers.discard(ticker)
                        self.no_sell_tickers.add(key)
                self.portfolio_dict[key] = new_amount
                changed_count += 1

        self.portfolio_price_snapshot = new_snapshot
        after_total = sum(float(v) for v in self.portfolio_dict.values())
        if changed_count:
            delta = after_total - before_total
            pct = (delta / before_total * 100.0) if before_total > 0 else 0.0
            self._append_current_equity_snapshot(replace_recent=True, replace_same_day=True)
            try:
                self._write_portfolio_file(quiet=True)
            except Exception as exc:
                self.log(f"Přecenění proběhlo, ale portfolio se nepodařilo uložit: {exc}")
            self.log(
                f"Portfolio přeceněno podle trhu: {changed_count} pozic, "
                f"{before_total:,.0f} -> {after_total:,.0f} CZK ({pct:+.2f} %)."
            )
            return True

        try:
            self._write_portfolio_file(quiet=True)
        except Exception:
            pass
        if skipped_count:
            self.log("Portfolio načteno; pro část pozic chyběla cenová kotva, takže se jen uložil aktuální snapshot.")
        return False

    def _fallback_record_from_data(self, ticker: str, fallback: dict) -> StockRecord | None:
        if not ENABLE_HOLDING_FALLBACK_PREDICTIONS:
            return None
        if not isinstance(fallback, dict) or not fallback:
            return None
        if fallback.get("not_found") or fallback.get("no_prediction"):
            return None
        try:
            pred = float(fallback.get("pred", 0.0))
            cvar10 = float(fallback.get("cvar10", -float(fallback.get("q10", 0.0))))
            downside = max(1e-6, _downside_tail_risk(cvar10))
            upside = float(fallback.get("q90", pred))
            record = StockRecord(
                ticker=normalize_ticker_symbol(ticker),
                sharpe=float(_starr_score(pred, downside)),
                forecast_pct=pred,
                std_pct=downside,
                upside_pct=upside,
                is_synthetic=True,
                horizon_data=fallback.get("horizon_data") if isinstance(fallback.get("horizon_data"), dict) else None,
                cvar10_pct=cvar10,
            )
            record.score = _record_objective_score(record, self._current_objective_config())
            return record
        except Exception:
            return None

    def _refresh_portfolio_tree(self, revalue_market: bool = False):
        self.portfolio_dict = self._coerce_holdings_dict(self.portfolio_dict)
        self._prune_no_sell_tickers()
        self._portfolio_refresh_seq += 1
        refresh_seq = self._portfolio_refresh_seq
        for item in self.port_tree.get_children():
            self.port_tree.delete(item)
        records_by_ticker = {normalize_ticker_symbol(record.ticker): record for record in self.records}
        missing_for_ema = [
            normalize_ticker_symbol(t)
            for t in self.portfolio_dict.keys()
            if normalize_ticker_symbol(t) not in records_by_ticker
        ]
        if missing_for_ema:
            try:
                for record in load_holding_ema_fallback_records(
                    self.base_dir,
                    missing_for_ema,
                    self._current_objective_config(),
                    lambda _msg: None,
                ):
                    records_by_ticker[normalize_ticker_symbol(record.ticker)] = record
            except Exception:
                pass
        if ENABLE_HOLDING_FALLBACK_PREDICTIONS:
            missing_fallback = [
                normalize_ticker_symbol(t)
                for t in self.portfolio_dict.keys()
                if normalize_ticker_symbol(t) not in records_by_ticker
                and normalize_ticker_symbol(t) not in self.fallback_prediction_cache
                and normalize_ticker_symbol(t) not in self.fallback_prediction_inflight
                and self.fallback_prediction_failures.get(normalize_ticker_symbol(t), 0) < self.fallback_max_retries
            ]
            if missing_fallback:
                by_scale: dict[float, list[str]] = {}
                for t in missing_fallback:
                    scale = float(self.fallback_uncertainty_by_ticker.get(t, 1.0))
                    by_scale.setdefault(scale, []).append(t)
                for scale, tickers in by_scale.items():
                    self._load_fallback_predictions_async(tickers, scale)
        
        def update_rows(metadata):
            if refresh_seq != self._portfolio_refresh_seq:
                return
            for item in self.port_tree.get_children():
                self.port_tree.delete(item)
            normalized_metadata = {normalize_ticker_symbol(k): v for k, v in (metadata or {}).items()}
            if revalue_market:
                self._revalue_portfolio_from_market(normalized_metadata)
            total_value = max(sum(self.portfolio_dict.values()), 1e-9)
            sorted_positions = sorted(self.portfolio_dict.items(), key=lambda x: x[1], reverse=True)
            for t, amt in sorted_positions:
                ticker = normalize_ticker_symbol(t)
                meta = normalized_metadata.get(ticker, {"name": ticker, "sector": "N/A", "price": 0.0, "mcap": "0"})
                record = records_by_ticker.get(ticker)
                fallback = self.fallback_prediction_cache.get(ticker)
                weight = (amt / total_value) * 100.0
                if record:
                    pred_val, q10_val, q90_val = _record_like_horizon_values(record)
                    source = "AI (starší)" if record.is_synthetic else "AI"
                    score = _format_score(record.score)
                    pred = _format_return_pct(pred_val)
                    q10 = _format_return_pct(q10_val)
                    q90 = _format_return_pct(q90_val)
                else:
                    if not ENABLE_HOLDING_FALLBACK_PREDICTIONS:
                        score, pred, q10, q90, source = "-", "-", "-", "-", "Bez predikce"
                    elif fallback and fallback.get("not_found"):
                        if self._ticker_has_market_evidence(ticker, normalized_metadata):
                            self.fallback_prediction_cache[ticker] = {"no_prediction": True}
                            score, pred, q10, q90, source = "-", "-", "-", "-", "Bez predikce"
                        else:
                            score, pred, q10, q90, source = "-", "-", "-", "-", "Ticker nenalezen"
                    elif fallback and fallback.get("no_prediction"):
                        score, pred, q10, q90, source = "-", "-", "-", "-", "Bez predikce"
                    elif fallback:
                        fallback_record = self._fallback_record_from_data(ticker, fallback)
                        if fallback_record:
                            fb_pred, fb_q10, fb_q90 = _record_like_horizon_values(fallback_record)
                        else:
                            fb_pred = fallback.get("52w_pred")
                            fb_q10 = fallback.get("52w_q10")
                            fb_q90 = fallback.get("52w_q90")
                        score = _format_score(fallback_record.score) if fallback_record else "-"
                        pred = _format_return_pct(fb_pred)
                        q10 = _format_return_pct(fb_q10)
                        q90 = _format_return_pct(fb_q90)
                        source = "Fallback historie"
                    else:
                        score, pred, q10, q90, source = "-", "-", "-", "-", "-"
                no_sell_mark = "☑" if ticker in self.no_sell_tickers else "☐"
                tags = ("no_sell",) if ticker in self.no_sell_tickers else ()
                self.port_tree.insert(
                    "",
                    tk.END,
                    values=(
                        ticker,
                        meta["name"],
                        meta["sector"],
                        f"{meta['price']:.2f}",
                        meta["mcap"],
                        f"{amt:.2f}",
                        f"{weight:.2f}",
                        score,
                        pred,
                        q10,
                        q90,
                        no_sell_mark,
                        source,
                    ),
                    tags=tags,
                )
                
        self.fetcher.fetch_async(
            list(self.portfolio_dict.keys()),
            lambda m: self.ui_queue.put(("call", update_rows, (m,))),
        )

    def _on_portfolio_row_selected(self, _event=None):
        selected = self.port_tree.selection()
        if not selected:
            self._sync_fallback_uncertainty_entry()
            return
        values = self.port_tree.item(selected[0], "values")
        if not values:
            self._sync_fallback_uncertainty_entry()
            return
        ticker = normalize_ticker_symbol(str(values[0]))
        amount = self.portfolio_dict.get(ticker, 0.0)
        source = str(values[-1]) if len(values) > 0 else ""
        self.add_ticker_var.set(ticker)
        self.add_amount_var.set(f"{amount:.2f}")
        self._sync_fallback_uncertainty_entry(ticker, source)

    def _load_fallback_predictions_async(self, tickers: list[str], uncertainty_scale: float = 1.5) -> None:
        if not ENABLE_HOLDING_FALLBACK_PREDICTIONS:
            return
        tickers = [normalize_ticker_symbol(t) for t in tickers if t]
        if not tickers:
            return
        self.fallback_prediction_inflight.update(tickers)
        generation = self._portfolio_state_generation

        def worker():
            data = self._compute_weekly_fallback_predictions(tickers, uncertainty_scale=uncertainty_scale)
            self.ui_queue.put(("fallback_ready", generation, data, tickers))

        threading.Thread(target=worker, daemon=True).start()

    def _download_weekly_close_prices(self, yf, pd, tickers: list[str]):
        prices = pd.DataFrame()
        try:
            prices_data = yf.download(tickers, period="10y", interval="1wk", auto_adjust=True, progress=False)
            if prices_data is not None and not prices_data.empty:
                close = prices_data["Close"] if "Close" in prices_data else prices_data.get("Adj Close")
                if close is not None:
                    if isinstance(close, pd.Series):
                        prices = close.to_frame(name=tickers[0])
                    else:
                        prices = close.copy()
                        prices.columns = [normalize_ticker_symbol(str(c)) for c in prices.columns]
        except Exception as exc:
            self.log(f"Yahoo fallback batch selhal: {exc}")

        missing = []
        for ticker in tickers:
            if ticker not in prices.columns or prices[ticker].dropna().empty:
                missing.append(ticker)

        for ticker in missing:
            try:
                hist = yf.Ticker(ticker).history(period="10y", interval="1wk", auto_adjust=True)
                if hist is None or hist.empty or "Close" not in hist:
                    continue
                series = hist["Close"].dropna()
                
                # Sjednotíme indexy odstraněním časových zón (naive dates),
                # aby se evropské a americké akcie správně spárovaly podle data (lokálních 00:00:00).
                if hasattr(prices.index, "tz") and prices.index.tz is not None:
                    prices.index = prices.index.tz_localize(None)
                if hasattr(series.index, "tz") and series.index.tz is not None:
                    series.index = series.index.tz_localize(None)
                        
                if len(series) >= 20:
                    if ticker in prices.columns:
                        prices = prices.drop(columns=[ticker])
                    prices = pd.concat([prices, series.rename(ticker)], axis=1)
            except Exception as exc:
                self.log(f"Yahoo fallback individuálně selhal pro {ticker}: {exc}")
        return prices

    def _compute_weekly_fallback_predictions(self, tickers: list[str], uncertainty_scale: float = 1.5) -> dict[str, dict[str, float]]:
        """Compute fallback predictions from historical Yahoo Finance data."""
        if not ENABLE_HOLDING_FALLBACK_PREDICTIONS:
            return {}
        pd = _load_pandas()
        yf = _load_yfinance()
        tickers = [normalize_ticker_symbol(t) for t in tickers if t]
        if not yf:
            return {}
        prices = self._download_weekly_close_prices(yf, pd, tickers)
        if prices is None or prices.empty:
            return {}

        horizons = {"1w": 1, "4w": 4, "13w": 13, "26w": 26, "52w": 52}
        result: dict[str, dict[str, float]] = {}
        for ticker in tickers:
            if ticker not in prices.columns:
                continue
            series = prices[ticker].dropna().astype(float)
            if len(series) < 20:
                continue
            horizon_data: dict[str, tuple[float, float, float, float]] = {}
            for hk, window in horizons.items():
                period_returns = series.pct_change(periods=window).dropna()
                if len(period_returns) < 12:
                    continue
                mean_pred = float(period_returns.mean())

                q10_threshold = period_returns.quantile(0.10)
                tail_10 = period_returns[period_returns <= q10_threshold]
                cvar10 = float(tail_10.mean()) if len(tail_10) > 0 else float(q10_threshold)
                abs_cvar10 = mean_pred - (mean_pred - cvar10) * uncertainty_scale

                raw_q90 = float(period_returns.quantile(0.90))
                abs_q90 = mean_pred + (raw_q90 - mean_pred) * uncertainty_scale
                uncertainty = max(mean_pred - abs_cvar10, abs_q90 - mean_pred, 0.02)

                horizon_data[hk] = (mean_pred, abs_cvar10, abs_q90, uncertainty)
            if not horizon_data:
                continue
            agg_pred, agg_cvar10, agg_up = _aggregate_horizon_values(horizon_data)
            h52 = horizon_data.get("52w", horizon_data[max(horizon_data, key=lambda k: horizons[k])])
            result[ticker] = {
                "pred": agg_pred,
                "q10": _downside_tail_risk(agg_cvar10),
                "cvar10": agg_cvar10,
                "q90": agg_up,
                "52w_pred": h52[0],
                "52w_q10": h52[1],
                "52w_q90": h52[2],
                "horizon_data": horizon_data,
            }
        return result

    def _build_fallback_records_for_holdings(
        self,
        holdings: dict[str, float],
        uncertainty_scale: float,
        objective_config: ObjectiveConfig | None = None,
    ) -> list[StockRecord]:
        if not ENABLE_HOLDING_FALLBACK_PREDICTIONS:
            return []
        objective_config = objective_config or self._current_objective_config()

        def apply_repricing(records: list[StockRecord]) -> list[StockRecord]:
            latest = find_latest_prediction_csv(self.base_dir, EMA_MODEL_PREDICTION_START_DATE)
            reference_date = latest[0] if latest else None
            self._apply_midweek_prediction_repricing(
                records,
                reference_date,
                objective_config=objective_config,
                log_summary=False,
            )
            return records

        known = {normalize_ticker_symbol(r.ticker) for r in self.records}
        missing = [normalize_ticker_symbol(t) for t in holdings.keys() if normalize_ticker_symbol(t) not in known]
        if not missing:
            return []

        records = load_holding_ema_fallback_records(
            self.base_dir,
            missing,
            objective_config,
            self.log,
        )
        resolved = {normalize_ticker_symbol(record.ticker) for record in records}
        remaining = [ticker for ticker in missing if ticker not in resolved]
        if not remaining:
            return apply_repricing(records)

        fallback_data = self._compute_weekly_fallback_predictions(remaining, uncertainty_scale=uncertainty_scale)
        for ticker in remaining:
            data = fallback_data.get(ticker)
            if not data:
                if self._ticker_has_market_evidence(ticker):
                    self.fallback_prediction_cache[ticker] = {"no_prediction": True}
                    self.fallback_prediction_failures[ticker] = self.fallback_max_retries
                continue
            downside = max(1e-6, float(data["q10"]))
            upside = float(data["q90"])
            horizon_data = data.get("horizon_data", {})
            record = StockRecord(
                ticker=ticker,
                sharpe=float(_starr_score(data["pred"], downside)),
                forecast_pct=float(data["pred"]),
                std_pct=downside,
                upside_pct=upside,
                is_synthetic=True,
                horizon_data=horizon_data if isinstance(horizon_data, dict) else None,
                cvar10_pct=float(data.get("cvar10", -downside)),
            )
            record.score = _record_objective_score(record, objective_config)
            records.append(record)
        return apply_repricing(records)

    def _add_holding(self):
        if self._blocked_action("Úprava portfolia"):
            return
        if self._is_trading212_linked_portfolio():
            messagebox.showinfo(
                "Trading 212",
                "Portfolio napojené na Trading 212 je v manageru jen zrcadlo brokeru.\n"
                "Pozice neupravuj ručně; aktuální stav se obnoví automaticky při načtení uloženého portfolia.",
                parent=self.root,
            )
            return
        t = normalize_ticker_symbol(self.add_ticker_var.get())
        amt_str = self.add_amount_var.get().strip()
        if not t or not amt_str: return
        try:
            amt = float(amt_str)
            self.portfolio_dict[t] = amt
            record_map = {normalize_ticker_symbol(r.ticker) for r in self.records}
            if ENABLE_HOLDING_FALLBACK_PREDICTIONS and t not in record_map:
                uncertainty = float(self.fallback_uncertainty_var.get() or 1.0)
                self.fallback_uncertainty_by_ticker[t] = uncertainty
                self.fallback_prediction_inflight.discard(t)
                self.fallback_prediction_cache.pop(t, None)
                self._load_fallback_predictions_async([t], uncertainty_scale=uncertainty)
            self._refresh_portfolio_tree()
            self.add_ticker_var.set("")
            self.add_amount_var.set("")
        except:
            messagebox.showerror("Chyba", "Neplatná částka")

    def _remove_holding(self):
        if self._blocked_action("Úprava portfolia"):
            return
        if self._is_trading212_linked_portfolio():
            messagebox.showinfo(
                "Trading 212",
                "Portfolio napojené na Trading 212 je v manageru jen zrcadlo brokeru.\n"
                "Pozice neodebírej ručně; aktuální stav se obnoví automaticky při načtení uloženého portfolia.",
                parent=self.root,
            )
            return
        t = normalize_ticker_symbol(self.add_ticker_var.get())
        if t in self.portfolio_dict:
            del self.portfolio_dict[t]
            self.portfolio_price_snapshot.pop(t, None)
            self.portfolio_broker_positions.pop(t, None)
            self.no_sell_tickers.discard(t)
            self._refresh_portfolio_tree()
            self.add_ticker_var.set("")

    def _new_portfolio(self):
        if self._blocked_action("Nové portfolio"):
            return
        self._reset_portfolio_session_state(clear_portfolio_data=True)
        self._refresh_portfolio_tree()


    def _compose_projection_from_holdings(self, holdings: dict[str, float], label: str, source: str, created_at: str, override_total: float | None = None):
        import datetime
        if not holdings:
            return None
        horizons = [1, 4, 13, 26, 52]
        rec_map = {normalize_ticker_symbol(r.ticker): r for r in self.records}
        needed_fallback = [normalize_ticker_symbol(t) for t in holdings if normalize_ticker_symbol(t) not in rec_map]
        if needed_fallback:
            if not ENABLE_HOLDING_FALLBACK_PREDICTIONS:
                self.log(
                    "Projekci neskládám, protože chybí modelová predikce pro: "
                    + ", ".join(sorted(needed_fallback))
                )
                return None
            ema_records = load_holding_ema_fallback_records(
                self.base_dir,
                needed_fallback,
                self._current_objective_config(),
                self.log,
            )
            for record in ema_records:
                rec_map[normalize_ticker_symbol(record.ticker)] = record
            needed_fallback = [
                normalize_ticker_symbol(t)
                for t in holdings
                if normalize_ticker_symbol(t) not in rec_map
            ]
        if needed_fallback:
            scale = float(self.fallback_uncertainty_var.get() or 1.0)
            fresh = self._compute_weekly_fallback_predictions(needed_fallback, scale)
            self.fallback_prediction_cache.update(fresh)
            unresolved = [
                ticker for ticker in needed_fallback
                if not isinstance(self.fallback_prediction_cache.get(ticker), dict)
                or not self.fallback_prediction_cache.get(ticker, {}).get("horizon_data")
            ]
            if unresolved:
                self.log(
                    "Projekci neskládám, protože fallback predikce chybí pro: "
                    + ", ".join(sorted(unresolved))
                )
                return None
        total_val = override_total if override_total is not None else (sum(float(v) for v in holdings.values()) + float(self.extra_cash_var.get() or 0.0))
        holdings_total = sum(float(v) for v in holdings.values())
        if total_val <= 0 or holdings_total <= 0:
            return None
            
        ticker_list = list(holdings.keys())
        w_list = [float(holdings[t]) / holdings_total for t in ticker_list]
        corr_matrix = build_correlation_matrix(ticker_list, self.return_cache, self.log)

        agg_p = {}
        agg_q10 = {}
        agg_q90 = {}

        for w in horizons:
            p_list = []
            c10_list = []
            q90_list = []
            
            for ticker in ticker_list:
                rec = rec_map.get(normalize_ticker_symbol(ticker))
                if rec and rec.horizon_data:
                    hdata = rec.horizon_data
                else:
                    hdata = self.fallback_prediction_cache.get(normalize_ticker_symbol(ticker), {}).get("horizon_data", {})
                
                default_p = rec.forecast_pct if rec else 0.0
                p, s, u, _ = _coerce_horizon_values(
                    hdata.get(f"{w}w"),
                    (default_p, default_p, default_p, 0.0),
                )
                p_list.append(float(p))
                c10_list.append(float(s))
                q90_list.append(float(u))
            
            p_port, q10_port, q90_port = _diversified_portfolio_quantiles(
                ticker_list, w_list, p_list, c10_list, q90_list, corr_matrix
            )
            agg_p[w] = p_port
            agg_q10[w] = q10_port
            agg_q90[w] = q90_port

        created_dt = datetime.datetime.fromisoformat(created_at)
        dates = [(created_dt + datetime.timedelta(weeks=w)).isoformat() for w in horizons]
        expected = [total_val * (1.0 + agg_p[w]) for w in horizons]
        q10_vals = [total_val * (1.0 + agg_q10[w]) for w in horizons]
        q90_vals = [total_val * (1.0 + agg_q90[w]) for w in horizons]
        return {
            "date": created_at,
            "label": label,
            "source": source,
            "start_value": total_val,
            "holdings": {normalize_ticker_symbol(t): float(v) for t, v in holdings.items()},
            "dates": dates,
            "expected": expected,
            "q10": q10_vals,
            "q90": q90_vals,
        }

    def _linked_trading212_pie(self) -> dict | None:
        link = self.portfolio_broker_link if isinstance(self.portfolio_broker_link, dict) else {}
        if link.get("broker") != "Trading 212":
            return None
        pie_id = str(link.get("pie_id") or "")
        if not pie_id and str(link.get("source_id") or "").startswith("pie:"):
            pie_id = str(link.get("source_id")).split(":", 1)[1]
        if not pie_id:
            return None
        result = dict(link)
        result["pie_id"] = pie_id
        result.setdefault("source_id", f"pie:{pie_id}")
        return result

    def _current_t212_broker_tickers_by_app(self) -> dict[str, str]:
        mapping: dict[str, str] = {}
        broker_positions = self.portfolio_broker_positions if isinstance(self.portfolio_broker_positions, dict) else {}
        for app_ticker, detail in broker_positions.items():
            key = normalize_ticker_symbol(app_ticker)
            if not isinstance(detail, dict):
                continue
            for position in detail.get("positions", []):
                if not isinstance(position, dict):
                    continue
                broker_ticker = str(position.get("broker_ticker") or "").strip()
                if broker_ticker:
                    self._remember_preferred_t212_broker_ticker(mapping, key, broker_ticker)
        settings = (self.portfolio_broker_link or {}).get("pie_settings")
        instrument_shares = settings.get("instrumentShares") if isinstance(settings, dict) else {}
        if isinstance(instrument_shares, dict):
            for broker_ticker in instrument_shares.keys():
                app_ticker = self._t212_app_ticker(str(broker_ticker))
                self._remember_preferred_t212_broker_ticker(mapping, app_ticker, str(broker_ticker))
        return mapping

    def _trading212_instrument_map(self, client: Trading212Client) -> dict[str, str]:
        mapping, _ = self._trading212_instrument_map_and_valid_brokers(client)
        return mapping

    def _trading212_instrument_cache_path(self, client: Trading212Client) -> Path:
        base_dir = Path(getattr(self, "base_dir", Path(".")))
        environment = re.sub(r"[^a-z0-9_-]+", "_", str(getattr(client, "environment", "live")).lower())
        return base_dir / "cache" / f"trading212_instruments_{environment}.json"

    def _trading212_us_universe_cache_path(self, client: Trading212Client) -> Path:
        base_dir = Path(getattr(self, "base_dir", Path(".")))
        environment = re.sub(r"[^a-z0-9_-]+", "_", str(getattr(client, "environment", "live")).lower())
        return base_dir / "cache" / f"trading212_us_universe_{environment}.json"

    def _trading212_us_universe_csv_path(self, client: Trading212Client) -> Path:
        base_dir = Path(getattr(self, "base_dir", Path(".")))
        environment = re.sub(r"[^a-z0-9_-]+", "_", str(getattr(client, "environment", "live")).lower())
        return base_dir / "cache" / f"trading212_us_universe_{environment}.csv"

    def _read_trading212_instrument_cache(self, client: Trading212Client) -> list[dict]:
        path = self._trading212_instrument_cache_path(client)
        try:
            if not path.exists():
                return []
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            instruments = payload.get("instruments") if isinstance(payload, dict) else None
            return [item for item in instruments if isinstance(item, dict)] if isinstance(instruments, list) else []
        except Exception:
            return []

    def _read_trading212_us_universe_cache(self, client: Trading212Client) -> list[dict]:
        path = self._trading212_us_universe_cache_path(client)
        try:
            if not path.exists():
                return []
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            rows = payload.get("items") if isinstance(payload, dict) else None
            if not isinstance(rows, list):
                return []
            instruments: list[dict] = []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                broker_ticker = str(row.get("broker_ticker") or "").strip()
                if not broker_ticker:
                    continue
                instruments.append(
                    {
                        "ticker": broker_ticker,
                        "name": str(row.get("name") or row.get("short_name") or "").strip(),
                        "shortName": str(row.get("short_name") or "").strip(),
                        "isin": str(row.get("isin") or "").strip(),
                        "currencyCode": str(row.get("currency") or "").strip(),
                        "exchange": str(row.get("exchange") or "").strip(),
                        "countryCode": str(row.get("country") or "").strip(),
                        "type": str(row.get("type") or "").strip(),
                        "tradable": True,
                    }
                )
            return instruments
        except Exception:
            return []

    def _build_trading212_us_universe_payload(
        self,
        client: Trading212Client,
        instruments: list[dict],
    ) -> dict:
        rows: list[dict[str, str]] = []
        by_app_ticker: dict[str, str] = {}
        by_app_instrument: dict[str, dict] = {}

        for instrument in instruments:
            if not isinstance(instrument, dict):
                continue
            broker_ticker = str(instrument.get("ticker") or "").strip()
            if not broker_ticker:
                continue
            if broker_ticker.upper() in TRADING212_BLOCKED_BROKER_TICKERS:
                continue
            if not self._t212_is_us_broker_ticker(broker_ticker, instrument=instrument):
                continue
            if not self._t212_instrument_is_tradeable(instrument):
                continue

            name = str(instrument.get("name") or "").strip()
            short_name = str(instrument.get("shortName") or "").strip()
            isin = str(instrument.get("isin") or "").strip().upper()
            app_ticker = self._t212_app_ticker(
                broker_ticker,
                name=name or short_name,
                short_name=short_name,
                isin=isin,
            )
            if not app_ticker:
                continue

            rows.append(
                {
                    "app_ticker": app_ticker,
                    "broker_ticker": broker_ticker,
                    "name": name,
                    "short_name": short_name,
                    "isin": isin,
                    "currency": self._t212_instrument_text(
                        instrument,
                        "currencyCode",
                        "currency",
                        "tradingCurrency",
                    ),
                    "exchange": self._t212_instrument_text(
                        instrument,
                        "exchange",
                        "exchangeCode",
                        "exchangeName",
                        "market",
                        "marketName",
                    ),
                    "country": self._t212_instrument_text(
                        instrument,
                        "countryCode",
                        "country",
                        "region",
                    ),
                    "type": self._t212_instrument_text(
                        instrument,
                        "type",
                        "instrumentType",
                        "category",
                    ),
                }
            )

            preferred = self._prefer_t212_broker_ticker(
                by_app_ticker.get(app_ticker),
                broker_ticker,
                app_ticker=app_ticker,
                current_instrument=by_app_instrument.get(app_ticker),
                candidate_instrument=instrument,
            )
            if preferred == broker_ticker:
                by_app_ticker[app_ticker] = broker_ticker
                by_app_instrument[app_ticker] = instrument

        rows.sort(key=lambda item: (item["app_ticker"], item["broker_ticker"]))
        return {
            "environment": getattr(client, "environment", ""),
            "updated_at": dt.datetime.now().isoformat(),
            "count": len(rows),
            "by_app_ticker": dict(sorted(by_app_ticker.items())),
            "items": rows,
        }

    def _write_trading212_us_universe_cache(self, client: Trading212Client, instruments: list[dict]) -> None:
        try:
            payload = self._build_trading212_us_universe_payload(client, instruments)
            json_path = self._trading212_us_universe_cache_path(client)
            csv_path = self._trading212_us_universe_csv_path(client)
            json_path.parent.mkdir(parents=True, exist_ok=True)
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, ensure_ascii=False, default=str)

            fieldnames = [
                "app_ticker",
                "broker_ticker",
                "name",
                "short_name",
                "isin",
                "currency",
                "exchange",
                "country",
                "type",
            ]
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(payload["items"])

            self.log(
                f"Trading 212: uložen US universe ({payload['count']} instrumentů) do {csv_path.name}."
            )
        except Exception as exc:
            self.log(f"Trading 212: US universe cache se nepodařilo uložit: {exc}")

    def _write_trading212_instrument_cache(self, client: Trading212Client, instruments: list[dict]) -> None:
        path = self._trading212_instrument_cache_path(client)
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(
                    {
                        "environment": getattr(client, "environment", ""),
                        "updated_at": dt.datetime.now().isoformat(),
                        "count": len(instruments),
                        "instruments": instruments,
                    },
                    f,
                    indent=2,
                    ensure_ascii=False,
                    default=str,
                )
            self._write_trading212_us_universe_cache(client, instruments)
        except Exception as exc:
            self.log(f"Trading 212: cache instrumentů se nepodařilo uložit: {exc}")

    @staticmethod
    def _t212_bool_field(instrument: dict, *keys: str) -> bool | None:
        for key in keys:
            if key not in instrument:
                continue
            value = instrument.get(key)
            if isinstance(value, bool):
                return value
            if isinstance(value, (int, float)):
                return bool(value)
            text = str(value).strip().lower()
            if text in {"true", "yes", "1"}:
                return True
            if text in {"false", "no", "0"}:
                return False
        return None

    @staticmethod
    def _t212_instrument_is_tradeable(instrument: dict) -> bool:
        if not isinstance(instrument, dict):
            return False
        positive_flag = PortfolioManagerApp._t212_bool_field(
            instrument,
            "tradable",
            "isTradable",
            "active",
            "isActive",
            "enabled",
            "isEnabled",
        )
        if positive_flag is False:
            return False
        negative_flag = PortfolioManagerApp._t212_bool_field(
            instrument,
            "delisted",
            "isDelisted",
            "suspended",
            "isSuspended",
            "blocked",
            "isBlocked",
        )
        if negative_flag is True:
            return False
        status = PortfolioManagerApp._t212_instrument_text(
            instrument,
            "status",
            "tradingStatus",
            "state",
            "instrumentStatus",
        )
        inactive_markers = ("DELIST", "SUSPEND", "INACTIVE", "REMOVED", "CLOSED", "DISABLED")
        if status and any(marker in status for marker in inactive_markers):
            return False
        return True

    def _trading212_instrument_map_and_valid_brokers(
        self,
        client: Trading212Client,
    ) -> tuple[dict[str, str], set[str]]:
        instruments: list[dict] = []
        api_error: Exception | None = None
        try:
            instruments = client.instruments()
            if instruments:
                self._write_trading212_instrument_cache(client, instruments)
                self.log(f"Trading 212: ověřeno {len(instruments)} instrumentů pro mapování tickerů.")
        except Exception as exc:
            api_error = exc
            instruments = self._read_trading212_instrument_cache(client)
            if instruments:
                self.log(
                    f"Trading 212: metadata instrumentů se teď nepodařilo načíst ({exc}); "
                    f"použiji cache s {len(instruments)} instrumenty."
                )
            else:
                instruments = self._read_trading212_us_universe_cache(client)
                if instruments:
                    self.log(
                        f"Trading 212: metadata instrumentů se teď nepodařilo načíst ({exc}); "
                        f"použiji US universe cache s {len(instruments)} instrumenty."
                    )
        if not instruments:
            message = (
                "Nepodařilo se ověřit dostupné Trading 212 instrumenty a není k dispozici cache. "
                "Update koláče proto nemůže bezpečně přidávat nové tickery."
            )
            if api_error is not None:
                raise Trading212ApiError(message) from api_error
            raise Trading212ApiError("Trading 212 nevrátil žádná metadata instrumentů.")

        try:
            full_cache_path = self._trading212_instrument_cache_path(client)
            universe_cache_path = self._trading212_us_universe_cache_path(client)
            should_refresh_universe = not universe_cache_path.exists()
            if not should_refresh_universe and full_cache_path.exists():
                should_refresh_universe = universe_cache_path.stat().st_mtime < full_cache_path.stat().st_mtime
            if should_refresh_universe:
                self._write_trading212_us_universe_cache(client, instruments)
        except Exception:
            pass

        mapping: dict[str, str] = {}
        valid_brokers: set[str] = set()
        for instrument in instruments:
            if not isinstance(instrument, dict):
                continue
            broker_ticker = str(instrument.get("ticker") or "").strip()
            if not broker_ticker:
                continue
            if broker_ticker.upper() in TRADING212_BLOCKED_BROKER_TICKERS:
                continue
            if not self._t212_is_us_broker_ticker(broker_ticker, instrument=instrument):
                continue
            if not self._t212_instrument_is_tradeable(instrument):
                continue
            valid_brokers.add(broker_ticker)
            app_ticker = self._t212_app_ticker(
                broker_ticker,
                name=str(instrument.get("name") or instrument.get("shortName") or ""),
                short_name=str(instrument.get("shortName") or ""),
                isin=str(instrument.get("isin") or ""),
            )
            self._remember_preferred_t212_broker_ticker(mapping, app_ticker, broker_ticker, instrument=instrument)
        return mapping, valid_brokers

    @staticmethod
    def _normalize_pie_shares(shares: dict[str, float], keep_zero: bool = False) -> dict[str, float]:
        cleaned: dict[str, float] = {}
        zero_tickers: list[str] = []
        for ticker, value in shares.items():
            broker_ticker = str(ticker or "").strip()
            if not broker_ticker:
                continue
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                continue
            if numeric > 0 and math.isfinite(numeric):
                cleaned[broker_ticker] = numeric
            elif keep_zero and math.isfinite(numeric):
                zero_tickers.append(broker_ticker)
        total = sum(cleaned.values())
        if total <= 0:
            return {ticker: 0.0 for ticker in zero_tickers} if keep_zero else {}
        normalized = {ticker: round(value / total, 4) for ticker, value in cleaned.items()}
        drift = round(1.0 - sum(normalized.values()), 4)
        if normalized and abs(drift) > 0:
            largest = max(normalized, key=normalized.get)
            normalized[largest] = round(max(0.0, normalized[largest] + drift), 4)
        if keep_zero:
            for ticker in zero_tickers:
                normalized.setdefault(ticker, 0.0)
        return normalized

    def _limit_t212_pie_share_keys(
        self,
        shares: dict[str, float],
        zero_priority: dict[str, float] | None = None,
        max_keys: int = 50,
    ) -> dict[str, float]:
        normalized = self._normalize_pie_shares(shares, keep_zero=True)
        positives = [(ticker, float(value)) for ticker, value in normalized.items() if float(value) > 0]
        zero_priority = zero_priority or {}
        zeros = [
            (float(zero_priority.get(ticker, 0.0) or 0.0), ticker)
            for ticker, value in normalized.items()
            if float(value) <= 0
        ]
        if len(positives) > max_keys:
            raise Trading212ApiError(
                f"Trading 212 koláč může mít maximálně {max_keys} aktivních pozic, "
                f"ale cílové váhy obsahují {len(positives)}."
            )

        room_for_zeros = max(0, max_keys - len(positives))
        zeros.sort(key=lambda item: (-item[0], item[1]))
        kept_zeros = zeros[:room_for_zeros]
        dropped_zeros = zeros[room_for_zeros:]
        if dropped_zeros:
            preview = ", ".join(ticker for _, ticker in dropped_zeros[:8])
            self.log(
                f"Trading 212: payload koláče omezen na {max_keys} položek; "
                f"{len(dropped_zeros)} nulových vah nebude posláno"
                + (f" ({preview}{'...' if len(dropped_zeros) > 8 else ''})." if preview else ".")
            )

        limited = {ticker: value for ticker, value in positives}
        for _, ticker in kept_zeros:
            limited.setdefault(ticker, 0.0)
        return self._normalize_pie_shares(limited, keep_zero=True)

    def _extract_t212_pie_shares(self, data) -> dict[str, float]:
        if not isinstance(data, dict):
            return {}
        for nested_key in ("target_response", "verified_detail", "target_payload", "payload"):
            nested = data.get(nested_key)
            if isinstance(nested, dict):
                nested_shares = self._extract_t212_pie_shares(nested)
                if nested_shares:
                    return nested_shares
        for container in (data.get("settings"), data):
            if not isinstance(container, dict):
                continue
            shares = container.get("instrumentShares")
            if isinstance(shares, dict) and shares:
                return self._normalize_pie_shares(shares, keep_zero=True)

        instruments = data.get("instruments")
        if isinstance(instruments, list):
            shares = {}
            for item in instruments:
                if not isinstance(item, dict):
                    continue
                ticker = str(item.get("ticker") or item.get("instrumentTicker") or "").strip()
                expected = item.get(
                    "expectedShare",
                    item.get("targetShare", item.get("share", item.get("weight"))),
                )
                if ticker and expected is not None:
                    shares[ticker] = expected
            if shares:
                return self._normalize_pie_shares(shares, keep_zero=True)
        return {}

    @staticmethod
    def _pie_shares_match(expected: dict[str, float], observed: dict[str, float], tolerance: float = 0.0005) -> bool:
        if not expected or not observed:
            return False
        meaningful_keys = {
            ticker
            for ticker in set(expected) | set(observed)
            if abs(float(expected.get(ticker, 0.0))) > tolerance or abs(float(observed.get(ticker, 0.0))) > tolerance
        }
        if not meaningful_keys:
            return False
        return all(abs(float(expected.get(ticker, 0.0)) - float(observed.get(ticker, 0.0))) <= tolerance for ticker in meaningful_keys)

    @staticmethod
    def _pie_share_diff_summary(expected: dict[str, float], observed: dict[str, float], limit: int = 8) -> str:
        keys = sorted(set(expected) | set(observed))
        rows = []
        for ticker in keys:
            exp = float(expected.get(ticker, 0.0))
            obs = float(observed.get(ticker, 0.0))
            rows.append((abs(exp - obs), ticker, exp, obs))
        rows.sort(reverse=True)
        parts = [
            f"{ticker}: posláno {exp * 100:.2f} %, Trading 212 {obs * 100:.2f} %"
            for _, ticker, exp, obs in rows[:limit]
        ]
        if len(rows) > limit:
            parts.append(f"... +{len(rows) - limit} dalších")
        return "; ".join(parts)

    def _save_trading212_pie_update_debug(
        self,
        pie_id,
        payload: dict,
        update_response=None,
        verified_detail=None,
        error: str | None = None,
    ):
        try:
            debug_path = self.base_dir / "portfolios" / "trading212_last_pie_update_debug.json"
            debug_path.parent.mkdir(parents=True, exist_ok=True)
            debug_payload = {
                "saved_at": dt.datetime.now().isoformat(),
                "pie_id": str(pie_id),
                "payload": payload,
                "update_response": update_response,
                "verified_detail": verified_detail,
                "error": error,
            }
            with open(debug_path, "w", encoding="utf-8") as f:
                json.dump(debug_payload, f, indent=2, ensure_ascii=False, default=str)
        except Exception as exc:
            self.log(f"Nepodařilo se uložit Trading 212 update diagnostiku: {exc}")

    @staticmethod
    def _temporary_t212_pie_name(name: str) -> str:
        base = PortfolioManagerApp._clean_temporary_t212_pie_name(name)
        now = dt.datetime.now()
        suffix = f" PM{now:%H%M%S}{now.microsecond // 10000:02d}"
        max_len = 50
        trimmed = base[: max(1, max_len - len(suffix))].rstrip()
        temporary = f"{trimmed}{suffix}"
        if temporary == base:
            temporary = f"{trimmed[: max(1, len(trimmed) - 1)]}{suffix}"
        return temporary

    @staticmethod
    def _clean_temporary_t212_pie_name(name: str) -> str:
        base = str(name or "Pie").replace('\u200B', '').strip() or "Pie"
        cleaned = re.sub(r"(?:\s+PM\d{6,8})+$", "", base).strip()
        return cleaned or "Pie"

    def _pie_payload_from_detail(self, detail: dict, fallback_payload: dict | None = None, override_name: str | None = None) -> dict:
        fallback_payload = dict(fallback_payload or {})
        settings = detail.get("settings") if isinstance(detail.get("settings"), dict) else {}
        shares = self._extract_t212_pie_shares(detail)
        if not shares:
            shares = self._normalize_pie_shares(fallback_payload.get("instrumentShares", {}), keep_zero=True)
        name = override_name if override_name is not None else (
            settings.get("name") or fallback_payload.get("name") or "Pie"
        )
        
        end_date = settings.get("endDate") or fallback_payload.get("endDate")
        if end_date is not None:
            end_date = str(end_date)
            
        goal = settings.get("goal") or fallback_payload.get("goal")
        if goal is not None:
            try:
                goal = float(goal)
            except (TypeError, ValueError):
                goal = None

        return {
            "dividendCashAction": settings.get("dividendCashAction") or fallback_payload.get("dividendCashAction") or "REINVEST",
            "endDate": end_date,
            "goal": goal,
            "icon": settings.get("icon") or fallback_payload.get("icon") or "Home",
            "instrumentShares": shares,
            "name": str(name or "Pie"),
        }

    def _t212_pie_payload_from_result(self, result: OptimizationResult, client: Trading212Client, link: dict) -> tuple[dict, list[str]]:
        raw_target_values: dict[str, float] = {}
        for trade in result.trades:
            value = float(trade.target_czk)
            if value <= 1.0:
                continue
            ticker = normalize_ticker_symbol(trade.ticker)
            raw_target_values[ticker] = raw_target_values.get(ticker, 0.0) + value
        if sum(raw_target_values.values()) <= 0:
            raise Trading212ApiError("Optimalizace nemá žádné cílové pozice pro update koláče.")

        broker_by_app = self._current_t212_broker_tickers_by_app()
        settings = link.get("pie_settings") if isinstance(link.get("pie_settings"), dict) else {}
        current_pie_shares = self._normalize_pie_shares(settings.get("instrumentShares") or {}, keep_zero=True)
        if not current_pie_shares:
            try:
                current_detail = client.pie_detail(link.get("pie_id"), retries=2, retry_wait_seconds=4.0)
                current_pie_shares = self._extract_t212_pie_shares(current_detail)
                detail_settings = current_detail.get("settings") if isinstance(current_detail.get("settings"), dict) else {}
                if detail_settings:
                    settings = {**settings, **detail_settings}
            except Exception as exc:
                self.log(f"Trading 212: nepodařilo se načíst aktuální nulové váhy koláče: {exc}")
        for broker_ticker in current_pie_shares:
            app_ticker = self._t212_app_ticker(str(broker_ticker))
            if app_ticker:
                self._remember_preferred_t212_broker_ticker(broker_by_app, app_ticker, str(broker_ticker))

        verified_broker_by_app, valid_broker_tickers = self._trading212_instrument_map_and_valid_brokers(client)
        unsupported_targets: list[str] = []
        target_values: dict[str, float] = {}
        for ticker, value in raw_target_values.items():
            current_broker = broker_by_app.get(ticker, "")
            verified_broker = verified_broker_by_app.get(ticker)
            alias_broker = TRADING212_APP_TICKER_BROKER_ALIASES.get(ticker)
            if alias_broker and alias_broker in valid_broker_tickers:
                if current_broker and current_broker != alias_broker:
                    self.log(
                        f"Trading 212: {ticker} použije broker alias {alias_broker} místo {current_broker}."
                    )
                broker_by_app[ticker] = alias_broker
                target_values[ticker] = value
                continue
            if verified_broker:
                if current_broker and current_broker != verified_broker:
                    self.log(
                        f"Trading 212: {ticker} použije ověřený ticker {verified_broker} místo {current_broker}."
                    )
                broker_by_app[ticker] = verified_broker
                target_values[ticker] = value
                continue
            if current_broker and current_broker in valid_broker_tickers and self._t212_is_us_broker_ticker(current_broker):
                target_values[ticker] = value
                continue
            unsupported_targets.append(ticker)

        if unsupported_targets:
            self.log(
                "Trading 212: vynechávám cílové tickery bez ověřeného obchodovatelného US instrumentu: "
                + ", ".join(unsupported_targets)
            )

        total_target = sum(target_values.values())
        if total_target <= 0:
            raise Trading212ApiError(
                "Po odfiltrování neověřených Trading 212 tickerů nezůstala žádná cílová pozice."
            )
        non_us_targets = [
            f"{ticker} -> {broker_by_app[ticker]}"
            for ticker in target_values
            if not self._t212_is_us_broker_ticker(broker_by_app.get(ticker, ""))
        ]
        if non_us_targets:
            raise Trading212ApiError(
                "Cílové portfolio je omezené na americkou burzu, ale Trading 212 mapování není US pro: "
                + ", ".join(non_us_targets)
            )

        raw_shares = {
            broker_by_app[ticker]: value / total_target
            for ticker, value in target_values.items()
        }
        for broker_ticker in current_pie_shares:
            raw_shares.setdefault(broker_ticker, 0.0)

        instrument_shares = self._normalize_pie_shares(raw_shares, keep_zero=True)
        if not instrument_shares:
            raise Trading212ApiError("Nepodařilo se sestavit cílové váhy koláče.")

        instrument_shares = self._limit_t212_pie_share_keys(
            instrument_shares,
            zero_priority=current_pie_shares,
            max_keys=50,
        )
        positive_share_count = sum(1 for value in instrument_shares.values() if float(value) > 0)
        if positive_share_count > 50:
            raise Trading212ApiError(
                f"Trading 212 koláč může mít maximálně 50 aktivních pozic, ale cílové váhy obsahují {positive_share_count}. "
                "Sniž cílový počet akcií nebo nastav větší odprodej některých pozic."
            )

        end_date = settings.get("endDate")
        if end_date is not None:
            end_date = str(end_date)
            
        goal = settings.get("goal")
        if goal is not None:
            try:
                goal = float(goal)
            except (TypeError, ValueError):
                goal = None
                
        payload = {
            "dividendCashAction": settings.get("dividendCashAction") or "REINVEST",
            "endDate": end_date,
            "goal": goal,
            "icon": settings.get("icon") or "Home",
            "instrumentShares": instrument_shares,
            "name": self._clean_temporary_t212_pie_name(
                settings.get("name") or str(link.get("source_label") or f"Pie {link.get('pie_id')}").replace("Koláč: ", "")
            ),
        }
        return payload, unsupported_targets

    def _apply_trading212_pie_weights_async(self, result: OptimizationResult, top: tk.Toplevel):
        link = self._linked_trading212_pie()
        if not link:
            messagebox.showerror("Trading 212", "Aktuální portfolio není napojené na konkrétní Trading 212 koláč.", parent=top)
            return
        environment = self._trading212_env_name(str(link.get("environment") or "live"))
        if not messagebox.askyesno(
            "Potvrdit změnu vah na Trading 212",
            "Tato akce odešle nové cílové váhy do skutečného Trading 212 koláče.\n\n"
            "Nebudou zadány žádné nákupní ani prodejní ordery. Změní se pouze složení/váhy koláče "
            "v Trading 212. Portfolio manager po této akci lokálně neprovede žádné obchody; "
            "skutečné pozice se do něj propíšou až dalším načtením stavu z Trading 212.\n\n"
            "Kvůli omezení Trading 212 Pies API program koláč krátce přejmenuje a potom ho vrátí "
            "na původní název; bez změny názvu API update vah odmítá.\n\n"
            f"Prostředí: {environment}\nKoláč: {link.get('source_label') or link.get('pie_id')}\n\n"
            "Opravdu pokračovat?",
            parent=top,
        ):
            return

        try:
            client = self._trading212_client_for_link(link)
        except Trading212ApiError as exc:
            messagebox.showerror("Trading 212", str(exc), parent=top)
            self._save_trading212_pie_update_debug(
                link.get("pie_id", ""),
                {},
                update_response={"stage": "client_error", "environment": environment},
                error=str(exc),
            )
            return

        self.log("Odesílám nové váhy koláče do Trading 212...")
        self._save_trading212_pie_update_debug(
            link.get("pie_id", ""),
            {},
            update_response={
                "stage": "confirmed",
                "environment": environment,
                "source_label": link.get("source_label"),
            },
        )

        def _safe_parent():
            try:
                if top.winfo_exists():
                    return top
            except (tk.TclError, Exception):
                pass
            return self.root

        def finish(payload, update_response, verified_detail, verify_error, restore_error, error):
            if error:
                self._save_trading212_pie_update_debug(
                    link["pie_id"],
                    payload or {},
                    update_response=update_response or {"stage": "failed_before_payload"},
                    verified_detail=verified_detail,
                    error=str(error),
                )
                self.log(f"Trading 212 update vah selhal: {error}")
                try:
                    messagebox.showerror("Trading 212", f"Update vah koláče selhal:\n{error}", parent=_safe_parent())
                except (tk.TclError, Exception):
                    pass
                return
            expected_shares = self._normalize_pie_shares(payload.get("instrumentShares", {}))
            observed_shares = (
                self._extract_t212_pie_shares(verified_detail)
                or self._extract_t212_pie_shares((update_response or {}).get("target_response") if isinstance(update_response, dict) else None)
                or self._extract_t212_pie_shares(update_response)
            )
            if observed_shares and verified_detail is None and isinstance(update_response, dict):
                verified_detail = update_response.get("target_response")
            self._save_trading212_pie_update_debug(
                link["pie_id"],
                payload,
                update_response=update_response,
                verified_detail=verified_detail,
                error=str(verify_error) if verify_error else None,
            )
            if not observed_shares:
                self.log("Trading 212 update: API odpovědělo, ale nepodařilo se ověřit cílové váhy koláče.")
                if verify_error:
                    self.log(f"Ověření detailu koláče selhalo: {verify_error}")
                try:
                    messagebox.showwarning(
                        "Trading 212",
                        "Trading 212 API odpovědělo bez chyby, ale program nedokázal ověřit, že se cílové váhy opravdu změnily.\n\n"
                        "Uložil jsem diagnostiku do portfolios/trading212_last_pie_update_debug.json.",
                        parent=_safe_parent(),
                    )
                except (tk.TclError, Exception):
                    pass
                return
            if not self._pie_shares_match(expected_shares, observed_shares):
                diff = self._pie_share_diff_summary(expected_shares, observed_shares)
                self.log(f"Trading 212 update vah nepotvrzen: {diff}")
                try:
                    messagebox.showwarning(
                        "Trading 212",
                        "Trading 212 API odpovědělo, ale ověřené váhy koláče neodpovídají odeslaným cílovým vahám.\n\n"
                        f"{diff}\n\n"
                        "Uložil jsem diagnostiku do portfolios/trading212_last_pie_update_debug.json.",
                        parent=_safe_parent(),
                    )
                except (tk.TclError, Exception):
                    pass
                return
            if restore_error:
                self.log(f"Trading 212 váhy byly změněny, ale návrat původního názvu selhal: {restore_error}")
            self.portfolio_broker_link = dict(self.portfolio_broker_link or {})
            settings = dict(self.portfolio_broker_link.get("pie_settings") or {})
            settings.update(payload)
            self.portfolio_broker_link["pie_settings"] = settings
            try:
                self._write_portfolio_file(quiet=True)
            except Exception as exc:
                self.log(f"Váhy byly odeslány, ale portfolio se nepodařilo uložit: {exc}")
            try:
                messagebox.showinfo(
                    "Trading 212",
                    "Nové cílové váhy koláče byly odeslány do Trading 212.\n"
                    "Žádné obchody nebyly zadány přes API a lokální portfolio nebylo ručně přepsáno."
                    + ("\n\nPozor: váhy se změnily, ale nepodařilo se vrátit původní název koláče." if restore_error else ""),
                    parent=_safe_parent(),
                )
            except (tk.TclError, Exception):
                pass
            try:
                top.destroy()
            except (tk.TclError, Exception):
                pass
            self.apply_btn.config(state=tk.DISABLED)
            self.last_result = None
            self._refresh_trading212_linked_portfolio_async("Obnovuji aktuální stav koláče z Trading 212...")

        def worker():
            payload = None
            update_response = None
            verified_detail = None
            verify_error = None
            restore_error = None
            try:
                update_response = {
                    "stage": "building_payload",
                    "environment": environment,
                    "source_label": link.get("source_label"),
                }
                self._save_trading212_pie_update_debug(link["pie_id"], {}, update_response=update_response)
                payload, _ = self._t212_pie_payload_from_result(result, client, link)
                payload = dict(payload)
                update_response = {
                    "stage": "payload_prepared",
                    "environment": environment,
                    "payload_key_count": len(payload.get("instrumentShares", {}) or {}),
                    "positive_key_count": sum(
                        1 for value in (payload.get("instrumentShares", {}) or {}).values() if float(value) > 0
                    ),
                }
                self._save_trading212_pie_update_debug(link["pie_id"], payload, update_response=update_response)
                
                try:
                    current_detail = client.pie_detail(link["pie_id"], retries=2, retry_wait_seconds=4.0)
                    current_name = str(current_detail.get("settings", {}).get("name") or "")
                except Exception:
                    current_name = str(link.get("pie_settings", {}).get("name") or payload.get("name") or "")

                clean_name = self._clean_temporary_t212_pie_name(current_name).replace('\u200B', '')
                
                if current_name.endswith('\u200B'):
                    payload["name"] = clean_name
                else:
                    payload["name"] = clean_name + '\u200B'
                original_name = clean_name
                expected = self._normalize_pie_shares(payload.get("instrumentShares", {}))
                self.log(
                    "Trading 212 payload cílových vah: "
                    + ", ".join(f"{ticker} {weight * 100:.2f}%" for ticker, weight in list(expected.items())[:12])
                    + (f" ... +{len(expected) - 12} dalších" if len(expected) > 12 else "")
                )
                update_response = {
                    "stage": "prepared",
                    "environment": environment,
                    "payload_key_count": len(payload.get("instrumentShares", {}) or {}),
                    "positive_key_count": sum(
                        1 for value in (payload.get("instrumentShares", {}) or {}).values() if float(value) > 0
                    ),
                }
                self._save_trading212_pie_update_debug(link["pie_id"], payload, update_response=update_response)
                self.log(f"Trading 212: posílám nové cílové váhy přímo do koláče '{original_name}'.")
                
                target_response = client.update_pie(link["pie_id"], payload)
                
                update_response.update({
                    "stage": "target_sent",
                    "target_payload": payload,
                    "target_response": target_response,
                })
                self._save_trading212_pie_update_debug(link["pie_id"], payload, update_response=update_response)

                verified_detail = target_response if self._extract_t212_pie_shares(target_response) else None
                if verified_detail is None:
                    time.sleep(6.0)
                    try:
                        verified_detail = client.pie_detail(link["pie_id"], retries=2, retry_wait_seconds=4.0)
                    except Exception as exc:
                        verify_error = exc

                observed_after_target = self._extract_t212_pie_shares(verified_detail)
                if not self._pie_shares_match(expected, observed_after_target):
                    try:
                        latest_detail = client.pie_detail(link["pie_id"], retries=2, retry_wait_seconds=4.0)
                        if isinstance(latest_detail, dict) and latest_detail:
                            verified_detail = latest_detail
                    except Exception:
                        pass
                    
                    self._save_trading212_pie_update_debug(
                        link["pie_id"],
                        payload,
                        update_response=update_response,
                        verified_detail=verified_detail,
                        error=str(verify_error) if verify_error else None,
                    )
                    self.ui_queue.put(("call", finish, (payload, update_response, verified_detail, verify_error, None, None)))
                    return
                self.ui_queue.put(("call", finish, (payload, update_response, verified_detail, verify_error, restore_error, None)))
            except Exception as exc:
                self.log(f"Trading 212 worker chyba: {exc}")
                self._save_trading212_pie_update_debug(
                    link["pie_id"],
                    payload or {},
                    update_response=update_response or {"stage": "worker_error_before_payload"},
                    verified_detail=verified_detail,
                    error=str(exc),
                )
                self.ui_queue.put(("call", finish, (payload, update_response, verified_detail, verify_error, restore_error, exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _open_apply_trades_dialog(self):
        if self._blocked_action("Aplikace optimalizovaného portfolia"):
            return
        if not self.last_result: 
            self.log("Varování: Žádné výsledky k aplikaci.")
            return
        linked_t212 = self._is_trading212_linked_portfolio()
        linked_pie = self._linked_trading212_pie()
        
        top = tk.Toplevel(self.root)
        top.title("Detail optimalizovaného portfolia")
        top.geometry("1000x750")
        top.configure(padx=10, pady=10)

        def close_dialog():
            try:
                top.destroy()
            finally:
                self._refresh_apply_button_state()

        top.protocol("WM_DELETE_WINDOW", close_dialog)

        notebook = ttk.Notebook(top)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # --- TAB 1: TRADES ---
        trades_frame = ttk.Frame(notebook, padding=10)
        notebook.add(trades_frame, text="Rozdíl proti realitě" if linked_t212 else "Navržené obchody")
        
        cols_t = ("ticker", "action", "amount_czk", "price", "shares", "fee")
        tree_t = ttk.Treeview(trades_frame, columns=cols_t, show="headings", height=15)
        headings_t = {
            "ticker": "Ticker", "action": "Akce", "amount_czk": "Částka (CZK)", 
            "price": "Odhad. Cena", "shares": "Kusy (odhad)", "fee": "Poplatek"
        }
        for c, l in headings_t.items():
            tree_t.heading(c, text=l)
            tree_t.column(c, anchor=tk.CENTER, width=120)
        
        sb_t = ttk.Scrollbar(trades_frame, orient=tk.VERTICAL, command=tree_t.yview)
        tree_t.configure(yscrollcommand=sb_t.set)
        tree_t.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb_t.pack(side=tk.RIGHT, fill=tk.Y)

        # --- TAB 2: FULL TARGET PORTFOLIO ---
        full_frame = ttk.Frame(notebook, padding=10)
        notebook.add(full_frame, text="Cílové nové portfolio")
        
        cols_f = ("ticker", "name", "czk", "weight", "score", "pred", "q10", "q90")
        tree_f = ttk.Treeview(full_frame, columns=cols_f, show="headings", height=15)
        headings_f = {
            "ticker": "Ticker", "name": "Název", "czk": "Cílová hodnota (CZK)", 
            "weight": "Podíl %", "score": "Skóre", "pred": "Predikce 52w %", "q10": "Q10 52w %", "q90": "Q90 52w %"
        }
        for c, l in headings_f.items():
            tree_f.heading(c, text=l)
            tree_f.column(c, anchor=tk.CENTER, width=130)
            
        sb_f = ttk.Scrollbar(full_frame, orient=tk.VERTICAL, command=tree_f.yview)
        tree_f.configure(yscrollcommand=sb_f.set)
        tree_f.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb_f.pack(side=tk.RIGHT, fill=tk.Y)

        trades_to_process = []
        total_buy = 0.0
        total_sell = 0.0
        target_total = self.last_result.target_total_czk or 1e-9
        
        sorted_trades = sorted(self.last_result.trades, key=lambda x: x.target_czk, reverse=True)
        
        for trade in sorted_trades:
            meta = self.fetcher.cache.get(trade.ticker, {"name": trade.ticker, "price": 0.0})
            if trade.target_czk > 1.0:
                wgt = (trade.target_czk / target_total) * 100.0
                pred_52w, q10_52w, q90_52w = _record_like_horizon_values(trade)
                tree_f.insert("", tk.END, values=(
                    trade.ticker,
                    meta.get("name", trade.ticker),
                    f"{trade.target_czk:,.2f}",
                    f"{wgt:.2f}%",
                    _format_score(trade.score),
                    _format_return_pct(pred_52w),
                    _format_return_pct(q10_52w),
                    _format_return_pct(q90_52w)
                ))

            if trade.action != "KEEP":
                amt = abs(trade.delta_czk)
                fee = amt * 0.02
                price = meta.get("price", 0.0)
                shares = amt / price if price > 0 else 0
                
                trades_to_process.append({"ticker": trade.ticker, "action": trade.action, "amount": amt, "fee": fee})
                if trade.action == "BUY": total_buy += amt
                else: total_sell += amt
                
                tree_t.insert("", tk.END, values=(
                    trade.ticker, trade.action, f"{amt:,.2f} CZK", 
                    f"{price:.2f}" if price > 0 else "-", 
                    f"{shares:.2f}" if shares > 0 else "-", 
                    f"{fee:,.2f} CZK"
                ))

        summary_frame = ttk.LabelFrame(top, text="Souhrn a akce", padding=10)
        summary_frame.pack(fill=tk.X, pady=5)
        
        net_change = total_sell - total_buy
        sum_lbl = f"Celkem nákup: {total_buy:,.2f} CZK | Celkem prodej: {total_sell:,.2f} CZK | Čistá změna hotovosti: {net_change:,.2f} CZK"
        ttk.Label(summary_frame, text=sum_lbl, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=10)
        if linked_pie:
            ttk.Label(
                summary_frame,
                text="Trading 212 koláč: manager změní jen cílové váhy; skutečný stav se pak načítá z brokeru.",
                foreground="#7a4b00",
            ).pack(side=tk.LEFT, padx=10)
        elif linked_t212:
            ttk.Label(
                summary_frame,
                text="Trading 212 portfolio: ruční aplikace obchodů je vypnutá; stav se obnoví při načtení portfolia.",
                foreground="#7a4b00",
            ).pack(side=tk.LEFT, padx=10)

        def apply_all():
            if self._blocked_action("Aplikace obchodů"):
                return
            if linked_pie:
                self._apply_trading212_pie_weights_async(self.last_result, top)
                return
            if linked_t212:
                messagebox.showinfo(
                    "Trading 212",
                    "Aktuální stav Trading 212 se obnoví automaticky při načtení uloženého portfolia.",
                    parent=top,
                )
                return
            if not messagebox.askyesno("Potvrzení", f"Opravdu si přejete aplikovat tyto obchody? Čistá změna hotovosti bude {net_change:,.2f} CZK."):
                return
            try:
                total_cash_change = 0.0
                for data in trades_to_process:
                    t, act, amt, fee = normalize_ticker_symbol(data["ticker"]), data["action"], data["amount"], data["fee"]
                    if act == "BUY":
                        self.portfolio_dict[t] = self.portfolio_dict.get(t, 0.0) + amt
                        total_cash_change -= (amt + fee)
                    elif act.startswith("SELL"):
                        val = min(self.portfolio_dict.get(t, 0.0), amt)
                        self.portfolio_dict[t] -= val
                        if self.portfolio_dict[t] < 1.0: self.portfolio_dict.pop(t, None)
                        total_cash_change += (val - fee)
                
                new_cash = float(self.extra_cash_var.get()) + total_cash_change
                self.extra_cash_var.set(f"{new_cash:.2f}")
                self._append_current_equity_snapshot(replace_recent=True)
                
                self._refresh_portfolio_tree()
                self._save_portfolio()
                messagebox.showinfo("Hotovo", "Obchody byly úspěšně aplikovány.")
                top.destroy()
                self.apply_btn.config(state=tk.DISABLED)
                self.last_result = None
            except Exception as e:
                messagebox.showerror("Chyba", f"Nepodařilo se aplikovat obchody: {e}")

        btn_frame = ttk.Frame(top)
        btn_frame.pack(fill=tk.X, pady=5)
        ttk.Button(btn_frame, text="Zrušit", command=close_dialog).pack(side=tk.RIGHT, padx=5)
        if linked_pie:
            action_text = "Změnit cílové váhy v Trading 212"
        elif not linked_t212:
            action_text = "Aplikovat obchody do reálu"
        else:
            action_text = None
        if action_text:
            ttk.Button(btn_frame, text=action_text, command=apply_all).pack(side=tk.RIGHT, padx=5)


    def _clear_projections_history(self):
        if self._blocked_action("Smazání historie"):
            return
        self.portfolio_projections = []
        self.portfolio_history = []
        self._save_portfolio()
        self.log("Historie predikcí a equity vymazána z portfolia.")

    def _on_show_graph_btn_click(self):
        if self._blocked_action("Zobrazení grafu"):
            return
        self._start_chart_computation(None)

    def _start_chart_computation(self, result):
        if self._is_ui_locked():
            self.log("Nyní nelze spustit grafy.")
            return
        self.log("Připravuji denní historická data pro graf...")
        self._set_busy(True, "Připravuji denní data pro graf...")
        import threading

        # Tk variables and widgets must only be touched from the main thread.
        # Snapshot plain Python values here; the worker only downloads/calculates.
        chart_context = {
            "holdings": dict(self.portfolio_dict),
            "extra_cash": _safe_float(self.extra_cash_var.get(), default=0.0),
            "history": list(self.portfolio_history),
            "projections": list(self.portfolio_projections),
            "broker_link": dict(self.portfolio_broker_link) if isinstance(self.portfolio_broker_link, dict) else {},
            "broker_positions": dict(self.portfolio_broker_positions) if isinstance(self.portfolio_broker_positions, dict) else {},
        }

        def worker():
            try:
                chart_data = self._compute_chart_data(result, chart_context)
                self.ui_queue.put(("charts_ready", chart_data))
            except Exception as e:
                self.ui_queue.put(("log", f"Chyba grafů: {e}"))
                self.ui_queue.put(("charts_ready", None)) # To clear busy state
        threading.Thread(target=worker, daemon=True).start()

    def _price_history_cache_key(self, ticker: str) -> str:
        label = str(ticker or "").strip().upper().replace("^", "INDEX_")
        label = re.sub(r"[^A-Z0-9._-]+", "_", label).strip("_")
        return label or "UNKNOWN"

    def _price_history_cache_paths(self, ticker: str) -> tuple[Path, Path]:
        cache_dir = self.base_dir / "cache" / "price_history"
        cache_key = self._price_history_cache_key(ticker)
        return cache_dir / f"{cache_key}_1d.csv", cache_dir / f"{cache_key}_1d.meta.json"

    def _read_price_history_meta(self, ticker: str) -> dict:
        _, meta_path = self._price_history_cache_paths(ticker)
        try:
            if meta_path.exists():
                with open(meta_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    return data if isinstance(data, dict) else {}
        except Exception:
            return {}
        return {}

    def _normalize_close_series(self, pd, series) -> object:
        if series is None:
            return pd.Series(dtype=float)
        try:
            values = pd.to_numeric(series, errors="coerce")
            index = pd.DatetimeIndex(pd.to_datetime(values.index, utc=True, errors="coerce")).tz_convert(None)
            cleaned = pd.Series(values.to_numpy(), index=index).dropna()
            cleaned = cleaned[~cleaned.index.isna()]
            if cleaned.empty:
                return pd.Series(dtype=float)
            cleaned.index = cleaned.index.normalize()
            cleaned = cleaned.groupby(level=0).last().sort_index()
            cleaned = cleaned[cleaned.map(lambda value: math.isfinite(float(value)))]
            return cleaned.astype(float)
        except Exception:
            return pd.Series(dtype=float)

    def _read_price_history_cache(self, pd, ticker: str):
        csv_path, _ = self._price_history_cache_paths(ticker)
        if not csv_path.exists():
            return pd.Series(dtype=float)
        try:
            df = pd.read_csv(csv_path, parse_dates=["date"])
            if df.empty or "close" not in df.columns:
                return pd.Series(dtype=float)
            series = pd.Series(df["close"].to_numpy(), index=df["date"])
            return self._normalize_close_series(pd, series)
        except Exception as exc:
            self.log(f"Cache historických cen pro {ticker} se nepodařilo načíst: {exc}")
            return pd.Series(dtype=float)

    def _write_price_history_cache(self, pd, ticker: str, series, fetched_start, fetched_end):
        series = self._normalize_close_series(pd, series)
        if series.empty:
            return
        csv_path, meta_path = self._price_history_cache_paths(ticker)
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame({"date": series.index, "close": series.values}).to_csv(csv_path, index=False)
        meta = {
            "ticker": ticker,
            "interval": "1d",
            "auto_adjust": True,
            "fetched_start": pd.Timestamp(fetched_start).normalize().date().isoformat(),
            "fetched_end": pd.Timestamp(fetched_end).normalize().date().isoformat(),
            "updated_at": dt.datetime.now().isoformat(),
        }
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2, ensure_ascii=False)

    def _cached_daily_close_prices(self, pd, tickers: list[str], start, end):
        target_start = pd.Timestamp(start).normalize()
        target_end = pd.Timestamp(end).normalize()
        cropped = {}
        for ticker in tickers:
            ticker = normalize_ticker_symbol(ticker)
            if not ticker:
                continue
            series = self._normalize_close_series(pd, self._read_price_history_cache(pd, ticker))
            if series.empty:
                continue
            series = series[(series.index >= target_start) & (series.index < target_end)]
            if not series.empty:
                cropped[ticker] = series
        if not cropped:
            return pd.DataFrame()
        return pd.DataFrame(cropped).sort_index()

    def _extract_close_frame(self, pd, downloaded, tickers: list[str]):
        frame = pd.DataFrame()
        if downloaded is None or getattr(downloaded, "empty", True):
            return frame
        try:
            if isinstance(downloaded.columns, pd.MultiIndex):
                close = None
                for level in range(downloaded.columns.nlevels):
                    level_values = {str(value) for value in downloaded.columns.get_level_values(level)}
                    if "Close" in level_values:
                        close = downloaded.xs("Close", axis=1, level=level)
                        break
                    if "Adj Close" in level_values:
                        close = downloaded.xs("Adj Close", axis=1, level=level)
                        break
                if close is None:
                    return frame
                if isinstance(close, pd.Series):
                    frame = close.to_frame(name=tickers[0] if tickers else "Close")
                else:
                    frame = close.copy()
                    if isinstance(frame.columns, pd.MultiIndex):
                        frame.columns = [str(col[-1]) for col in frame.columns]
            else:
                if "Close" in downloaded.columns:
                    frame = downloaded["Close"].to_frame(name=tickers[0] if tickers else "Close")
                elif "Adj Close" in downloaded.columns:
                    frame = downloaded["Adj Close"].to_frame(name=tickers[0] if tickers else "Close")
        except Exception:
            return pd.DataFrame()

        rename = {}
        for column in list(frame.columns):
            column_label = str(column)
            for ticker in tickers:
                if column_label == ticker or normalize_ticker_symbol(column_label) == normalize_ticker_symbol(ticker):
                    rename[column] = ticker
                    break
        frame = frame.rename(columns=rename)
        return frame

    def _download_daily_close_prices_cached(self, yf, pd, tickers: list[str], start, end):
        requested = []
        for ticker in tickers:
            ticker = str(ticker or "").strip()
            if ticker and ticker not in requested:
                requested.append(ticker)
        if not requested:
            return pd.DataFrame()

        target_start = pd.Timestamp(start).normalize()
        target_end = pd.Timestamp(end).normalize()
        cached_series = {}
        to_download = []
        for ticker in requested:
            series = self._read_price_history_cache(pd, ticker)
            cached_series[ticker] = series
            meta = self._read_price_history_meta(ticker)
            try:
                fetched_start = pd.Timestamp(meta.get("fetched_start")).normalize()
                fetched_end = pd.Timestamp(meta.get("fetched_end")).normalize()
                covered = fetched_start <= target_start and fetched_end >= target_end
            except Exception:
                covered = False
            if series.empty or not covered:
                to_download.append(ticker)

        if to_download:
            self.log(
                f"Stahuji denní historii z Yahoo Finance pro {len(to_download)} z {len(requested)} symbolů "
                f"({target_start.date()} až {target_end.date()})."
            )
            chunk_size = 60
            for offset in range(0, len(to_download), chunk_size):
                chunk = to_download[offset:offset + chunk_size]
                try:
                    downloaded = yf.download(
                        chunk if len(chunk) > 1 else chunk[0],
                        start=target_start,
                        end=target_end,
                        interval="1d",
                        auto_adjust=True,
                        progress=False,
                        group_by="ticker",
                        threads=True,
                    )
                    close_frame = self._extract_close_frame(pd, downloaded, chunk)
                except Exception as exc:
                    self.log(f"Yahoo denní historie selhala pro část symbolů: {exc}")
                    close_frame = pd.DataFrame()

                for ticker in chunk:
                    downloaded_series = (
                        self._normalize_close_series(pd, close_frame[ticker])
                        if ticker in close_frame.columns else pd.Series(dtype=float)
                    )
                    if downloaded_series.empty:
                        continue
                    existing_series = self._normalize_close_series(pd, cached_series.get(ticker))
                    pieces = []
                    if not existing_series.empty:
                        pieces.append(existing_series)
                    if not downloaded_series.empty:
                        pieces.append(downloaded_series)
                    if not pieces:
                        continue
                    merged = pieces[0] if len(pieces) == 1 else pd.concat(pieces)
                    merged = self._normalize_close_series(pd, merged)
                    cached_series[ticker] = merged
                    try:
                        self._write_price_history_cache(pd, ticker, merged, target_start, target_end)
                    except Exception as exc:
                        self.log(f"Cache historických cen pro {ticker} se nepodařilo uložit: {exc}")
            self.log("Cache denních historických cen je aktualizovaná.")
        else:
            self.log(f"Denní historická data pro graf načtena z cache ({len(requested)} symbolů).")

        cropped = {}
        for ticker, series in cached_series.items():
            series = self._normalize_close_series(pd, series)
            if series.empty:
                continue
            series = series[(series.index >= target_start) & (series.index < target_end)]
            if not series.empty:
                cropped[ticker] = series
        if not cropped:
            return pd.DataFrame()
        return pd.DataFrame(cropped).sort_index()

    def _latest_yahoo_price(self, yf, ticker: str) -> float | None:
        try:
            info = getattr(yf.Ticker(ticker), "fast_info", {}) or {}
            for key in ("last_price", "lastPrice", "regular_market_price", "regularMarketPrice"):
                try:
                    value = info.get(key) if hasattr(info, "get") else getattr(info, key, None)
                except Exception:
                    value = None
                price = self._positive_float(value)
                if price:
                    return float(price)
        except Exception:
            pass
        try:
            cache = getattr(getattr(self, "fetcher", None), "cache", {}) or {}
            cached = cache.get(normalize_ticker_symbol(ticker), {}) if isinstance(cache, dict) else {}
            price = self._positive_float(cached.get("price")) if isinstance(cached, dict) else None
            if price:
                return float(price)
        except Exception:
            pass
        return None

    def _append_live_price_point(self, pd, yf, series, ticker: str, now: dt.datetime):
        series = self._normalize_close_series(pd, series)
        if series.empty:
            return series
        try:
            if series.index[-1].date() >= now.date():
                return series
        except Exception:
            return series

        live_price = self._latest_yahoo_price(yf, ticker)
        if not live_price:
            return series

        live_point = pd.Series([float(live_price)], index=pd.DatetimeIndex([pd.Timestamp(now)]))
        return pd.concat([series, live_point]).sort_index()

    def _build_offline_chart_data(self, pd, result, chart_context: dict | None, reason: str = ""):
        chart_context = chart_context or {}
        now = dt.datetime.now()

        def normalize_holdings_map(raw_holdings) -> dict[str, float]:
            if not isinstance(raw_holdings, dict):
                return {}
            holdings = {}
            for ticker, value in raw_holdings.items():
                try:
                    numeric = float(value)
                except (TypeError, ValueError):
                    continue
                if math.isfinite(numeric) and numeric > 0:
                    holdings[normalize_ticker_symbol(ticker)] = numeric
            return holdings

        current_holdings = normalize_holdings_map(chart_context.get("holdings"))
        extra_cash = _safe_float(chart_context.get("extra_cash", 0.0), default=0.0)
        current_total = sum(current_holdings.values()) + extra_cash
        portfolio_history = list(chart_context.get("history") or self.portfolio_history)
        portfolio_projections = list(chart_context.get("projections") or self.portfolio_projections)

        history_points_by_day: dict[object, dict] = {}

        def add_history_point(timestamp, value: float, flow: float = 0.0, is_snapshot: bool = False):
            try:
                numeric_value = float(value)
            except (TypeError, ValueError):
                return
            if not math.isfinite(numeric_value) or numeric_value <= 0.0:
                return
            day = pd.Timestamp(timestamp).normalize()
            key = day.to_pydatetime()
            existing = history_points_by_day.get(key)
            if existing is None or is_snapshot or not existing.get("_is_snapshot"):
                history_points_by_day[key] = {
                    "timestamp": key,
                    "value": numeric_value,
                    "flow": float(flow),
                    "_is_snapshot": bool(is_snapshot),
                }

        for item in portfolio_history:
            if not isinstance(item, dict) or not item.get("timestamp"):
                continue
            try:
                timestamp = dt.datetime.fromisoformat(str(item["timestamp"]))
                value = float(item.get("total_equity", 0.0))
                flow = _safe_float(item.get("external_flow_czk", 0.0), default=0.0)
                add_history_point(timestamp, value, flow, is_snapshot=True)
            except Exception:
                continue

        if current_total > 0.0:
            add_history_point(now, current_total, 0.0, is_snapshot=False)

        history_points = sorted(history_points_by_day.values(), key=lambda point: point["timestamp"])
        hist_dates = [point["timestamp"] for point in history_points]
        hist_equity = [float(point["value"]) for point in history_points]
        hist_external_flows = [float(point.get("flow", 0.0)) for point in history_points]

        hist_return_pct = []
        if hist_equity:
            twr_factor = 1.0
            previous_equity = max(hist_equity[0], 1e-9)
            hist_return_pct.append(0.0)
            for value, flow in zip(hist_equity[1:], hist_external_flows[1:]):
                adjusted_end = float(value) - float(flow)
                if previous_equity > 0 and adjusted_end > 0:
                    twr_factor *= adjusted_end / previous_equity
                hist_return_pct.append((twr_factor - 1.0) * 100.0)
                previous_equity = max(float(value), 1e-9)

        display_start = pd.Timestamp(now) - pd.DateOffset(years=1)
        display_candidates = [pd.Timestamp(value) for value in hist_dates]
        for proj in portfolio_projections:
            if not isinstance(proj, dict):
                continue
            try:
                display_candidates.append(pd.Timestamp(dt.datetime.fromisoformat(str(proj.get("date", "")))))
            except Exception:
                pass
        if display_candidates:
            earliest = min(display_candidates)
            if earliest < display_start:
                display_start = earliest - pd.DateOffset(months=1)

        warnings = []
        if reason:
            warnings.append(
                f"Graf běží v omezeném režimu: {reason}. Zobrazuji uložené snapshoty a predikce bez benchmarku."
            )
        if not hist_dates and not portfolio_projections:
            warnings.append("Graf nemá uloženou historii ani predikce, takže zatím není co vykreslit.")

        return {
            "hist_dates": hist_dates,
            "hist_equity": hist_equity,
            "hist_return_pct": hist_return_pct,
            "hist_external_flows": hist_external_flows,
            "bench_dates": [],
            "bench_values": [],
            "bench_return_pct": [],
            "synth_dates": [],
            "synth_equity": [],
            "synth_return_pct": [],
            "current_total": current_total,
            "now": now,
            "display_start": display_start.to_pydatetime(),
            "projections": portfolio_projections,
            "current_projection": None,
            "chart_warnings": warnings,
            "offline_mode": True,
        }

    def _compute_chart_data(self, result, chart_context: dict | None = None):
        np = _load_numpy()
        pd = _load_pandas()
        chart_context = chart_context or {}
        yf = _load_yfinance()

        if not yf:
            return self._build_offline_chart_data(pd, result, chart_context, "yfinance není dostupné")

        raw_current_holdings = dict(chart_context.get("holdings") or self.portfolio_dict)

        def normalize_holdings_map(raw_holdings) -> dict[str, float]:
            if not isinstance(raw_holdings, dict):
                return {}
            holdings = {}
            for ticker, value in raw_holdings.items():
                try:
                    numeric = float(value)
                except (TypeError, ValueError):
                    continue
                if math.isfinite(numeric) and numeric > 0:
                    holdings[normalize_ticker_symbol(ticker)] = numeric
            return holdings

        current_holdings = normalize_holdings_map(raw_current_holdings)
        extra_cash = float(chart_context.get("extra_cash", 0.0))
        portfolio_history = list(chart_context.get("history") or self.portfolio_history)
        portfolio_projections = list(chart_context.get("projections") or self.portfolio_projections)
        broker_link = chart_context.get("broker_link")
        if not isinstance(broker_link, dict):
            broker_link = self.portfolio_broker_link if isinstance(self.portfolio_broker_link, dict) else {}
        broker_positions = chart_context.get("broker_positions")
        if not isinstance(broker_positions, dict):
            broker_positions = self.portfolio_broker_positions if isinstance(self.portfolio_broker_positions, dict) else {}
        is_degiro_chart = broker_link.get("broker") == "Degiro" or any(
            isinstance(detail, dict) and detail.get("broker") == "Degiro"
            for detail in broker_positions.values()
        )
        current_total = sum(current_holdings.values()) + extra_cash

        valid_tickers = list(current_holdings.keys())
        now = dt.datetime.now()
        end_date = pd.Timestamp(now).normalize() + pd.Timedelta(days=1)
        oldest_allowed = end_date - pd.DateOffset(years=10)
        start_date = end_date - pd.DateOffset(years=5)

        context_dates = []
        for h in portfolio_history:
            if isinstance(h, dict) and h.get("timestamp"):
                try:
                    context_dates.append(pd.Timestamp(dt.datetime.fromisoformat(str(h["timestamp"]))))
                except Exception:
                    pass
        for proj in portfolio_projections:
            if not isinstance(proj, dict):
                continue
            for raw_date in [proj.get("date"), *list(proj.get("dates", []) if isinstance(proj.get("dates", []), list) else [])]:
                try:
                    context_dates.append(pd.Timestamp(dt.datetime.fromisoformat(str(raw_date))))
                except Exception:
                    pass
        if context_dates:
            start_date = min(start_date, min(context_dates) - pd.DateOffset(months=3))
        if start_date < oldest_allowed:
            start_date = oldest_allowed

        history_holdings_for_download: list[str] = []
        for h in portfolio_history:
            if isinstance(h, dict):
                history_holdings_for_download.extend(normalize_holdings_map(h.get("holdings")).keys())

        chart_tickers = list(valid_tickers) + history_holdings_for_download
        if result is not None:
            try:
                chart_tickers.extend(t.ticker for t in result.trades if getattr(t, "target_czk", 0.0) > 0)
            except Exception:
                pass
        for proj in portfolio_projections:
            if isinstance(proj, dict) and isinstance(proj.get("holdings"), dict):
                chart_tickers.extend(str(ticker) for ticker in proj["holdings"].keys())

        bench_tickers = ["^GSPC"]
        all_tickers = list(dict.fromkeys([normalize_ticker_symbol(ticker) for ticker in chart_tickers if str(ticker).strip()] + bench_tickers))
        data = self._download_daily_close_prices_cached(yf, pd, all_tickers, start_date, end_date)
        if data.empty:
            return self._build_offline_chart_data(pd, result, chart_context, "historické ceny nejsou dostupné")

        for ticker in [t for t in all_tickers if t != "^GSPC"]:
            if ticker not in data.columns:
                continue
            live_series = self._append_live_price_point(pd, yf, data[ticker], ticker, now)
            if live_series.empty:
                continue
            combined_index = data.index.union(live_series.index).sort_values()
            if len(combined_index) != len(data.index):
                data = data.reindex(combined_index)
            data[ticker] = live_series.reindex(data.index)
            
        sp500 = data["^GSPC"].dropna() if "^GSPC" in data.columns else None
        if sp500 is not None and not sp500.empty:
            sp500 = self._append_live_price_point(pd, yf, sp500, "^GSPC", now)

        def clean_snapshot_holdings(raw_holdings) -> dict[str, float]:
            return normalize_holdings_map(raw_holdings)

        history_snapshots: list[dict] = []
        if portfolio_history:
            for h in portfolio_history:
                if not isinstance(h, dict) or "timestamp" not in h or "total_equity" not in h:
                    continue
                try:
                    ts = dt.datetime.fromisoformat(str(h["timestamp"]))
                    value = float(h["total_equity"])
                    if not math.isfinite(value) or value <= 0:
                        continue
                    flow = _safe_float(h.get("external_flow_czk", 0.0), default=0.0)
                    snapshot = {
                        "timestamp": ts,
                        "value": value,
                        "flow": flow,
                        "holdings": clean_snapshot_holdings(h.get("holdings")),
                        "cash": _safe_float(h.get("extra_cash_czk", h.get("cash_czk", extra_cash)), default=extra_cash),
                    }
                    history_snapshots.append(snapshot)
                except Exception:
                    pass

        history_snapshots.sort(key=lambda item: item["timestamp"])

        def anchor_price(ticker: str, anchor_day):
            if ticker not in data.columns:
                return None
            series = self._normalize_close_series(pd, data[ticker])
            if series.empty:
                return None
            before = series[series.index <= anchor_day]
            if not before.empty:
                return float(before.iloc[-1])
            after = series[series.index >= anchor_day]
            if not after.empty:
                return float(after.iloc[0])
            return None

        def price_at_or_before(ticker: str, timestamp):
            if ticker not in data.columns:
                return None
            series = self._normalize_close_series(pd, data[ticker])
            if series.empty:
                return None
            target = pd.Timestamp(timestamp)
            before = series[series.index <= target]
            if not before.empty:
                return float(before.iloc[-1])
            after = series[series.index >= target]
            if not after.empty:
                return float(after.iloc[0])
            return None

        def snapshot_value_at(snapshot: dict, timestamp) -> float | None:
            holdings = snapshot.get("holdings", {}) if isinstance(snapshot.get("holdings"), dict) else {}
            if not holdings:
                return None
            cash = float(snapshot.get("cash", 0.0) or 0.0)
            base_total = cash + sum(float(value) for value in holdings.values())
            if base_total <= 0:
                return None
            scale = float(snapshot["value"]) / base_total
            anchor_day = pd.Timestamp(snapshot["timestamp"]).normalize()
            total = float(cash)
            for ticker, amount in holdings.items():
                amount = float(amount)
                price_at_anchor = anchor_price(ticker, anchor_day)
                latest_price = price_at_or_before(ticker, timestamp)
                if not price_at_anchor or not latest_price or price_at_anchor <= 0:
                    total += amount
                    continue
                total += amount * (float(latest_price) / float(price_at_anchor))
            return float(total * scale) if math.isfinite(total * scale) else None

        def daily_values_from_snapshot(snapshot: dict, daily_index, next_snapshot: dict | None = None):
            daily_index = pd.DatetimeIndex(daily_index)
            if daily_index.empty:
                return pd.Series(dtype=float)
            holdings = snapshot.get("holdings", {}) if isinstance(snapshot.get("holdings"), dict) else {}
            if not holdings:
                if next_snapshot is None:
                    return pd.Series(float(snapshot["value"]), index=daily_index, dtype=float)
                start_ts = pd.Timestamp(snapshot["timestamp"])
                end_ts = pd.Timestamp(next_snapshot["timestamp"])
                total_seconds = max((end_ts - start_ts).total_seconds(), 1.0)
                start_value = float(snapshot["value"])
                end_value = float(next_snapshot["value"])
                values = []
                for day in daily_index:
                    fraction = min(1.0, max(0.0, (pd.Timestamp(day) - start_ts).total_seconds() / total_seconds))
                    values.append(start_value + (end_value - start_value) * fraction)
                return pd.Series(values, index=daily_index, dtype=float)
            cash = float(snapshot.get("cash", 0.0) or 0.0)
            base_total = cash + sum(float(value) for value in holdings.values())
            scale = float(snapshot["value"]) / base_total if base_total > 0 else 1.0
            values = pd.Series(float(cash), index=daily_index, dtype=float)
            extended_index = data.index.union(daily_index).sort_values()
            anchor_day = pd.Timestamp(snapshot["timestamp"]).normalize()
            for ticker, amount in holdings.items():
                amount = float(amount)
                price_at_anchor = anchor_price(ticker, anchor_day)
                if not price_at_anchor or price_at_anchor <= 0 or ticker not in data.columns:
                    values = values.add(amount, fill_value=0.0)
                    continue
                prices = (
                    self._normalize_close_series(pd, data[ticker])
                    .reindex(extended_index)
                    .ffill()
                    .reindex(daily_index)
                )
                ratios = (prices / price_at_anchor).replace([np.inf, -np.inf], np.nan).fillna(1.0)
                values = values.add(ratios.astype(float) * amount, fill_value=0.0)
            return values * scale

        if is_degiro_chart and history_snapshots:
            latest_snapshot = history_snapshots[-1]
            estimated_current_total = snapshot_value_at(latest_snapshot, now)
            if estimated_current_total and estimated_current_total > 0:
                current_total = float(estimated_current_total)

        live_snapshot = {
            "timestamp": now,
            "value": float(current_total),
            "flow": 0.0,
            "holdings": dict(current_holdings),
            "cash": extra_cash,
        }
        should_append_live = True
        if history_snapshots:
            latest = history_snapshots[-1]
            try:
                same_moment = abs((now - latest["timestamp"]).total_seconds()) < 60
                same_value = abs(float(latest["value"]) - float(current_total)) < 1.0
                same_holdings = latest.get("holdings") == live_snapshot["holdings"] and abs(
                    float(latest.get("cash", 0.0)) - extra_cash
                ) < 1.0
                should_append_live = not (same_moment and same_value and same_holdings)
            except Exception:
                should_append_live = True
        if should_append_live:
            history_snapshots.append(live_snapshot)

        history_points_by_day: dict[object, dict] = {}

        def add_history_point(timestamp, value: float, flow: float = 0.0, is_snapshot: bool = False):
            if not math.isfinite(float(value)):
                return
            day = pd.Timestamp(timestamp).normalize()
            key = day.to_pydatetime()
            existing = history_points_by_day.get(key)
            if existing is None or is_snapshot or not existing.get("_is_snapshot"):
                history_points_by_day[key] = {
                    "timestamp": key,
                    "value": float(value),
                    "flow": float(flow),
                    "_is_snapshot": bool(is_snapshot),
                }

        for idx, snapshot in enumerate(history_snapshots):
            add_history_point(
                snapshot["timestamp"],
                float(snapshot["value"]),
                float(snapshot.get("flow", 0.0)),
                is_snapshot=True,
            )
            if idx + 1 >= len(history_snapshots):
                continue
            anchor_day = pd.Timestamp(snapshot["timestamp"]).normalize()
            next_day = pd.Timestamp(history_snapshots[idx + 1]["timestamp"]).normalize()
            daily_index = data.index[(data.index > anchor_day) & (data.index < next_day)]
            daily_values = daily_values_from_snapshot(snapshot, daily_index, history_snapshots[idx + 1])
            for day, value in daily_values.items():
                add_history_point(day, float(value), 0.0, is_snapshot=False)

        history_points = sorted(history_points_by_day.values(), key=lambda point: point["timestamp"])
        hist_dates = [point["timestamp"] for point in history_points]
        hist_equity = [float(point["value"]) for point in history_points]
        hist_external_flows = [float(point.get("flow", 0.0)) for point in history_points]

        hist_return_pct = []
        if hist_equity:
            twr_factor = 1.0
            previous_equity = max(hist_equity[0], 1e-9)
            hist_return_pct.append(0.0)
            for value, flow in zip(hist_equity[1:], hist_external_flows[1:]):
                adjusted_end = float(value) - float(flow)
                if previous_equity > 0 and adjusted_end > 0:
                    twr_factor *= adjusted_end / previous_equity
                hist_return_pct.append((twr_factor - 1.0) * 100.0)
                previous_equity = max(float(value), 1e-9)

        bench_dates = []
        bench_values = []
        bench_return_pct = []
        if sp500 is not None and not sp500.empty:
            bench_dates = sp500.index.tolist()
            if hist_dates:
                first_date = pd.Timestamp(hist_dates[0])
                if first_date < sp500.index[0]: first_date = sp500.index[0]
                if first_date > sp500.index[-1]: first_date = sp500.index[-1]
                
                idx = sp500.index.get_indexer([first_date], method='nearest')[0]
                anchor_val_bench = sp500.iloc[idx]
                bench_values = (sp500 / anchor_val_bench * hist_equity[0]).tolist()
                bench_return_pct = ((sp500 / anchor_val_bench - 1.0) * 100.0).tolist()
            else:
                bench_values = (sp500 / sp500.iloc[-1] * current_total).tolist()
                bench_return_pct = ((sp500 / sp500.iloc[-1] - 1.0) * 100.0).tolist()

        anchor_dt = hist_dates[0] if hist_dates else now
        anchor_val = hist_equity[0] if hist_equity else current_total
        synthetic_anchor_snapshot = next(
            (snapshot for snapshot in history_snapshots if snapshot.get("holdings")),
            None,
        )

        synth_dates = []
        synth_equity = []
        synth_return_pct = []
        if synthetic_anchor_snapshot:
            try:
                synth_holdings = dict(synthetic_anchor_snapshot.get("holdings") or {})
                synth_cash = float(synthetic_anchor_snapshot.get("cash", 0.0) or 0.0)
                synth_tickers = list(synth_holdings.keys())
                available_tickers = [ticker for ticker in synth_tickers if ticker in data.columns]
                if available_tickers:
                    anchor_dt = synthetic_anchor_snapshot["timestamp"]
                    anchor_val = float(synthetic_anchor_snapshot["value"])
                    anchor_day = pd.Timestamp(anchor_dt).normalize()
                    raw = data[available_tickers].loc[data.index < anchor_day].copy()
                    if isinstance(raw, pd.Series): raw = raw.to_frame(name=available_tickers[0])
                    port_series = None
                    for t in synth_tickers:
                        if t in raw.columns:
                            s = raw[t].dropna()
                            if s.empty: continue
                            rel = s / s.iloc[-1]
                            weight = synth_holdings.get(t, 0.0)
                            contrib = rel * weight
                            port_series = contrib if port_series is None else port_series.add(contrib, fill_value=0.0)
                    
                    if port_series is not None:
                        scale = anchor_val / (port_series.iloc[-1] + synth_cash) if (port_series.iloc[-1] + synth_cash) > 0 else 1.0
                        synth_dates = port_series.index.tolist()
                        synth_equity = ((port_series + synth_cash) * scale).tolist()
                        if synth_dates and pd.Timestamp(synth_dates[-1]) < pd.Timestamp(anchor_dt):
                            synth_dates.append(anchor_dt)
                            synth_equity.append(float(anchor_val))
                        if anchor_val > 0:
                            synth_return_pct = [((float(value) / anchor_val) - 1.0) * 100.0 for value in synth_equity]
            except: pass

        def enrich_projection_actuals(proj):
            if not isinstance(proj, dict):
                return proj
            enriched = dict(proj)
            holdings = normalize_holdings_map(enriched.get("holdings"))
            recovered_cash = None
            start_value = _safe_float(enriched.get("start_value"), default=0.0)
            try:
                projection_dt = dt.datetime.fromisoformat(str(enriched.get("date", "")))
            except Exception:
                return enriched
            if not holdings and self._projection_source(enriched) == "current":
                nearby_snapshots = [
                    snapshot
                    for snapshot in history_snapshots
                    if snapshot.get("holdings")
                    and abs((snapshot["timestamp"] - projection_dt).total_seconds()) <= 6 * 3600
                ]
                if nearby_snapshots:
                    nearest = min(
                        nearby_snapshots,
                        key=lambda snapshot: abs((snapshot["timestamp"] - projection_dt).total_seconds()),
                    )
                    holdings = normalize_holdings_map(nearest.get("holdings"))
                    recovered_cash = float(nearest.get("cash", 0.0) or 0.0)
            if not holdings or start_value <= 0.0:
                return enriched

            anchor_day = pd.Timestamp(projection_dt).normalize()
            now_ts = pd.Timestamp(now)
            daily_index = data.index[(data.index >= anchor_day) & (data.index <= now_ts)]
            if daily_index.empty:
                return enriched

            holdings_total = sum(float(value) for value in holdings.values())
            cash = max(0.0, float(recovered_cash)) if recovered_cash is not None else max(0.0, start_value - holdings_total)
            base_total = cash + holdings_total
            scale = start_value / base_total if base_total > 0.0 else 1.0
            values = pd.Series(float(cash), index=daily_index, dtype=float)
            extended_index = data.index.union(daily_index).sort_values()

            for ticker, amount in holdings.items():
                amount = float(amount)
                if ticker not in data.columns:
                    values = values.add(amount, fill_value=0.0)
                    continue
                price_at_anchor = anchor_price(ticker, anchor_day)
                if not price_at_anchor or price_at_anchor <= 0:
                    values = values.add(amount, fill_value=0.0)
                    continue
                prices = (
                    self._normalize_close_series(pd, data[ticker])
                    .reindex(extended_index)
                    .ffill()
                    .reindex(daily_index)
                )
                ratios = (prices / price_at_anchor).replace([np.inf, -np.inf], np.nan).fillna(1.0)
                values = values.add(ratios.astype(float) * amount, fill_value=0.0)

            values = (values * scale).dropna()
            if values.empty:
                return enriched

            actual_dates = [projection_dt]
            actual_values = [float(start_value)]
            projection_ts = pd.Timestamp(projection_dt)
            for day, value in values.items():
                day_ts = pd.Timestamp(day)
                if day_ts <= projection_ts:
                    continue
                actual_day_dt = day_ts.to_pydatetime().replace(
                    hour=projection_dt.hour,
                    minute=projection_dt.minute,
                    second=projection_dt.second,
                    microsecond=projection_dt.microsecond,
                )
                if actual_day_dt <= projection_dt:
                    actual_day_dt = actual_day_dt + dt.timedelta(days=1)
                actual_dates.append(actual_day_dt)
                actual_values.append(float(value))

            if len(actual_dates) > 1:
                enriched["actual_dates"] = [value.isoformat() for value in actual_dates]
                enriched["actual_values"] = actual_values
            return enriched

        enriched_projections = [
            enrich_projection_actuals(proj)
            for proj in portfolio_projections
        ]

        return {
            "hist_dates": hist_dates,
            "hist_equity": hist_equity,
            "hist_return_pct": hist_return_pct,
            "hist_external_flows": hist_external_flows,
            "bench_dates": bench_dates,
            "bench_values": bench_values,
            "bench_return_pct": bench_return_pct,
            "synth_dates": synth_dates,
            "synth_equity": synth_equity,
            "synth_return_pct": synth_return_pct,
            "current_total": current_total,
            "now": now,
            "display_start": (pd.Timestamp(now) - pd.DateOffset(years=1)).to_pydatetime(),
            "projections": enriched_projections,
            "current_projection": None
        }

    def _build_projection_from_result(self, result: OptimizationResult | None, total_val: float):
        if result is None:
            return None
        import datetime
        horizons = [1, 4, 13, 26, 52]
        valid_trades = [t for t in result.trades if t.target_czk > 0]
        ticker_list = [t.ticker for t in valid_trades]
        total_target = sum(t.target_czk for t in valid_trades) or 1e-9
        w_list = [t.target_czk / total_target for t in valid_trades]
        
        corr_matrix = build_correlation_matrix(ticker_list, self.return_cache, self.log)

        expected_by_horizon = {}
        q10_by_horizon = {}
        q90_by_horizon = {}

        for w in horizons:
            p_list = []
            c10_list = []
            q90_list = []
            
            for trade in valid_trades:
                hdata = trade.horizon_data or {}
                default_p = trade.forecast_pct
                p, s, u, _ = _coerce_horizon_values(
                    hdata.get(f"{w}w"),
                    (default_p, default_p, default_p, 0.0),
                )
                p_list.append(float(p))
                c10_list.append(float(s))
                q90_list.append(float(u))

            p_port, q10_port, q90_port = _diversified_portfolio_quantiles(
                ticker_list, w_list, p_list, c10_list, q90_list, corr_matrix
            )
            expected_by_horizon[w] = p_port
            q10_by_horizon[w] = q10_port
            q90_by_horizon[w] = q90_port
                
        now = dt.datetime.now()
        dates = [(now + datetime.timedelta(weeks=w)).isoformat() for w in horizons]
        expected = [total_val * (1.0 + expected_by_horizon[w]) for w in horizons]
        q10_vals = [total_val * (1.0 + q10_by_horizon[w]) for w in horizons]
        q90_vals = [total_val * (1.0 + q90_by_horizon[w]) for w in horizons]
        
        return {
            "date": now.isoformat(), 
            "start_value": total_val, 
            "holdings": {normalize_ticker_symbol(t.ticker): float(max(0.0, t.target_czk)) for t in valid_trades},
            "dates": dates, 
            "expected": expected,
            "q10": q10_vals,
            "q90": q90_vals,
            "label": "Aktuální Návrh"
        }


    def _display_charts(self, chart_data):
        if not chart_data:
            self._set_busy(False, "Graf se nepodařilo připravit.")
            message = "Graf nelze zobrazit, protože nejsou dostupná historická data ani uložené predikce."
            self.log(message)
            messagebox.showwarning("Graf", message)
            return

        self._set_busy(False, "Grafy připraveny.")
        for warning in chart_data.get("chart_warnings", []) or []:
            self.log(str(warning))

        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from matplotlib.patches import FancyBboxPatch
        from matplotlib.widgets import CheckButtons
        import datetime
        import pandas as pd

        plt.rcParams['font.sans-serif'] = ['Segoe UI', 'Arial', 'sans-serif']
        plt.rcParams['axes.unicode_minus'] = False
        
        fig, ax = plt.subplots(figsize=(15, 8))
        fig.patch.set_facecolor('#fdfdfd')
        ax.set_facecolor('#ffffff')
        
        toggle_items = []
        used_toggle_labels = set()
        layer_state: dict[str, bool] = {}
        active_mode = {"value": "czk"}
        zero_line = ax.axhline(0, color="#777777", linewidth=1.0, linestyle=":", alpha=0.55, visible=False)
        history_anchors = {}
        for h_date, h_value, h_pct in zip(
            chart_data.get("hist_dates", []),
            chart_data.get("hist_equity", []),
            chart_data.get("hist_return_pct", []),
        ):
            try:
                anchor_dt = h_date.to_pydatetime() if hasattr(h_date, "to_pydatetime") else h_date
                if isinstance(anchor_dt, dt.datetime):
                    history_anchors[anchor_dt.date()] = (anchor_dt, float(h_value), float(h_pct))
            except Exception:
                pass

        def set_artist_visible(artists, visible):
            for artist in artists:
                try:
                    artist.set_visible(visible)
                except Exception:
                    pass

        def add_toggle(label, groups, visible=True, color=None, delete_identity=None):
            if not groups:
                return
            if not isinstance(groups, dict):
                groups = {"czk": list(groups), "pct": []}
            if not groups.get("czk") and not groups.get("pct"):
                return
            base = str(label).strip() or "Vrstva"
            final_label = base
            suffix = 2
            while final_label in used_toggle_labels:
                final_label = f"{base} #{suffix}"
                suffix += 1
            used_toggle_labels.add(final_label)
            layer_state[final_label] = bool(visible)
            for mode, artists in groups.items():
                set_artist_visible(artists, bool(visible) and mode == active_mode["value"])
            toggle_items.append((final_label, groups, visible, color or "#333333", delete_identity))

        def projection_label(proj, index):
            raw_label = str(proj.get("label", f"Predikce {index + 1}"))
            pretty = {
                "Původní portfolio": "Původní",
                "Optimalizované portfolio": "Optimalizované",
                "Aktuální Návrh": "Aktuální návrh",
            }.get(raw_label, raw_label)
            try:
                created = dt.datetime.fromisoformat(str(proj.get("date", ""))).strftime("%d.%m. %H:%M")
                return f"{created} | {pretty}"
            except Exception:
                return pretty

        def current_xlim():
            now_ts = pd.Timestamp(chart_data["now"])
            display_start = pd.Timestamp(chart_data.get("display_start", now_ts - pd.DateOffset(years=5)))
            return display_start, now_ts + pd.DateOffset(years=1)

        def autoscale_y_to_xlim():
            x0, x1 = ax.get_xlim()
            values = []

            def date_num(value):
                try:
                    return mdates.date2num(pd.Timestamp(value).to_pydatetime())
                except Exception:
                    try:
                        return float(value)
                    except Exception:
                        return None

            for artist in ax.get_children():
                try:
                    if not artist.get_visible():
                        continue
                except Exception:
                    continue
                if artist is zero_line and active_mode["value"] == "pct":
                    values.append(0.0)
                    continue
                if hasattr(artist, "get_xdata") and hasattr(artist, "get_ydata"):
                    try:
                        for x_raw, y_raw in zip(artist.get_xdata(), artist.get_ydata()):
                            x_num = date_num(x_raw)
                            y_val = float(y_raw)
                            if x_num is not None and x0 <= x_num <= x1 and math.isfinite(y_val):
                                values.append(y_val)
                    except Exception:
                        pass

            if not values:
                return
            ymin = min(values)
            ymax = max(values)
            if not math.isfinite(ymin) or not math.isfinite(ymax):
                return
            if abs(ymax - ymin) <= 1e-12:
                pad = max(abs(ymax) * 0.05, 1.0)
            else:
                pad = (ymax - ymin) * 0.06
            ax.set_ylim(ymin - pad, ymax + pad)

        def apply_mode():
            mode = active_mode["value"]
            for label, groups, _, _, _ in toggle_items:
                enabled = layer_state.get(label, True)
                for group_mode, artists in groups.items():
                    set_artist_visible(artists, enabled and group_mode == mode)
            zero_line.set_visible(mode == "pct")
            ax.set_ylabel("Hodnota v CZK" if mode == "czk" else "Cash-flow neutral zisk (%)", fontsize=12, labelpad=10)
            ax.set_title(
                "Vizualizace Vývoje Portfolia a AI Predikcí" if mode == "czk"
                else "Cash-flow neutral výkonnost a AI predikce",
                fontsize=18,
                pad=25,
                fontweight='bold',
                color='#333',
            )
            try:
                ax.relim(visible_only=True)
            except TypeError:
                ax.relim()
            ax.autoscale_view(scalex=False, scaley=True)
            ax.set_xlim(*current_xlim())
            autoscale_y_to_xlim()
            fig.canvas.draw_idle()

        if chart_data.get("synth_dates"):
            ln, = ax.plot(
                chart_data["synth_dates"],
                chart_data["synth_equity"],
                color='teal',
                linestyle='-',
                linewidth=1.6,
                alpha=0.32,
                label='Syntetická historie',
            )
            pct_group = []
            if chart_data.get("synth_return_pct"):
                ln_pct, = ax.plot(
                    chart_data["synth_dates"],
                    chart_data["synth_return_pct"],
                    color='teal',
                    linestyle='-',
                    linewidth=1.6,
                    alpha=0.32,
                    label='Syntetická historie (%)',
                )
                pct_group.append(ln_pct)
            add_toggle('Syntetická historie', {"czk": [ln], "pct": pct_group}, visible=True, color='teal')

        if chart_data.get("bench_dates"):
            ln, = ax.plot(
                chart_data["bench_dates"],
                chart_data["bench_values"],
                color='#555555',
                linestyle='-',
                linewidth=1.8,
                alpha=0.45,
                label='S&P 500 Benchmark',
            )
            pct_group = []
            if chart_data.get("bench_return_pct"):
                ln_pct, = ax.plot(
                    chart_data["bench_dates"],
                    chart_data["bench_return_pct"],
                    color='#555555',
                    linestyle='-',
                    linewidth=1.8,
                    alpha=0.45,
                    label='S&P 500 Benchmark (%)',
                )
                pct_group.append(ln_pct)
            add_toggle('S&P 500 Benchmark', {"czk": [ln], "pct": pct_group}, visible=True, color='#555555')

        if chart_data.get("hist_dates"):
            ln, = ax.plot(
                chart_data["hist_dates"],
                chart_data["hist_equity"],
                color='#0052cc',
                linewidth=2.5,
                marker='o' if len(chart_data.get("hist_dates", [])) <= 2 else None,
                markersize=5,
                label='Reálná historie',
                zorder=10,
            )
            pct_group = []
            if chart_data.get("hist_return_pct"):
                ln_pct, = ax.plot(
                    chart_data["hist_dates"],
                    chart_data["hist_return_pct"],
                    color='#0052cc',
                    linewidth=2.5,
                    marker='o' if len(chart_data.get("hist_dates", [])) <= 2 else None,
                    markersize=5,
                    label='Reálná historie (%)',
                    zorder=10,
                )
                pct_group.append(ln_pct)
            add_toggle('Reálná historie', {"czk": [ln], "pct": pct_group}, visible=True, color='#0052cc')

        colors = ['#28a745', '#fd7e14', '#6f42c1', '#e83e8c', '#20c997', '#ffc107']
        
        def plot_proj(proj, color, display_label, is_current=False, show_fan=True):
            try:
                raw_date = dt.datetime.fromisoformat(proj["date"])
                raw_start_val = float(proj["start_value"])
                if raw_start_val <= 0:
                    return {}
                p_date = raw_date
                p_val = raw_start_val
                scale = 1.0
                date_shift = datetime.timedelta(0)
                anchor_return_pct = 0.0
                anchor = history_anchors.get(raw_date.date())
                if anchor and raw_start_val > 0:
                    p_date, p_val, anchor_return_pct = anchor
                    scale = p_val / raw_start_val
                    date_shift = p_date - raw_date
                
                px = [p_date] + [dt.datetime.fromisoformat(d) + date_shift for d in proj.get("dates", [])]
                py = [p_val] + [float(v) * scale for v in proj.get("expected", [])]
                ratio_y = [1.0] + [float(v) / raw_start_val for v in proj.get("expected", [])]
                anchor_factor = 1.0 + (anchor_return_pct / 100.0)
                py_pct = [(anchor_factor * ratio - 1.0) * 100.0 for ratio in ratio_y]
                
                ln, = ax.plot(
                    px,
                    py,
                    color=color,
                    linestyle='--',
                    linewidth=2.4 if is_current else 1.8,
                    alpha=0.92 if is_current else 0.72,
                    label=display_label,
                    zorder=5 if is_current else 4,
                )
                ln_pct, = ax.plot(
                    px,
                    py_pct,
                    color=color,
                    linestyle='--',
                    linewidth=2.4 if is_current else 1.8,
                    alpha=0.92 if is_current else 0.72,
                    label=f"{display_label} (%)",
                    zorder=5 if is_current else 4,
                )
                
                czk_group = [ln]
                pct_group = [ln_pct]

                actual_dates = proj.get("actual_dates", [])
                actual_values = proj.get("actual_values", [])
                if isinstance(actual_dates, list) and isinstance(actual_values, list) and len(actual_dates) == len(actual_values):
                    actual_pairs = []
                    for raw_actual_date, raw_actual_value in zip(actual_dates, actual_values):
                        try:
                            actual_pairs.append((
                                dt.datetime.fromisoformat(str(raw_actual_date)) + date_shift,
                                float(raw_actual_value),
                            ))
                        except Exception:
                            continue
                    if len(actual_pairs) > 1:
                        actual_x = [item[0] for item in actual_pairs]
                        actual_raw_values = [item[1] for item in actual_pairs]
                        actual_y = [value * scale for value in actual_raw_values]
                        actual_pct = [
                            (anchor_factor * (value / raw_start_val) - 1.0) * 100.0
                            for value in actual_raw_values
                        ]
                        actual_ln, = ax.plot(
                            actual_x,
                            actual_y,
                            color=color,
                            linestyle='-',
                            linewidth=2.1,
                            alpha=0.9,
                            label=f"{display_label} skutečnost",
                            zorder=6,
                        )
                        actual_ln_pct, = ax.plot(
                            actual_x,
                            actual_pct,
                            color=color,
                            linestyle='-',
                            linewidth=2.1,
                            alpha=0.9,
                            label=f"{display_label} skutečnost (%)",
                            zorder=6,
                        )
                        czk_group.append(actual_ln)
                        pct_group.append(actual_ln_pct)
                
                if show_fan and "q10" in proj and "q90" in proj:
                    q10 = [p_val] + [float(v) * scale for v in proj["q10"]]
                    q90 = [p_val] + [float(v) * scale for v in proj["q90"]]
                    if len(q10) == len(px) and len(q90) == len(px):
                        fan = ax.fill_between(px, q10, q90, color=color, alpha=0.07)
                        czk_group.append(fan)
                    q10_pct = [(anchor_factor * ratio - 1.0) * 100.0 for ratio in ([1.0] + [float(v) / raw_start_val for v in proj["q10"]])]
                    q90_pct = [(anchor_factor * ratio - 1.0) * 100.0 for ratio in ([1.0] + [float(v) / raw_start_val for v in proj["q90"]])]
                    if len(q10_pct) == len(px) and len(q90_pct) == len(px):
                        fan_pct = ax.fill_between(px, q10_pct, q90_pct, color=color, alpha=0.07)
                        pct_group.append(fan_pct)
                
                if len(py) > 1:
                    pct_change = (ratio_y[-1] - 1.0) * 100
                    sign = "+" if pct_change >= 0 else ""
                    ann = ax.annotate(f"{sign}{pct_change:.1f}%", xy=(px[-1], py[-1]), xytext=(5, 0), 
                                      textcoords="offset points", color=color, fontsize=10, fontweight='bold')
                    ann_pct = ax.annotate(f"{sign}{pct_change:.1f}%", xy=(px[-1], py_pct[-1]), xytext=(5, 0),
                                          textcoords="offset points", color=color, fontsize=10, fontweight='bold')
                    czk_group.append(ann)
                    pct_group.append(ann_pct)
                
                return {"czk": czk_group, "pct": pct_group}
            except: return {}

        projections = self._chart_projections(chart_data.get("projections", []))
        visible_from = max(0, len(projections) - 4)
        for i, proj in enumerate(projections):
            display_label = projection_label(proj, i)
            default_visible = i >= visible_from
            color = colors[i % len(colors)]
            groups = plot_proj(proj, color, display_label, show_fan=True)
            projection_source = self._projection_source(proj)
            delete_identity = (
                self._projection_identity(proj)
                if projection_source in {"current", "optimized"}
                else None
            )
            add_toggle(display_label, groups, visible=default_visible, color=color, delete_identity=delete_identity)

        if not toggle_items:
            message = "Graf nemá žádnou vykreslitelnou vrstvu. Ulož portfolio nebo spusť optimalizaci a zkus graf znovu."
            self.log(message)
            messagebox.showwarning("Graf", message)
            plt.close(fig)
            return

        ax.grid(True, linestyle='--', alpha=0.5)
        date_locator = mdates.AutoDateLocator(minticks=5, maxticks=9)
        date_formatter = mdates.AutoDateFormatter(date_locator)
        date_formatter.defaultfmt = "%Y-%m-%d"
        date_formatter.scaled[365.0] = "%Y"
        date_formatter.scaled[30.0] = "%Y-%m"
        date_formatter.scaled[1.0] = "%d.%m."
        date_formatter.scaled[1.0 / 24.0] = "%d.%m. %H:%M"
        date_formatter.scaled[1.0 / (24.0 * 60.0)] = "%H:%M"
        ax.xaxis.set_major_locator(date_locator)
        ax.xaxis.set_major_formatter(date_formatter)

        ax.set_xlim(*current_xlim())

        if toggle_items:
            labels = [item[0] for item in toggle_items]
            visibility = [item[2] for item in toggle_items]
            label_colors = [item[3] for item in toggle_items]
            delete_identities = [item[4] for item in toggle_items]
            plt.subplots_adjust(right=0.96, left=0.08, top=0.92, bottom=0.16)
            panel_height = min(0.56, max(0.20, 0.07 + 0.035 * len(labels)))
            panel_width = min(0.50, max(0.32, 0.010 * max(len(label) for label in labels)))
            try:
                rax = ax.inset_axes([0.015, 0.965 - panel_height, panel_width, panel_height], zorder=20)
            except Exception:
                rax = fig.add_axes([0.095, 0.92 - panel_height, panel_width, panel_height])
            rax.set_facecolor('#f9f9f9')
            rax.patch.set_alpha(0.88)
            check = CheckButtons(rax, labels, visibility)
            color_key_artists = []
            check_labels = list(getattr(check, "labels", []))
            fig._portfolio_del_buttons = []
            
            for idx, color in enumerate(label_colors):
                label_text = labels[idx]
                try:
                    y = check_labels[idx].get_position()[1]
                    check_labels[idx].set_x(0.35)
                except Exception:
                    y = 1.0 - ((idx + 0.5) / max(len(label_colors), 1))
                ln, = rax.plot([0.20, 0.30], [y, y], color=color, linewidth=2.4, transform=rax.transAxes)
                color_key_artists.append(ln)

                # Add delete button next to saved current/optimized portfolio projections.
                delete_identity = delete_identities[idx]
                if delete_identity:
                    try:
                        from matplotlib.widgets import Button
                        check_labels[idx].set_x(0.33)
                        del_ax = rax.inset_axes([0.88, y - 0.030, 0.080, 0.060])
                        del_btn = Button(del_ax, 'X', color='#ffe6e6', hovercolor='#ffcccc')
                        del_btn.label.set_color('#cc0000')
                        del_btn.label.set_fontsize(8)
                        del_btn.label.set_fontweight('bold')
                        
                        def make_delete_callback(target_identity):
                            def delete_callback(event):
                                self.portfolio_projections = [
                                    p
                                    for p in self.portfolio_projections
                                    if self._projection_identity(p) != target_identity
                                ]
                                
                                chart_projs = chart_data.get("projections", [])
                                if isinstance(chart_projs, list):
                                    chart_data["projections"] = [
                                        p
                                        for p in chart_projs
                                        if self._projection_identity(p) != target_identity
                                    ]

                                try:
                                    self._write_portfolio_file(quiet=True)
                                except Exception as exc:
                                    self.log(f"Nepodařilo se uložit portfolio po smazání projekce: {exc}")
                                
                                plt.close(fig)
                                self._display_charts(chart_data)
                            return delete_callback
                        
                        del_btn.on_clicked(make_delete_callback(delete_identity))
                        fig._portfolio_del_buttons.append((del_btn, del_ax))
                    except Exception as e:
                        self.log(f"Chyba při vykreslení tlačítka smazat pro {label_text}: {e}")

            def toggle_visibility(label):
                if label in layer_state:
                    layer_state[label] = not layer_state[label]
                    apply_mode()

            check.on_clicked(toggle_visibility)
            for text, color in zip(check_labels, label_colors):
                text.set_fontsize(8 if len(labels) > 10 else 9)
                text.set_color(color)
                text.set_fontweight("bold")
            rax.set_title("Vrstvy / legenda", fontsize=10, fontweight='bold')
            fig._portfolio_checkbuttons = check
            fig._portfolio_checkbuttons_ax = rax
            fig._portfolio_color_keys = color_key_artists
        else:
            plt.subplots_adjust(right=0.96, left=0.08, top=0.92, bottom=0.16)

        mode_ax = fig.add_axes([0.486, 0.052, 0.064, 0.032])
        mode_ax.set_axis_off()
        mode_ax.set_xlim(0.0, 1.0)
        mode_ax.set_ylim(0.0, 1.0)
        toggle_bg = FancyBboxPatch(
            (0.02, 0.10),
            0.96,
            0.80,
            boxstyle="round,pad=0.02,rounding_size=0.30",
            facecolor="#e9edf3",
            edgecolor="#aab4c0",
            linewidth=1.0,
            transform=mode_ax.transAxes,
        )
        toggle_knob = FancyBboxPatch(
            (0.05, 0.16),
            0.44,
            0.68,
            boxstyle="round,pad=0.02,rounding_size=0.24",
            facecolor="#ffffff",
            edgecolor="#8aa0b8",
            linewidth=1.0,
            transform=mode_ax.transAxes,
        )
        mode_ax.add_patch(toggle_bg)
        mode_ax.add_patch(toggle_knob)
        czk_text = mode_ax.text(
            0.27,
            0.50,
            "CZK",
            ha="center",
            va="center",
            fontsize=7,
            fontweight="bold",
            color="#1f2933",
            transform=mode_ax.transAxes,
        )
        pct_text = mode_ax.text(
            0.74,
            0.50,
            "%",
            ha="center",
            va="center",
            fontsize=8,
            fontweight="bold",
            color="#667085",
            transform=mode_ax.transAxes,
        )

        def refresh_mode_button_label():
            if active_mode["value"] == "pct":
                toggle_knob.set_x(0.51)
                czk_text.set_color("#667085")
                pct_text.set_color("#1f2933")
            else:
                toggle_knob.set_x(0.05)
                czk_text.set_color("#1f2933")
                pct_text.set_color("#667085")

        def toggle_mode(event):
            if event.inaxes is not mode_ax:
                return
            active_mode["value"] = "pct" if active_mode["value"] == "czk" else "czk"
            refresh_mode_button_label()
            apply_mode()

        mode_connection = fig.canvas.mpl_connect("button_press_event", toggle_mode)
        fig._portfolio_mode_toggle_artists = (toggle_bg, toggle_knob, czk_text, pct_text)
        fig._portfolio_mode_toggle_connection = mode_connection
        fig._portfolio_mode_toggle_ax = mode_ax
        refresh_mode_button_label()
        apply_mode()

        plt.show()


def main() -> None:
    root = tk.Tk()
    PortfolioManagerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
